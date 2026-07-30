"""One-shot: turn Kyle's `$Project Info Sheet- Job Name.xlsx` into a template
this tool can fill.

Run it when the master changes, then commit the result:

    python backend/prepare_info_sheet_template.py

WHY THIS EXISTS — the dropdowns do not survive a round-trip.

Six of the seven dropdowns on the Info Sheet are *x14 extension* validations
(`x14:dataValidation`, Microsoft's 2009 namespace) rather than plain ones. Excel
uses that form whenever a list points at another worksheet, which all six do
(they read the `Packet ` tab). openpyxl cannot represent them: loading the master
prints "Data Validation extension is not supported and will be removed", and
saving drops them. Fill the master directly and the estimator gets a hand-off
sheet where Market, Primary Floor, State, Payment Terms, Lead Source and every
Y/N cell are free-text.

Re-adding them as literal inline lists does not work either: joined, the market
segments run past Excel's 255-character cap for an inline list.

So this script rebuilds the same six dropdowns in the *plain* `dataValidation`
form, which openpyxl does round-trip, using the one construction that is legal
there and understood by every Excel version — a defined name. The list values
are copied verbatim off `Packet ` onto a hidden `Lists` sheet, each range gets a
workbook-level name, and each validation points at that name. `Packet ` itself
then has no remaining references and is removed (it only ever existed to feed
these lists).

Nothing else about the workbook is touched: the SOV, Foundation Import, Invoice
and Deposit tabs keep their cross-sheet formulas, and the Info Sheet keeps its
layout, number formats, merges and logo. Kyle's master is opened read-only.
"""
from __future__ import annotations

import sys
import warnings
from pathlib import Path

import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, quote_sheetname

MASTER = Path(
    r"C:\Users\Admin\Treadwell Dropbox\2023 Treadwell Team Folder\Projects"
    r"\$xxx New Project\$Project Info Sheet- Job Name.xlsx"
)
OUT = Path(__file__).resolve().parent / "templates" / "project_info_sheet.xlsx"

LISTS_SHEET = "Lists"
SOURCE_SHEET = "Packet "  # trailing space is real

# Each dropdown: the column on `Packet ` holding its values, the name we give
# the range, and the Info Sheet cells that use it. Ranges start at row 3 (rows
# 1-2 are headers) and stop at the last non-empty row, so the picker has no
# trailing blanks — the master's B16 validation ran to H27 with six empties.
DROPDOWNS = [
    ("H", "MarketList",     "B16"),
    ("K", "FloorList",      "B17"),
    ("O", "StateList",      "B19"),
    ("M", "TermsList",      "B60"),
    ("E", "LeadSourceList", "B62"),
    ("J", "YNList",         "B59 B61 B63:B64 B66:B68"),
]


def read_list(ws, col: str) -> list[str]:
    """Values in `col` from row 3 down to the last non-empty one, verbatim."""
    out, blanks = [], []
    for row in range(3, ws.max_row + 1):
        v = ws[f"{col}{row}"].value
        if v is None or str(v).strip() == "":
            blanks.append(v)          # hold: might be a gap, not the end
            continue
        out.extend([""] * len(blanks))
        blanks.clear()
        out.append(str(v))
    return out


def main(master: Path = MASTER, out: Path = OUT) -> None:
    if not master.is_file():
        sys.exit(f"master not found: {master}")

    with warnings.catch_warnings():
        # The x14 removal warning is the whole reason for this script.
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(master)

    if SOURCE_SHEET not in wb.sheetnames:
        sys.exit(f"{master.name} has no {SOURCE_SHEET!r} tab — nothing to lift")
    packet = wb[SOURCE_SHEET]
    info = wb["Info Sheet"]

    lists = wb.create_sheet(LISTS_SHEET)
    lists["A1"] = "Dropdown sources — copied from the master's 'Packet' tab."
    lists["A2"] = "Do not edit here; edit the master and re-run prepare_info_sheet_template.py."

    for i, (col, name, targets) in enumerate(DROPDOWNS):
        values = read_list(packet, col)
        if not values:
            sys.exit(f"{name}: column {col} of {SOURCE_SHEET!r} is empty")
        letter = get_column_letter(i + 1)
        for j, v in enumerate(values):
            lists[f"{letter}{j + 4}"] = v
        ref = (f"{quote_sheetname(LISTS_SHEET)}!${letter}$4:"
               f"${letter}${len(values) + 3}")
        wb.defined_names.add(DefinedName(name, attr_text=ref))

        dv = DataValidation(type="list", formula1=name, allowBlank=True,
                            showInputMessage=True, showErrorMessage=True)
        info.add_data_validation(dv)
        for target in targets.split():
            dv.add(target)

    lists.sheet_state = "hidden"
    del wb[SOURCE_SHEET]
    info["B13"] = None      # the master's "xxNAMExx" placeholder

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    verify(out)
    print(f"wrote {out}  ({out.stat().st_size:,} bytes)")


def verify(path: Path) -> None:
    """Fail loudly rather than ship a template whose dropdowns went missing."""
    wb = openpyxl.load_workbook(path)
    info = wb["Info Sheet"]

    assert SOURCE_SHEET not in wb.sheetnames, "Packet tab survived"
    assert wb[LISTS_SHEET].sheet_state == "hidden", "Lists tab is visible"
    for tab in ("Info Sheet", "SOV", "Foundation Import", "Invoice"):
        assert tab in wb.sheetnames, f"lost the {tab} tab"
    assert wb["Foundation Import"]["A1"].value == "='Info Sheet'!B14", \
        "Foundation Import stopped pointing at the Info Sheet"

    by_formula = {dv.formula1: dv for dv in info.data_validations.dataValidation}
    for _col, name, targets in DROPDOWNS:
        assert name in wb.defined_names, f"{name} was not defined"
        dv = by_formula.get(name)
        assert dv is not None, f"{name} has no validation on the Info Sheet"
        for cell in targets.split():
            assert cell in str(dv.sqref), f"{name} does not cover {cell}"

    market = [c[0].value for c in wb[LISTS_SHEET]["A4":f"A{wb[LISTS_SHEET].max_row}"]
              if c[0].value is not None]
    assert market[0] == "-Select-" and "Religious" in market, \
        f"market list looks wrong: {market[:3]}"
    print(f"verified: {len(DROPDOWNS)} dropdowns, {len(market)} market segments, "
          f"tabs = {wb.sheetnames}")


if __name__ == "__main__":
    main(Path(sys.argv[1]) if len(sys.argv) > 1 else MASTER,
         Path(sys.argv[2]) if len(sys.argv) > 2 else OUT)
