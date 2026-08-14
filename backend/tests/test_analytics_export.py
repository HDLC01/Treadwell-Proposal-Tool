"""The Analytics page as a workbook — and the Trailing-12 sheet as Kyle's own file.

Hanz, 2026-08-14: "we need to add a prenthesis…" — no, this one: "Then we need option to download
the analytics page on a Excel sheet based on the filters that are on. ALso they need a trailing
12th month for this. Basically what they do with this is the fill it out manually. The Excel file
is in here name 'Trailing 12TH MONTH' for your basis."

So the Trailing-12 sheet is not "a table of the same numbers" — it is a REPLACEMENT for a tab in a
workbook Kyle already keeps, and the test for it is a cell-for-cell comparison against that file
(committed at the repo root). If a label loses its literal double quotes or a formula becomes a
value, the download stops dropping into his archive and he goes back to typing it by hand.

THE BROWSER SENDS THE NUMBERS. analytics-core.js is the only thing that computes a figure on that
page, and the server cannot see the viewer's filters, so re-deriving totals here would be two
engines on one question. That makes this endpoint's real job VALIDATION — hence the 422 block.
"""
import io
import pathlib

import openpyxl
import pytest
from fastapi.testclient import TestClient

import main

client = TestClient(main.app)

# A SANITIZED copy of Kyle's "Trailing 12TH MONTH.xlsx": same tab name, same labels, same formulas,
# same number formats, same column widths — every figure replaced with an obvious fabrication.
#
# His real workbook stays out of git. This repository is public, and that file is Treadwell's win
# rate and revenue by trade: exactly the numbers a competitor would want. The layout is what these
# tests need to pin, and the layout is not sensitive.
LAYOUT = pathlib.Path(__file__).resolve().parent / "fixtures" / "trailing12_layout.xlsx"
# The real thing, when a developer happens to have it beside the checkout: the sanitized fixture is
# generated FROM it, so a drift check is worth running locally even though CI cannot.
KYLE = pathlib.Path(__file__).resolve().parents[2] / "Trailing 12TH MONTH.xlsx"

# Distinct numbers per column, so a transposed or copy-pasted column cannot pass.
_T12 = {
    "as_of": "2026-08-15", "w15_from": "2025-05-15", "w90_from": "2026-05-17",
    "columns": [
        {"label": "All Bids", "won_amount": 7927983.0, "submitted_amount": 56377337.0,
         "sub90_amount": 11855702.0, "n_awarded": 174.0, "n_submitted": 734.0, "n_sub90": 159.0},
        {"label": "Gyp", "won_amount": 1.0, "submitted_amount": 2.0, "sub90_amount": 3.0,
         "n_awarded": 4.0, "n_submitted": 5.0, "n_sub90": 6.0},
        {"label": "Epoxy", "won_amount": 10.0, "submitted_amount": 20.0, "sub90_amount": 30.0,
         "n_awarded": 40.0, "n_submitted": 50.0, "n_sub90": 60.0},
        {"label": "Polish", "won_amount": 100.0, "submitted_amount": 200.0, "sub90_amount": 300.0,
         "n_awarded": 400.0, "n_submitted": 500.0, "n_sub90": 600.0},
    ],
}


def _payload(**over):
    body = {
        "generated_at": "2026-08-15T00:00:00Z",
        "filters": "Dates: Year to date (2026-01-01 to 2026-08-15) · Trades: Epoxy · 812 of 3,400 projects",
        "pull_window": {"from": None, "to": None},
        "truncated": False,
        "tabs": [
            {"name": "Overview", "tables": [
                {"title": "Overview", "columns": [{"label": "Metric"}, {"label": "Value"}],
                 "rows": [["Won amount", {"v": 50000.0, "t": "money"}],
                          ["Win % by amount", {"v": 0.8333, "t": "pct"}],
                          ["# Awarded projects", {"v": 3, "t": "int"}],
                          ["Nothing to show", {"v": None, "t": "money"}]]}]},
            {"name": "Trades", "tables": [
                {"title": "Trades", "columns": [{"label": "Trade"}, {"label": "Submitted amount"}],
                 "rows": [["Epoxy", {"v": 60000.0, "t": "money"}]]}]},
        ],
        "trailing12": _T12,
    }
    body.update(over)
    return body


def _download(body):
    """POST the payload, follow the returned link, hand back the parsed workbook.

    Loaded WITHOUT data_only, so formulas read back as their "=..." strings — with it, openpyxl
    returns the cached values, which a freshly built file does not have, and every formula
    assertion would read None and pass for the wrong reason."""
    r = client.post("/api/analytics/export", json=body)
    assert r.status_code == 200, r.text
    url = r.json()["xlsx_download_url"]
    got = client.get(url)
    assert got.status_code == 200, got.text
    return openpyxl.load_workbook(io.BytesIO(got.content)), got


# ── Kyle's sheet, against Kyle's file ────────────────────────────────────────
def test_the_trailing_sheet_matches_kyles_layout_cell_for_cell():
    """Every label, every heading, every formula, against the layout he actually keeps."""
    wb, _ = _download(_payload())
    mine = wb["Trailing 12"]
    his = openpyxl.load_workbook(LAYOUT)["04.21.26"]

    coords = ["A1", "A3", "B3", "E3", "H3", "K3",
              "A4", "A5", "A8", "A12", "A13", "A16",
              "B4", "B5", "B6", "B8", "B10", "B12", "B13", "B14", "B16", "B18", "B20", "B21",
              "E4", "E5", "E10", "H4", "H10", "K4", "K10",
              "C6", "C10", "C14", "C18", "C20", "C21",
              "F6", "F10", "F14", "F18", "F20", "F21",
              "I6", "I10", "I14", "I18", "I20", "I21",
              "L6", "L10", "L14", "L18", "L20", "L21"]
    wrong = [(c, his[c].value, mine[c].value) for c in coords
             if str(his[c].value).strip() != str(mine[c].value).strip()]
    assert not wrong, "cells that no longer match Kyle's workbook: %r" % (wrong,)


def test_the_labels_keep_their_literal_quotes():
    """He typed them with quote characters, and the sheet is meant to look like the ones beside
    it. Easy to "tidy" away, invisible in a screenshot, wrong in his archive."""
    wb, _ = _download(_payload())
    ws = wb["Trailing 12"]
    assert ws["B4"].value == '"Won Amount"'
    assert ws["B5"].value == '"Total Submitted Amount"'
    assert ws["B12"].value == '"# Awarded Projects"'
    assert ws["B13"].value == '"# Submitted Projects"'
    assert ws["B8"].value == 'Last 90 days "Total Submitted Amount"'
    assert ws["B16"].value == 'Last 90 days "# Submitted Projects"'


def test_every_column_carries_the_raw_sums_it_was_sent():
    """Distinct values per column, so a transposition or a copy-pasted column letter shows up."""
    wb, _ = _download(_payload())
    ws = wb["Trailing 12"]
    for col, letter in zip(_T12["columns"], ("C", "F", "I", "L")):
        assert ws[letter + "4"].value == col["won_amount"], letter
        assert ws[letter + "5"].value == col["submitted_amount"], letter
        assert ws[letter + "8"].value == col["sub90_amount"], letter
        assert ws[letter + "12"].value == col["n_awarded"], letter
        assert ws[letter + "13"].value == col["n_submitted"], letter
        assert ws[letter + "16"].value == col["n_sub90"], letter


def test_the_derived_cells_are_live_formulas_not_our_arithmetic():
    """THE POINT OF SENDING RAW SUMS. If these were values, the workbook would hold our answer AND
    his, and they would part company the moment he edited a number. Also why the payload carries no
    ratios at all — there is exactly one place the percentages come from, and it is Excel."""
    wb, _ = _download(_payload())
    ws = wb["Trailing 12"]
    for L in ("C", "F", "I", "L"):
        assert ws[L + "6"].value == "={0}4/{0}5".format(L)
        assert ws[L + "10"].value == "={0}4/({0}5-{0}8)".format(L)      # the win% he reads
        assert ws[L + "14"].value == "={0}12/{0}13".format(L)
        assert ws[L + "18"].value == "={0}12/({0}13-{0}16)".format(L)   # the other one
        assert ws[L + "20"].value == "={0}5/{0}13".format(L)
        assert ws[L + "21"].value == "={0}4/{0}12".format(L)


def test_the_number_formats_come_from_his_workbook():
    """Excel ACCOUNTING, not a plain currency format: "$" flush left, digits aligned right, zero as
    "-". A pasted tab that formats differently reads as a different document.

    Compared against HIS FILE, not against our own constant. Asserting
    `number_format == FMT_ACC0` was the first version of this test and it was worthless — change
    the constant and both sides of the comparison move together, which a mutation proved by
    swapping in a plain "$#,##0" and passing."""
    wb, _ = _download(_payload())
    mine = wb["Trailing 12"]
    his = openpyxl.load_workbook(LAYOUT)["04.21.26"]
    # Column C only. In his workbook the trade columns are EMPTY on most runs — he fills All Bids
    # first — so their cells never got a format and read back as "General". Ours formats all four
    # the same way, which is the improvement, not a mismatch. The next assertion is what pins that.
    for coord in ("C4", "C5", "C8", "C12", "C13", "C16",
                  "C6", "C10", "C14", "C18", "C20", "C21"):
        assert mine[coord].number_format == his[coord].number_format, (
            "%s: %r vs his %r" % (coord, mine[coord].number_format, his[coord].number_format))


def test_all_four_columns_are_formatted_alike():
    """Whatever column C does, F, I and L do. He fills All Bids first and often leaves the trade
    columns for later, so in his own file they are unformatted — a workbook where three of four
    columns lose their currency symbol the moment they get numbers would be worse than his."""
    wb, _ = _download(_payload())
    ws = wb["Trailing 12"]
    for row in (4, 5, 8, 6, 10, 14, 18, 20, 21):
        want = ws["C" + str(row)].number_format
        for letter in ("F", "I", "L"):
            got = ws[letter + str(row)].number_format
            assert got == want, "row %d: %s is %r but C is %r" % (row, letter, got, want)


def test_the_accounting_formats_are_accounting_formats():
    """The literal strings, so nothing here is measured against a constant that can move with it.
    An accounting format is recognisable: the currency symbol is padded away from the digits with
    `* `, and a zero shows as a dash."""
    wb, _ = _download(_payload())
    ws = wb["Trailing 12"]
    assert ws["C4"].number_format.startswith('_("$"* #,##0_)')
    assert '"-"??' in ws["C4"].number_format, "a zero no longer shows as a dash"
    assert ws["C8"].number_format.startswith('_("$"* #,##0.00_)'), "the cents variant was lost"
    for row in (6, 10, 14, 18):
        assert ws["C" + str(row)].number_format == "0%"


def test_the_run_date_is_a_real_date():
    """His own newest tabs put the run date in B2 as a date, which is what makes a stack of these
    sortable. A string would look identical and sort as text."""
    import datetime as dt
    wb, _ = _download(_payload())
    b2 = wb["Trailing 12"]["B2"]
    assert isinstance(b2.value, (dt.date, dt.datetime))
    assert (b2.value.date() if isinstance(b2.value, dt.datetime) else b2.value) == \
        dt.date(2026, 8, 15)
    assert b2.number_format == "mm-dd-yy"


def test_a_garbled_as_of_does_not_break_the_download():
    """It arrives as a string from the browser. A bad one should cost the date cell, not the file."""
    body = _payload()
    body["trailing12"] = dict(_T12, as_of="not-a-date")
    wb, _ = _download(body)
    assert wb["Trailing 12"]["B2"].value == "not-a-date"


# ── the dashboard tabs ──────────────────────────────────────────────────────
def test_one_sheet_per_tab_plus_the_trailing_sheet():
    wb, _ = _download(_payload())
    assert wb.sheetnames == ["Overview", "Trades", "Trailing 12"]


def test_a_sheet_says_which_filters_produced_it():
    """A saved file with no provenance is a number without a question. Hanz asked for the export to
    reflect "the filters that are on", so the file has to say what they were."""
    wb, _ = _download(_payload())
    header = wb["Overview"]["A1"].value
    assert "Year to date" in header and "Trades: Epoxy" in header
    assert "812 of 3,400 projects" in header


def test_typed_cells_get_their_formats_and_a_blank_stays_blank():
    """None must not become 0: the page shows "—" for a missing ratio, and a zero in a workbook
    reads as a measured result."""
    wb, _ = _download(_payload())
    ws = wb["Overview"]
    vals = {ws.cell(row=r, column=1).value: (ws.cell(row=r, column=2).value,
                                             ws.cell(row=r, column=2).number_format)
            for r in range(1, ws.max_row + 1)}
    assert vals["Won amount"][0] == 50000.0
    assert vals["Won amount"][1] == main.analytics_export.FMT_MONEY
    assert vals["Win % by amount"] == (0.8333, "0%")
    assert vals["# Awarded projects"][1] == main.analytics_export.FMT_INT
    assert vals["Nothing to show"][0] is None


def test_the_cap_warning_travels_into_the_file():
    """`truncated` means the dashboard did not load every bid, and the ones dropped are not chosen
    by date — a trailing window can lose rows silently. Someone reading the file later has no way
    to know that unless it says so."""
    wb, _ = _download(_payload(truncated=True))
    assert "CAPPED" in wb["Overview"]["A1"].value


def test_the_filename_carries_the_run_date():
    _, resp = _download(_payload())
    assert "attachment" in resp.headers["content-disposition"]
    assert "2026-08-15" in resp.headers["content-disposition"]


def test_an_empty_payload_still_produces_an_openable_file():
    """Nothing to export is not an error — and a workbook with zero sheets will not open at all."""
    wb, _ = _download({"tabs": [], "trailing12": None})
    assert wb.sheetnames == ["Analytics"]


# ── validation: staff input, so 422 and never 500 ───────────────────────────
@pytest.mark.parametrize("mutate,why", [
    ({"nope": 1}, "an unknown top-level key"),
    ({"tabs": [{"name": "X", "tables": [{"columns": [], "rows": [], "bogus": 2}]}]},
     "an unknown key inside a table"),
    ({"tabs": [{"name": "X", "tables": [{"columns": [], "rows": [[{"v": "abc", "t": "money"}]]}]}]},
     "a non-numeric cell value"),
    ({"tabs": [{"name": "X", "tables": [{"columns": [], "rows": [[{"v": 1, "t": "weird"}]]}]}]},
     "an unknown cell type"),
    ({"tabs": [{"name": "X", "tables": [{"columns": [], "rows": [[]] * 2001}]}]},
     "more rows than the cap"),
    ({"tabs": [{"name": "X", "tables": [{} for _ in range(13)]}]},
     "more tables than the cap"),
    ({"trailing12": dict(_T12, columns=_T12["columns"] * 2)}, "more than four trade columns"),
])
def test_a_malformed_payload_is_refused_not_a_crash(mutate, why):
    """openpyxl will happily write a cell Excel then refuses to open, so every bound lives at the
    door. These are all 422 — a 500 would say "the server is broken" about a bad request."""
    r = client.post("/api/analytics/export", json=_payload(**mutate))
    assert r.status_code == 422, "%s produced %s" % (why, r.status_code)


@pytest.mark.parametrize("literal", ["NaN", "Infinity", "-Infinity"])
def test_a_non_finite_number_is_refused(literal):
    """These cannot come from our own page — JSON.stringify turns them into null, and even the test
    client refuses to encode them. They arrive only in a RAW body using the non-standard literals
    Python's json parser accepts, which any other holder of a staff token can send. Left alone,
    openpyxl writes a cell that makes Excel refuse to open the whole file."""
    raw = ('{"tabs": [{"name": "X", "tables": [{"columns": [], '
           '"rows": [[{"v": %s, "t": "int"}]]}]}]}' % literal)
    r = client.post("/api/analytics/export", content=raw,
                    headers={"Content-Type": "application/json"})
    assert r.status_code == 422, r.text


@pytest.mark.parametrize("literal", ["NaN", "Infinity"])
def test_a_non_finite_trailing_sum_is_refused(literal):
    raw = ('{"trailing12": {"as_of": "2026-08-15", "w15_from": "", "w90_from": "", '
           '"columns": [{"label": "X", "won_amount": %s, "submitted_amount": 1, '
           '"sub90_amount": 1, "n_awarded": 1, "n_submitted": 1, "n_sub90": 1}]}}' % literal)
    r = client.post("/api/analytics/export", content=raw,
                    headers={"Content-Type": "application/json"})
    assert r.status_code == 422, r.text


def test_long_strings_are_truncated_rather_than_refused():
    """A filter sentence naming forty companies is legitimate; it just cannot be unbounded."""
    wb, _ = _download(_payload(filters="x" * 5000))
    assert len(wb["Overview"]["A1"].value) < 1000


@pytest.mark.skipif(not KYLE.exists(), reason="Kyle's own workbook is not beside this checkout")
def test_the_fixture_has_not_drifted_from_kyles_real_workbook():
    """The fixture is a sanitized copy, so it can go stale if he changes his sheet. Skipped in CI
    (his file is deliberately not in git) and loud on any machine that has it.

    Compares LAYOUT ONLY — labels, formulas, headings. Never his figures, which is the whole
    reason the fixture exists."""
    fixture = openpyxl.load_workbook(LAYOUT)["04.21.26"]
    real = openpyxl.load_workbook(KYLE)["04.21.26"]
    drift = []
    for row in real.iter_rows(min_row=1, max_row=25, max_col=12):
        for cell in row:
            import datetime as _dt
            if cell.value is None or isinstance(cell.value, (int, float, _dt.date, _dt.datetime)):
                continue                          # figures, dates and blanks are not the contract
            if str(fixture[cell.coordinate].value) != str(cell.value):
                drift.append(cell.coordinate)
    assert not drift, (
        "Kyle's workbook layout has changed at %r — regenerate "
        "backend/tests/fixtures/trailing12_layout.xlsx" % (drift,))
