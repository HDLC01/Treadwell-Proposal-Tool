"""Excel sheet-protection on the generated .xlsx (rate/markup/tax cell locking).

fill_estimate must lock exactly the LOCK_MAP cells (GP %, Hard Bid Discount,
Superintendent, Soft Costs, Contingency, Sales Tax %, Kansas Remodel Tax %, Bond)
per sheet layout and leave every other cell editable, with sheet protection ON
(no password) for the lock-mapped sheets and OFF for everything else. Copies
inherit their source's layout; the alternate tab and display-label renames must
not shake the protection loose (it's applied to worksheet objects, post-rename).
"""
import io

from openpyxl import load_workbook

import estimate_writer as ew


def _wb(data):
    return load_workbook(io.BytesIO(data))


def _locked_addrs(ws, addrs):
    return {a: bool(ws[a].protection.locked) for a in addrs}


# ── every lock-mapped template sheet is protected with exactly its set ──
def test_lock_map_sheets_protected_with_correct_cells():
    wb = _wb(ew.fill_estimate({"project_name": "P"}))
    for name, addrs in ew.LOCK_MAP.items():
        sheets = ([s for s in wb.sheetnames if s.lower().startswith("gyp")]
                  if name == "Gyp" else [name])
        assert sheets, f"no sheet found for layout {name!r}"
        for sheet in sheets:
            ws = wb[sheet]
            assert ws.protection.sheet is True, f"{sheet}: protection not enabled"
            assert not ws.protection.password, f"{sheet}: unexpected password"
            got = _locked_addrs(ws, addrs)
            assert all(got.values()), f"{sheet}: unlocked rate cells {got}"


def test_all_other_cells_stay_editable():
    wb = _wb(ew.fill_estimate({"project_name": "P", "sqft": 12000}))
    for sheet, lock_key in (("Epoxy", "Epoxy"), ("Polish", "Polish")):
        ws = wb[sheet]
        locked = set(ew.LOCK_MAP[lock_key])
        # Every cell that actually exists in the file must be explicitly
        # unlocked except the lock-map addrs. (Virgin cells materialized by
        # iter_rows on the RELOADED wb report openpyxl's locked-by-default
        # Protection, but Excel resolves them through the Normal named style —
        # asserted unlocked below — so they're editable in the real app.)
        stray = [c.coordinate for row in ws.iter_rows() for c in row
                 if c.protection.locked and c.coordinate not in locked
                 and (c.value is not None or c.has_style)]
        assert not stray, f"{sheet}: unexpectedly locked {stray[:10]}"


def test_virgin_cells_inherit_unlocked_normal_style():
    # Cells that don't exist in the file (blank rows/columns) inherit the
    # workbook "Normal" style under Excel's protection rules; fill_estimate
    # flips its locked flag off so Kyle can still type in blank cells on a
    # protected sheet. Assert the flag survives the save/load roundtrip.
    wb = _wb(ew.fill_estimate({"project_name": "P"}))
    normal = next(s for s in wb._named_styles if s.name == "Normal")
    assert normal.protection.locked is False


def test_non_lock_sheets_not_protected():
    wb = _wb(ew.fill_estimate({"project_name": "P"}))
    for sheet in ("Takeoff", "Stnd Alts", "validation", "Unit Layouts"):
        assert wb[sheet].protection.sheet is not True, f"{sheet}: should be unprotected"


# ── copies inherit the SOURCE layout ─────────────────────────────────────
def test_copy_inherits_source_lock_layout():
    wb = _wb(ew.fill_estimate({}, tab_copies=[{"id": "Copy1", "source": "Polish"}]))
    ws = wb["Copy1"]
    assert ws.protection.sheet is True
    assert all(_locked_addrs(ws, ew.LOCK_MAP["Polish"]).values())


def test_copy_of_copy_resolves_through_chain():
    wb = _wb(ew.fill_estimate({}, tab_copies=[
        {"id": "Copy1", "source": "Leveling"},
        {"id": "Copy2", "source": "Copy1"},
    ]))
    ws = wb["Copy2"]
    assert ws.protection.sheet is True
    assert all(_locked_addrs(ws, ew.LOCK_MAP["Leveling"]).values())


def test_gyp_copy_uses_gyp_set():
    wb = _wb(ew.fill_estimate({}, tab_copies=[{"id": "GypCopy", "source": 'Gyp (USG 1-8")'}]))
    ws = wb["GypCopy"]
    assert ws.protection.sheet is True
    got = _locked_addrs(ws, ew.LOCK_MAP["Gyp"])
    assert all(got.values()), got
    # Contingency on Gyp is col E, not D — D76 must stay editable.
    assert not ws["D76"].protection.locked


# ── protection survives the renames (alternate tab + display labels) ────
def test_alternate_rename_keeps_protection():
    wb = _wb(ew.fill_estimate({}, alternate={"sf": 9000, "material_sub": 4000,
                                             "label": "MMA alt"}))
    assert "Alternate System" in wb.sheetnames
    ws = wb["Alternate System"]
    assert ws.protection.sheet is True
    assert all(_locked_addrs(ws, ew.LOCK_MAP["Epoxy blank"]).values())


def test_tab_label_rename_keeps_protection():
    wb = _wb(ew.fill_estimate({"project_name": "P"},
                              tab_labels={"Epoxy": "Grooming Room"}))
    ws = wb["Grooming Room"]
    assert ws.protection.sheet is True
    assert all(_locked_addrs(ws, ew.LOCK_MAP["Epoxy"]).values())


def test_renamed_copy_keeps_protection():
    wb = _wb(ew.fill_estimate(
        {"project_name": "P"},
        tab_copies=[{"id": "Copy1", "source": "Epoxy"}],
        tab_labels={"Copy1": "Exam Room"},
    ))
    ws = wb["Exam Room"]
    assert ws.protection.sheet is True
    assert all(_locked_addrs(ws, ew.LOCK_MAP["Epoxy"]).values())


# ── the user's own edits still land in locked cells (tool writes bypass) ─
def test_cell_values_write_into_locked_cell_still_lands():
    # The 🔒-unlock flow on the frontend sends the edit via cell_values; sheet
    # protection guards EXCEL editing only, never the tool's own writes.
    wb = _wb(ew.fill_estimate({}, cell_values={"Epoxy!B75": 0.05}))
    ws = wb["Epoxy"]
    assert ws["B75"].value == 0.05
    assert ws["B75"].protection.locked
    assert ws.protection.sheet is True


def test_frontend_lock_map_parity():
    # The backend LOCK_MAP must mirror frontend/js/estimate-review.js
    # LOCKED_CELLS + GYP_LOCKED — same layouts, same addresses.
    import json
    import pathlib
    import re

    js = (pathlib.Path(__file__).resolve().parents[2]
          / "frontend" / "js" / "estimate-review.js").read_text(encoding="utf-8")
    m = re.search(r"const LOCKED_CELLS = (\{.*?\});", js, re.S)
    assert m, "LOCKED_CELLS not found in estimate-review.js"
    fe = json.loads(re.sub(r",(\s*[}\]])", r"\1", m.group(1)))   # tolerate trailing commas
    g = re.search(r"const GYP_LOCKED = (\[.*?\]);", js, re.S)
    assert g, "GYP_LOCKED not found in estimate-review.js"
    fe["Gyp"] = json.loads(g.group(1))
    assert fe == ew.LOCK_MAP


# ── user lock/unlock overrides (the "Lock cell" toolbar button) ──────────
def test_user_lock_adds_to_defaults():
    wb = _wb(ew.fill_estimate({}, lock_overrides={"Epoxy": {"lock": ["E20"]}}))
    ws = wb["Epoxy"]
    assert ws["E20"].protection.locked                      # user-locked
    assert all(_locked_addrs(ws, ew.LOCK_MAP["Epoxy"]).values())   # presets intact
    assert ws.protection.sheet is True


def test_user_lock_on_virgin_cell_materializes_and_locks():
    # A cell that doesn't exist in the template gets created + locked.
    wb = _wb(ew.fill_estimate({}, lock_overrides={"Epoxy": {"lock": ["J50"]}}))
    assert wb["Epoxy"]["J50"].protection.locked


def test_user_unlock_removes_a_default():
    wb = _wb(ew.fill_estimate({}, lock_overrides={"Epoxy": {"unlock": ["B73"]}}))
    ws = wb["Epoxy"]
    assert ws["B73"].protection.locked is not True          # unlocked by the user
    # the other presets stay locked, sheet still protected
    assert all(ws[a].protection.locked for a in ew.LOCK_MAP["Epoxy"] if a != "B73")
    assert ws.protection.sheet is True


def test_unlock_all_defaults_disables_protection():
    wb = _wb(ew.fill_estimate({}, lock_overrides={"Epoxy": {"unlock": list(ew.LOCK_MAP["Epoxy"])}}))
    assert wb["Epoxy"].protection.sheet is not True          # nothing left to lock -> off


def test_user_lock_protects_an_unmapped_sheet():
    wb = _wb(ew.fill_estimate({}, lock_overrides={"Takeoff": {"lock": ["B5"]}}))
    tk = wb["Takeoff"]
    assert tk.protection.sheet is True
    assert tk["B5"].protection.locked
    # sheets with neither presets nor user locks stay unprotected
    assert wb["validation"].protection.sheet is not True


def test_empty_and_junk_overrides_preserve_today_behavior():
    base = _wb(ew.fill_estimate({"project_name": "P"}))
    for lo in (None, {}, {"Epoxy": "not-a-dict"}, {"Epoxy": {"lock": ["A1:B2", "Epoxy!A1", "ZZZZ99", 5, None]}},
               {"x" * 40: {"lock": ["A1"]}}):
        wb = _wb(ew.fill_estimate({"project_name": "P"}, lock_overrides=lo))
        # Epoxy presets unchanged; no crash; junk addrs ignored.
        assert all(_locked_addrs(wb["Epoxy"], ew.LOCK_MAP["Epoxy"]).values())
    # a bare range/prefixed/oversized addr never becomes a lock
    wb = _wb(ew.fill_estimate({}, lock_overrides={"Epoxy": {"lock": ["A1:B2", "Epoxy!A1", "ZZZZ99"]}}))
    # A1 must NOT be locked (none of those are valid single-cell current-coord addrs)
    assert wb["Epoxy"]["A1"].protection.locked is not True


def test_user_lock_uses_current_coords_while_presets_translate():
    # An insert at row 30 shifts the Epoxy default B73 -> B74. A user lock is
    # given in CURRENT coords ("B90") and must apply as-is (NOT translated).
    wb = _wb(ew.fill_estimate(
        {}, tab_structs=[{"sheet": "Epoxy", "kind": "insert_rows", "at": 30, "count": 1}],
        lock_overrides={"Epoxy": {"lock": ["B90"]}}))
    ws = wb["Epoxy"]
    assert ws["B90"].protection.locked                       # user addr as-is
    assert ws["B74"].protection.locked                       # default B73 shifted down 1
    assert ws["B73"].protection.locked is not True           # old default coord now free


def test_user_unlock_uses_current_coords():
    # After the same insert, unlocking the SHIFTED default coord (B74) removes it.
    wb = _wb(ew.fill_estimate(
        {}, tab_structs=[{"sheet": "Epoxy", "kind": "insert_rows", "at": 30, "count": 1}],
        lock_overrides={"Epoxy": {"unlock": ["B74"]}}))
    assert wb["Epoxy"]["B74"].protection.locked is not True


def test_copy_overrides_apply_to_the_copy_only():
    wb = _wb(ew.fill_estimate(
        {}, tab_copies=[{"id": "Copy1", "source": "Epoxy"}],
        lock_overrides={"Copy1": {"lock": ["E20"], "unlock": ["B73"]}}))
    copy, epoxy = wb["Copy1"], wb["Epoxy"]
    assert copy["E20"].protection.locked                     # user lock on the copy
    assert copy["B73"].protection.locked is not True         # user unlock on the copy
    assert epoxy["E20"].protection.locked is not True        # source untouched
    assert epoxy["B73"].protection.locked                    # source preset intact


def test_user_lock_survives_display_rename():
    wb = _wb(ew.fill_estimate(
        {}, tab_labels={"Epoxy": "Grooming Room"},
        lock_overrides={"Epoxy": {"lock": ["E20"]}}))
    ws = wb["Grooming Room"]
    assert ws["E20"].protection.locked
    assert ws.protection.sheet is True


def test_addr_in_both_lock_and_unlock_unlock_wins():
    wb = _wb(ew.fill_estimate({}, lock_overrides={"Epoxy": {"lock": ["E20"], "unlock": ["E20"]}}))
    assert wb["Epoxy"]["E20"].protection.locked is not True


def test_norm_lock_overrides_tolerates_non_list_values():
    # The lock/unlock VALUE itself may be junk (not a list) on a corrupted
    # draft or hostile request — must coerce to empty, never raise on the slice.
    for bad in ({"x": 1}, 5, True, 3.5, "A1"):
        out = ew._norm_lock_overrides({"Epoxy": {"lock": bad, "unlock": bad}})
        assert out == {}                                     # nothing valid -> sheet dropped
    # mixed: junk lock value, valid unlock list
    out = ew._norm_lock_overrides({"Epoxy": {"lock": 5, "unlock": ["B73"]}})
    assert out == {"Epoxy": {"lock": [], "unlock": ["B73"]}}
    # and it still generates (no 500) with a non-list value
    wb = _wb(ew.fill_estimate({}, lock_overrides={"Epoxy": {"lock": {"x": 1}}}))
    assert all(_locked_addrs(wb["Epoxy"], ew.LOCK_MAP["Epoxy"]).values())


def test_norm_lock_overrides_normalizes_and_caps():
    out = ew._norm_lock_overrides({
        "Epoxy": {"lock": ["e20", " b30 ", "b30", "A1:B2", "junk"], "unlock": ["c40"]},
        "x" * 40: {"lock": ["A1"]},                          # long key clamped to 31
        "Bad": "not-a-dict",
        "Empty": {"lock": [], "unlock": []},                 # dropped (nothing to say)
    })
    assert out["Epoxy"]["lock"] == ["E20", "B30"]            # upper+strip+dedup, junk gone
    assert out["Epoxy"]["unlock"] == ["C40"]
    assert ("x" * 31) in out and ("x" * 40) not in out
    assert "Bad" not in out and "Empty" not in out
    # cap: >64 sheets is truncated
    big = {f"S{i}": {"lock": ["A1"]} for i in range(100)}
    assert len(ew._norm_lock_overrides(big)) <= 64


def test_generate_endpoint_survives_malformed_lock_overrides():
    from fastapi.testclient import TestClient
    import main
    client = TestClient(main.app)
    body = {"work_type": "epoxy", "audience": "Direct", "values": {"project_name": "P"},
            "lock_overrides": {"Epoxy": {"lock": ["A1:B2", 5, None, {"x": 1}]}, "Bad": "nope"}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
