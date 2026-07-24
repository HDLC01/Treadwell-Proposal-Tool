"""Proposal Review PRICE-line DISPLAY overrides (price_overrides).

Editable base bid / option / manual price-line label+amount that override the
proposal .docx's shown TEXT only — never cell_values, the .xlsx, or the totals.
Pattern mirrors system_overrides (test_proposal_editor.py). The .docx-rendering
helper (_rendered) and value dict mirror the rest of the suite.
"""
import io

from docx import Document
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import main

client = TestClient(main.app)

_MC = "{http://schemas.openxmlformats.org/markup-compatibility/2006}"


def _rendered(docx_bytes):
    """Text of paragraphs Word actually renders (mc:Choice copy), excluding the
    legacy mc:Fallback duplicate — same helper used across the suite."""
    d = Document(io.BytesIO(docx_bytes))
    out = []
    for p in d.element.xpath("//w:p"):
        if any(True for _ in p.iterancestors(f"{_MC}Fallback")):
            continue
        t = "".join(x.text or "" for x in p.xpath(".//w:t")).strip()
        if t:
            out.append(t)
    return "\n".join(out)


_VALS = {
    "job_name": "J", "city_state": "C", "bid_date_formatted": "6/19/26",
    "base_bid_formatted": "$58,523.00", "material_tax_formatted": "$2,639.00",
    "state_name": "Kansas", "total_formatted": "$36,763.00",
    "system_name": "MACRO", "texture": "OP", "epoxy_sf": "12,000",
    "scope_notes": "demo + install", "schedule_notes": "~5 days",
    "work_description": "w", "site_visit_date": "6/19", "disposal": "d",
    "site_visit_phrase": "per site visit on 6/19/26",
    "base_tax_phrase": "(material sales tax INCLUDED)",
    "exclusions": "standard exclusions",
}


def _rooms(base_total=36763):
    """A base bid + one priced option (id 'Copy1', computed 'Quartz …' / $11,126)."""
    return [
        {"id": "Epoxy", "is_base": True, "name": "Base",
         "system_desc": "MACRO Flake Single Broadcast", "bid": {"total": base_total, "remodel": 0}},
        {"id": "Copy1", "is_base": False, "show": True, "base_total": base_total,
         "price_mode": "total", "option_desc": "Quartz Double Broadcast",
         "system_desc": "Quartz Double Broadcast", "bid": {"total": 11126, "remodel": 0}},
    ]


# ── option label + amount override (docx changes, .xlsx untouched) ─────────
def test_option_override_label_and_amount_applied_xlsx_untouched():
    body = {"work_type": "epoxy", "audience": "Direct", "values": dict(_VALS),
            "cell_values": {"Epoxy!E20": 12345}, "rooms": _rooms(),
            "price_overrides": {"options": {"Copy1": {"label": "Custom Option Line", "amount": "$9,999"}}}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    assert "Custom Option Line" in blob                 # label overridden
    assert "$9,999" in blob                              # amount overridden
    assert "Quartz Double Broadcast" not in blob         # computed label replaced everywhere
    assert "$11,126" not in blob                         # computed amount replaced
    # The .xlsx is generated independently of price_overrides (display-only).
    wb = load_workbook(io.BytesIO(client.get(r.json()["xlsx_download_url"]).content))
    assert wb["Epoxy"]["E20"].value == 12345


def test_option_override_unknown_id_is_noop():
    body = {"work_type": "epoxy", "audience": "Direct", "values": dict(_VALS), "rooms": _rooms(),
            "price_overrides": {"options": {"DoesNotExist": {"label": "X", "amount": "$1"}}}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    assert "Quartz Double Broadcast" in blob             # computed option unchanged
    assert "$11,126" in blob


# ── single_bid base amount + tax phrase (INCLUDED layout) ──────────────────
def test_single_bid_override_amount_and_tax_phrase_included_layout():
    body = {"work_type": "epoxy", "audience": "Direct", "values": dict(_VALS),
            "price_overrides": {"single_bid": {"amount": "$40,000", "tax_phrase": "(custom tax note)"}}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    assert "$40,000" in blob                             # base amount overridden
    assert "(custom tax note)" in blob                   # tax phrase overridden
    # INCLUDED stays a single all-in line: the override must NOT inject the
    # itemized (engine-computed) Material Sales Tax / separate Total math rows.
    assert "Material Sales Tax" not in blob
    assert not any(l.strip().endswith("– Total") for l in blob.split("\n"))


# ── blank override reverts to the computed value ───────────────────────────
def test_override_blank_reverts_to_computed():
    body = {"work_type": "epoxy", "audience": "Direct", "values": dict(_VALS), "rooms": _rooms(),
            "price_overrides": {"options": {"Copy1": {"label": "   ", "amount": ""}},
                                "single_bid": {"amount": "   "}}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    assert "Quartz Double Broadcast" in blob             # blank option override -> computed
    assert "$11,126" in blob
    assert "$36,763" in blob                             # blank base override -> computed base


# ── manual override: positional by ORIGINAL index, survives a filtered hole ─
def test_manual_override_positional_alignment_with_hole():
    price_lines = [
        {"label": "Line A", "amount": 100},
        {"label": "", "amount": 0},                      # filtered out -> hole at index 1
        {"label": "Line C", "amount": 300},
    ]
    body = {"work_type": "epoxy", "audience": "Direct", "values": dict(_VALS),
            "price_lines": price_lines,
            "price_overrides": {"manual": [{}, {}, {"label": "Overridden C", "amount": "$333"}]}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    assert "Line A" in blob and "$100" in blob           # untouched
    assert "Overridden C" in blob and "$333" in blob     # index-2 override landed on Line C's row
    assert "Line C" not in blob                          # (a built-index apply would miss this)


# ── malformed price_overrides must never 500 ───────────────────────────────
def test_malformed_price_overrides_never_500():
    body = {"work_type": "epoxy", "audience": "Direct", "values": dict(_VALS), "rooms": _rooms(),
            "price_lines": [{"label": "Add for X", "amount": 500}],
            "price_overrides": {
                "options": {"Copy1": "not-a-dict", "Bad": {"label": {"n": 1}, "amount": ["x"]}, "Num": 42},
                "manual": ["junk", 42, None, {"label": {"n": 1}}, {"amount": ["x"]}],
                "single_bid": {"amount": ["x"], "tax_phrase": {"n": 1}},
            }}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    assert "Quartz Double Broadcast" in blob             # nothing malformed applied -> computed stands


# ── unit: _sanitize_price_overrides caps + coerces (never raises) ──────────
def test_sanitize_price_overrides_caps_and_coerces():
    out = main._sanitize_price_overrides({
        "options": {
            "A": {"label": 42, "amount": "  $9,999  "},   # int coerces, amount strips
            "B": {"label": {"n": 1}, "amount": ["x"]},     # nested dropped -> empty -> not included
            "C": {"label": "", "amount": "   "},           # blank -> empty -> not included (revert)
            "D": {"label": "keep", "bogus": "x"},          # unknown key ignored
        },
        "manual": [{"label": "m0"}, "not-a-dict", {"amount": ["x"]}, {"label": True}],
        "single_bid": {"amount": "$40,000", "tax_phrase": "(note)", "bogus": "x"},
    })
    assert out["options"] == {"A": {"label": "42", "amount": "$9,999"}, "D": {"label": "keep"}}
    assert out["manual"] == [{"label": "m0"}, {}, {}, {}]  # index-preserving {} placeholders
    assert out["single_bid"] == {"amount": "$40,000", "tax_phrase": "(note)"}
    # options cap
    big = {"options": {str(i): {"label": f"L{i}"} for i in range(300)}}
    assert len(main._sanitize_price_overrides(big)["options"]) == main._PRICE_OVERRIDES_MAX
    # per-field length cap
    long_one = main._sanitize_price_overrides({"single_bid": {"amount": "z" * 999}})
    assert len(long_one["single_bid"]["amount"]) == main._PRICE_OVERRIDE_FIELD_MAXLEN
    # non-dict input never raises
    assert main._sanitize_price_overrides("nope") == {
        "options": {}, "manual": [], "single_bid": {}, "rows": {}, "alternate": {}}


# ── unit: tax-row + alternate sanitize (new keys) ──────────────────────────
def test_sanitize_rows_and_alternate():
    out = main._sanitize_price_overrides({
        "rows": {
            "sales_tax": {"amount": "  $2,000 ", "label": 7},     # coerce + strip
            "remodel":   {"amount": "", "label": "   "},           # blank -> dropped
            "total":     {"amount": "$9", "label": "Grand Total"},
            "bogus_row": {"amount": "$1"},                         # unknown key ignored
        },
        "alternate": {"name": "Alt Sys", "total_amount": "$5,000", "bogus": "x",
                      "total_label": {"n": 1}},                    # nested field dropped
    })
    assert out["rows"] == {"sales_tax": {"amount": "$2,000", "label": "7"},
                           "total": {"amount": "$9", "label": "Grand Total"}}
    assert out["alternate"] == {"name": "Alt Sys", "total_amount": "$5,000"}
    # legacy payload (no rows/alternate) still yields the empty new keys
    legacy = main._sanitize_price_overrides({"single_bid": {"amount": "$1"}})
    assert legacy["rows"] == {} and legacy["alternate"] == {}


# ── tax rows: amount + label overridden in the docx, .xlsx untouched ───────
def test_tax_row_overrides_amount_and_label_brokenout_epoxy():
    vals = dict(_VALS); vals["tax_inclusion"] = "BROKEN_OUT"
    body = {"work_type": "epoxy", "audience": "Direct", "values": vals,
            "cell_values": {"Epoxy!E20": 12345},
            "remodel": [{"amount_formatted": "$1,200.00"}],
            "price_overrides": {"rows": {
                "sales_tax": {"amount": "$2,000", "label": "Sales Tax Custom"},
                "remodel":   {"amount": "$1,500", "label": "Remodel Custom"},
                "total":     {"amount": "$99,999", "label": "Grand Total"}}}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    for s in ("$2,000", "Sales Tax Custom", "$1,500", "Remodel Custom", "$99,999", "Grand Total"):
        assert s in blob, f"missing override {s!r}"
    assert "$2,639.00" not in blob      # computed material tax replaced
    assert "$36,763.00" not in blob     # computed total replaced
    # .xlsx is generated independently of price_overrides (display-only).
    wb = load_workbook(io.BytesIO(client.get(r.json()["xlsx_download_url"]).content))
    assert wb["Epoxy"]["E20"].value == 12345


# ── gate: gyp is NOT touched by a stale row override (paragraph-channel rows) ─
def test_tax_row_override_gate_leaves_gyp_computed():
    vals = {"work_type": "gyp", "job_name": "J", "city_state": "Branson, MO",
            "bid_date_formatted": "7/10/26", "gyp_soft_sf": "100", "gyp_hard_sf": "50",
            "gyp_corridor_sf": "25", "base_bid_formatted": "$98,000.00",
            "material_tax_formatted": "$5,364.00", "tax_amount_formatted": "$0.00",
            "total_formatted": "$103,364.00", "estimator_name": "Kyle"}
    body = {"work_type": "gyp", "audience": "Direct", "values": vals,
            "price_overrides": {"rows": {"sales_tax": {"amount": "$777", "label": "HACKED"}}}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    assert "$5,364.00" in blob and "Material Sales Tax" in blob  # computed row intact
    assert "HACKED" not in blob and "$777" not in blob           # gate blocked the override


# ── included mode: a total override must not leak into the base line ───────
def test_total_override_no_leak_in_included_mode():
    vals = dict(_VALS)  # default INCLUDED layout (no tax_inclusion)
    body = {"work_type": "epoxy", "audience": "Direct", "values": vals,
            "price_overrides": {"rows": {"total": {"amount": "$1"}}}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    # Included mode: the single base line carries the COMPUTED total, never the
    # override (the total row itself is hidden/stripped in included mode).
    assert "$36,763.00" in blob
    assert "$1 " not in blob and "$1\n" not in blob


# ── alternate block: name + rows overridden in docx; .xlsx tab computed ────
def test_alternate_block_overrides_docx_only():
    acb = {"alternate_full_bid": {"total_base_bid": 20000.0, "remodel_tax": 1000.0},
           "alternate": {"label": "Computed Alt", "sf": 5000}}
    body = {"work_type": "epoxy", "audience": "Direct", "values": dict(_VALS),
            "alternate_computed_bid": acb, "alternate_label": "Computed Alt",
            "price_overrides": {"alternate": {
                "name": "Premium Alternate", "flooring_amount": "$18,500",
                "flooring_label": "Premium flooring as described above",
                "total_amount": "$21,000", "total_label": "Alt Grand Total"}}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    for s in ("Premium Alternate", "$18,500", "Premium flooring as described above",
              "$21,000", "Alt Grand Total"):
        assert s in blob, f"missing alternate override {s!r}"
    assert "Computed Alt" not in blob     # name replaced


# ── combo: overrides ride pre-applied combo_options strings ────────────────
def test_combo_options_overridden_strings_render():
    body = {"work_type": "combo", "audience": "Direct", "values": dict(_VALS),
            "combo_options": [
                {"amount_formatted": "$12,345", "label": "Option 1: Custom Epoxy line"},
                {"amount_formatted": "$67,890", "label": "Total"}]}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    assert "$12,345" in blob and "Option 1: Custom Epoxy line" in blob
    assert "$67,890" in blob


def test_malformed_rows_alternate_never_500():
    body = {"work_type": "epoxy", "audience": "Direct", "values": dict(_VALS),
            "price_overrides": {"rows": "nope", "alternate": ["bad"], "combo": {"x": 1}}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text


# ── option override composes with the "Options:" label + ordering ──────────
def test_option_override_composes_with_options_label():
    body = {"work_type": "epoxy", "audience": "Direct", "values": dict(_VALS), "rooms": _rooms(),
            "price_overrides": {"options": {"Copy1": {"label": "Renamed Option", "amount": "$9,999"}}}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    lines = _rendered(client.get(r.json()["docx_download_url"]).content).split("\n")
    assert lines.count("Base Bid") == 1                  # single base bid
    assert "Options:" in lines                            # label still present
    base_i = lines.index("Base Bid")
    opt_i = lines.index("Options:")
    assert base_i < opt_i                                 # ordered after the base
    after = lines[opt_i + 1:]
    assert any("Renamed Option" in l and "$9,999" in l for l in after)


# ── single_bid.desc: base-line description override (in-doc text swap) ──────
def test_single_bid_desc_override_replaces_base_description():
    body = {"work_type": "epoxy", "audience": "Direct", "values": dict(_VALS), "rooms": _rooms(),
            "price_overrides": {"single_bid": {"desc": "Custom epoxy system as described above"}}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    assert "Custom epoxy system as described above" in blob   # base description overridden
    assert "Epoxy flooring as described above" not in blob     # template's static default replaced
    assert "Base Bid" in blob                                  # base line still renders
    assert "$36,763" in blob                                   # desc-only override leaves amount intact


def test_base_description_default_preserved_without_override():
    body = {"work_type": "epoxy", "audience": "Direct", "values": dict(_VALS), "rooms": _rooms()}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    assert "Epoxy flooring as described above" in blob         # no override -> template wording kept
