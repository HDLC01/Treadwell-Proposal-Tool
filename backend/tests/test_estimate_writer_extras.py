"""Extra-material write-back: custom lines land in the sheet's spare =B*C
rows, the formula stays intact, and overflow past the 6 native rows lumps
into one 'Misc materials' line (so the .xlsx formulas never break)."""
import io

import openpyxl

import estimate_writer


def _epoxy(xlsx_bytes):
    return openpyxl.load_workbook(io.BytesIO(xlsx_bytes))["Epoxy"]


def test_extras_written_to_spare_rows_with_formula_intact():
    extras = [
        {"label": "Super Stick", "qty": 2, "unit_price": 50},
        {"label": "Floor Graphic", "qty": 1, "unit_price": 850},
    ]
    ws = _epoxy(estimate_writer.fill_estimate({}, extras=extras))
    # spare rows in order: 23, 27, 28, 32, 33, 39
    assert ws["A23"].value == "Super Stick"
    assert ws["B23"].value == 2
    assert ws["C23"].value == 50
    assert ws["A27"].value == "Floor Graphic"
    assert ws["B27"].value == 1
    assert ws["C27"].value == 850
    # D stays the template's =B*C formula (we never overwrite the amount)
    assert str(ws["D23"].value).replace(" ", "") == "=B23*C23"


def test_blank_extras_are_ignored():
    extras = [{"label": "", "qty": "", "unit_price": ""},
              {"label": "Real One", "qty": 1, "unit_price": 10}]
    ws = _epoxy(estimate_writer.fill_estimate({}, extras=extras))
    # the real one takes the FIRST spare row (blanks skipped before serialization)
    assert ws["A23"].value == "Real One"


def test_overflow_lumps_into_last_spare_row():
    extras = [{"label": f"M{i}", "qty": 1, "unit_price": 100} for i in range(8)]  # 8 > 6
    ws = _epoxy(estimate_writer.fill_estimate({}, extras=extras))
    # first 5 itemized (rows 23,27,28,32,33), remainder lumped into row 39
    assert ws["A23"].value == "M0"
    assert ws["A33"].value == "M4"
    assert str(ws["A39"].value).startswith("Misc materials")
    # 3 leftover lines x $100, written as qty 1 x unit price 300
    assert ws["B39"].value == 1
    assert ws["C39"].value == 300


def test_no_extras_leaves_spare_rows_blank():
    ws = _epoxy(estimate_writer.fill_estimate({}))
    for r in (23, 27, 28, 32, 33, 39):
        assert ws[f"A{r}"].value in (None, "")
