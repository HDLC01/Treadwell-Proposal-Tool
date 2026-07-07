"""Google-Sheets-style typed formulas (Phase 1).

A "=…" string typed into the grid must persist to the generated .xlsx as a
REAL formula — but only when it passes the conservative validator (charset +
function whitelist). Injection shapes (WEBSERVICE/HYPERLINK exfil, cmd| DDE,
external-workbook refs) keep the apostrophe-escape they always had.
"""
import io

from openpyxl import load_workbook

import estimate_writer as ew


# ── validator ────────────────────────────────────────────────────────────
def test_safe_formulas_accepted():
    for f in (
        "=C60+C59",
        "=SUM(D18:D39)",
        "='Exam Room'!E20*0.5",
        '=IF(B5="yes",1,0)',
        "=ROUND(D70*1.05,2)",
        "=E20*0.33+MAX(B41,0)",
        "=Polish!D64-Epoxy!D70",
    ):
        assert ew._is_safe_formula(f), f
        assert ew._coerce(f) == f, f


def test_dangerous_formulas_escaped():
    for f in (
        '=WEBSERVICE("http://evil")',
        '=HYPERLINK("http://x","y")',
        "=cmd|' /C calc'!A0",
        "=[Book1]Sheet1!A1",
        '=INDIRECT("A"&B1)',          # not whitelisted — classic guard bypass
        '=IMPORTXML("http://x","/")',
    ):
        assert not ew._is_safe_formula(f), f
        assert ew._coerce(f) == "'" + f, f


def test_non_formula_coercion_unchanged():
    assert ew._coerce("12000") == 12000
    assert ew._coerce("0.07") == 0.07
    assert ew._coerce("hello") == "hello"
    assert ew._coerce("@mention") == "'@mention"
    assert ew._coerce("+1 (913) 555") == "'+1 (913) 555"


# ── persistence through fill_estimate ────────────────────────────────────
def test_typed_formula_lands_as_formula():
    data = ew.fill_estimate({}, cell_values={"Epoxy!E61": "=C60+C59",
                                             "Epoxy!F61": '=WEBSERVICE("http://evil")'})
    ws = load_workbook(io.BytesIO(data))["Epoxy"]
    assert ws["E61"].value == "=C60+C59"                      # real formula
    assert ws["F61"].value == "'=WEBSERVICE(\"http://evil\")"  # escaped text
