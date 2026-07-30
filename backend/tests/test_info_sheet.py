"""Project Info Sheet — the ops hand-off workbook.

Three things are worth pinning here, because each fails silently rather than
loudly:

  1. **The template.** Its dropdowns only survive because
     `prepare_info_sheet_template.py` rebuilt them around defined names. A
     regenerated or hand-edited template that loses them still opens fine, still
     fills, and still downloads — the estimator just gets free-text where the
     market segment should be, and Foundation gets a category nobody can report
     on. The market list in particular is Kyle's, verbatim, typo included.

  2. **The prefill mapping.** A wrong cell here is a plausible-looking number in
     an accounting document. Tax exempt is the sharp edge: the estimate stores
     "is this taxable", the info sheet asks the opposite question.

  3. **What fill refuses to write.** The other four tabs read the Info Sheet
     through formulas, so a stray key landing on B18 or on a label breaks
     Foundation Import and the Invoice, not this page.
"""
import io
import pathlib
import re

import openpyxl
import pytest
from fastapi.testclient import TestClient

import estimate_writer as ew
import info_sheet_writer as isw
import main

client = TestClient(main.app)


def _draft(**data):
    base = {
        "project_name": "Westport Commons", "address": "1200 Main St",
        "city": "Kansas City", "state": "MO", "zip": "64105",
        "audience": "Direct", "work_type": "epoxy",
        "contact_name": "Rita Alvarez", "contact_phone": "816-555-0134",
        "contact_email": "rita@acme.com",
        "system_name": "Treadwell MACRO Flake",
        "proposal_lump_sum": 82496.0,
    }
    base.update(data)
    return {"id": "d1", "owner_email": "kyle@wetreadwell.com", "data": base}


# ── 1. The template ───────────────────────────────────────────────────
def test_every_dropdown_survived_the_template_rebuild():
    """openpyxl drops the master's x14 validations on save. If these six are
    missing, the sheet still works and every picker is silently free-text."""
    drops = isw.read_sheet(isw.SHEET)["dropdowns"]
    for addr in ("B16", "B17", "B19", "B60", "B62", "B59", "B61", "B63", "B64",
                 "B66", "B67", "B68", "F33"):
        assert drops.get(addr), f"{addr} lost its dropdown"
    assert drops["B59"] == ["N", "Y"]


def test_market_segments_are_kyles_list_verbatim():
    """Hanz: do not add, delete or reword the Project Class options. The
    "Industial" typo is in the master and stays until he fixes it there."""
    market = isw.read_sheet(isw.SHEET)["dropdowns"]["B16"]
    assert market[0] == "-Select-"
    assert len(market) == 19
    for segment in ("Animal Care", "Correctional", "Food & Bev Manufacturing",
                    "Industial & Manufacturing", "Multifamily & Sr. Living",
                    "Religious", "Technology"):
        assert segment in market, segment


def test_the_dropdown_source_tab_is_gone_but_the_working_tabs_remain():
    wb = openpyxl.load_workbook(isw.TEMPLATE_PATH)
    assert "Packet " not in wb.sheetnames        # only ever fed the validations
    assert wb["Lists"].sheet_state == "hidden"
    for tab in ("Info Sheet", "SOV", "Foundation Import", "Invoice"):
        assert tab in wb.sheetnames


def test_the_other_tabs_still_read_the_info_sheet():
    """Foundation Import and the Invoice are pure formula mirrors — that is the
    whole reason this tool only writes one tab."""
    wb = openpyxl.load_workbook(isw.TEMPLATE_PATH)
    assert wb["Foundation Import"]["A1"].value == "='Info Sheet'!B14"   # job no.
    assert wb["Foundation Import"]["E1"].value == "='Info Sheet'!B57"   # contract
    assert wb["Invoice"]["C14"].value == "='Info Sheet'!B15"            # project


def test_no_cell_is_marked_read_only():
    """Every cell is editable now, like the estimate grid. A derived cell is
    signalled by shipping its formula, not by being locked out."""
    cells = {c["addr"]: c for c in isw.read_sheet(isw.SHEET)["cells"]}
    assert not any(c.get("readOnly") for c in cells.values())


def test_formula_cells_ship_their_formula_text():
    """The grid shows the computed value at rest and the formula on focus, so
    tabbing through a derived cell cannot flatten it into a constant."""
    cells = {c["addr"]: c for c in isw.read_sheet(isw.SHEET)["cells"]}
    for addr in ("B18", "F21", "B65", "B69", "B71"):
        assert cells[addr].get("isFormula"), f"{addr} lost its formula flag"
        assert str(cells[addr].get("formula", "")).startswith("="), addr


# ── 1b. Where the cost figures come from ──────────────────────────────
_JS = (pathlib.Path(__file__).resolve().parents[2]
       / "frontend" / "js" / "estimate-review.js").read_text(encoding="utf-8")


@pytest.mark.parametrize("layout,sheet", [
    ("Epoxy", "Epoxy"), ("Polish", "Polish"), ("Gyp", 'Gyp (USG 1-8")'),
])
def test_the_cost_cells_are_the_ones_the_estimate_sheet_labels(layout, sheet):
    """B58 and I58 are read off the estimate via `state.cost_snapshot`, which
    snapshots two cells per layout. Kyle labels both on the sheet, one column to
    the right — so assert against his label rather than trusting a coordinate
    that was transcribed by hand and would otherwise fail as a plausible number.
    """
    # Scope to TOTAL_CELLS — AREA_SF_CELLS earlier in the file has its own
    # `Epoxy: {…}` and would match first.
    totals = re.search(r"const TOTAL_CELLS = \{(.*?)\n\};", _JS, re.S).group(1)
    block = re.search(rf"^\s*{layout}:\s*\{{(.*?)\}}", totals, re.S | re.M).group(1)
    coords = dict(re.findall(r'(\w+):\s*"([A-Z]+\d+)"', block))
    ws = openpyxl.load_workbook(ew.TEMPLATE_PATH)[sheet]

    def label_beside(addr):
        col, row = re.match(r"([A-Z]+)(\d+)", addr).groups()
        return str(ws[f"{chr(ord(col) + 1)}{row}"].value or "")

    assert 'Estimated Costs" w/taxes + fees' in label_beside(coords["costs"])
    assert "Man Hour Budget" in label_beside(coords["man_hours"])


# ── 2. The prefill ────────────────────────────────────────────────────
def test_fills_the_project_and_contact_block():
    pf = isw.build_prefill(_draft())
    assert pf["B13"] == "Kyle"                  # no estimator_name yet
    assert pf["B15"] == "Westport Commons"
    assert pf["B19"] == "MO - Missouri"
    assert (pf["B20"], pf["B21"], pf["D21"]) == ("1200 Main St", "Kansas City", "64105")
    assert pf["B25"] == "Kansas City, MO 64105"
    assert (pf["B29"], pf["B30"], pf["B31"]) == ("Rita Alvarez", "816-555-0134", "rita@acme.com")
    assert pf["B57"] == 82496.0


def test_a_typed_estimator_name_beats_the_account_it_was_saved_under():
    pf = isw.build_prefill(_draft(estimator_name="Troy Holmes"))
    assert pf["B13"] == "Troy Holmes"


def test_an_autopilot_draft_has_no_estimator_yet():
    """The lead autopilot creates drafts before anyone owns them. Printing
    "Autopilot" in the Estimator / Sales Rep box reads as a person's name on a
    document accounting files."""
    d = {"owner_email": "autopilot", "data": _draft()["data"]}
    assert "B13" not in isw.build_prefill(d)


def test_gc_jobs_bill_the_contractor_direct_jobs_bill_the_owner():
    assert isw.build_prefill(_draft())["B23"] == "Westport Commons"
    gc = isw.build_prefill(_draft(audience="GC", architect="Titan Construction"))
    assert gc["B23"] == "Titan Construction"


def test_tax_exempt_is_the_inverse_of_the_estimates_taxable_flag():
    """The estimate asks "Taxable?"; the info sheet asks "Tax Exempt?". Getting
    this backwards tells accounting to chase a certificate on a taxable job, or
    to charge tax on an exempt one."""
    taxed = isw.build_prefill(_draft(cell_values={"Epoxy!B6": "Yes"}))
    assert taxed["B66"] == "N"
    exempt = isw.build_prefill(_draft(cell_values={"Epoxy!B6": "No"}))
    assert exempt["B66"] == "Y"


def test_prevailing_wage_and_remodel_tax_carry_across():
    pf = isw.build_prefill(_draft(cell_values={"Epoxy!D5": "Yes", "Epoxy!D6": "Yes"}))
    assert pf["B63"] == "Y" and pf["B67"] == "Y"


def test_an_untouched_flag_leaves_the_templates_default_alone():
    pf = isw.build_prefill(_draft())
    for addr in ("B63", "B66", "B67"):
        assert addr not in pf


def test_a_gyp_job_never_inherits_epoxys_taxable_answer():
    """In the estimate template `Gyp (USG 1-8")!B8` (Taxable) is its own literal
    "Yes", while D7/D8 (prevailing wage, remodel) really are `=Epoxy!…` mirrors.
    The gyp bid's sales tax comes from `=IF($B$8="no",…)` on the gyp tab, and
    nothing there reads Epoxy!B6 — but AI Autofill writes every flag to hardcoded
    `Epoxy!` keys whatever the work type. Falling through printed "Tax Exempt? Y"
    on a job that was priced WITH tax, and told Foundation the same."""
    d = _draft(work_type="gyp", base_tab_id='Gyp (USG 1-8")',
               cell_values={"Epoxy!B6": "No"})          # autofill's epoxy answer
    assert "B66" not in isw.build_prefill(d)
    # The gyp tab's own answer is still honoured when somebody set it.
    d2 = _draft(work_type="gyp", base_tab_id='Gyp (USG 1-8")',
                cell_values={'Gyp (USG 1-8")!B8': "No", "Epoxy!B6": "Yes"})
    assert isw.build_prefill(d2)["B66"] == "Y"


def test_a_gyp_job_does_inherit_the_two_flags_that_are_real_mirrors():
    """D7/D8 on the gyp tab are literally `=Epoxy!D5` / `=Epoxy!D6`, so reading
    epoxy for those is reading the same cell, not guessing."""
    d = _draft(work_type="gyp", base_tab_id='Gyp (USG 1-8")',
               cell_values={"Epoxy!D5": "Yes", "Epoxy!D6": "Yes"})
    pf = isw.build_prefill(d)
    assert pf["B63"] == "Y" and pf["B67"] == "Y"


def test_the_gyp_tabs_own_wage_and_remodel_answers_win():
    d = _draft(work_type="gyp", base_tab_id='Gyp (USG 1-8")',
               cell_values={'Gyp (USG 1-8")!D7': "Yes", 'Gyp (USG 1-8")!D8': "No",
                            "Epoxy!D5": "No", "Epoxy!D6": "Yes"})
    pf = isw.build_prefill(d)
    assert pf["B63"] == "Y" and pf["B67"] == "N"


def test_gyp_reads_its_own_taxable_cell_and_falls_back_to_epoxy():
    """The gyp layout sits two rows lower and keeps its own Taxable cell, but
    mirrors Prevailing Wage off epoxy by formula — so an unset gyp key legitimately
    falls through rather than reading the wrong row."""
    gyp = _draft(work_type="gyp", cell_values={
        'Gyp (USG 1-8")!B8': "No",     # its own
        "Epoxy!D5": "Yes",             # mirrored
    })
    pf = isw.build_prefill(gyp)
    assert pf["B66"] == "Y"            # exempt
    assert pf["B63"] == "Y"            # prevailing wage found on epoxy
    assert pf["B17"] == "Gypsum Cement Underlayment"


def test_an_option_tabs_product_name_cannot_decide_primary_floor():
    """B17 drives B18 Division, which is how the job is filed in Foundation.
    Pooling every tab's product names let a quartz alternate outrank the flake
    system that was actually sold."""
    d = _draft(work_type="epoxy", base_tab_id="Epoxy",
               system_name="", sheet_area={"epoxy_sf": 8000},
               priced_tabs=[
                   {"id": "Epoxy", "kind": "base", "role": "epoxy",
                    "sys_names": ["MACRO Flake Single Broadcast", ""],
                    "sf": {"epoxy_sf": 8000}},
                   {"id": "Epoxy-2", "kind": "copy", "role": "epoxy",
                    "sys_names": ["Decorative Quartz Double Broadcast", ""],
                    "sf": {"epoxy_sf": 8000}},
               ])
    assert isw.build_prefill(d)["B17"] == "Epoxy - Flake"


@pytest.mark.parametrize("system,expected", [
    ("Treadwell MACRO Flake", "Epoxy - Flake"),
    ("Treadwell Decorative Quartz", "Epoxy - Quartz"),
    ("Treadwell Poly-Crete SLB", "Epoxy - Urethane Cement"),
    # A hybrid flake system is a urethane-cement floor — the urethane test has
    # to win over the flake one, which is why the match order is fixed.
    ("Flake with hybrid blend, no base", "Epoxy - Urethane Cement"),
    ("Something we have never installed", ""),
])
def test_primary_floor_is_read_off_the_system(system, expected):
    assert isw.build_prefill(_draft(system_name=system)).get("B17", "") == expected


def test_polish_variants_are_matched_too():
    pf = isw.build_prefill(_draft(work_type="polish", system_name="Polished Concrete, Salt & Pepper"))
    assert pf["B17"] == "Polish - S&P"


def test_the_designated_base_tab_beats_the_intake_work_type():
    """The estimator can price several tabs and nominate one. The hand-off has to
    describe the bid that was actually sent."""
    d = _draft(work_type="epoxy", base_tab_id="polish-2",
               system_name="Polished Concrete - Cream",
               priced_tabs=[{"id": "Epoxy", "kind": "base", "role": "epoxy"},
                            {"id": "polish-2", "role": "polish"}])
    assert isw.build_prefill(d)["B17"] == "Polish - Cream"


def test_areas_come_from_the_sheet_and_fall_back_to_intake():
    sheet = isw.build_prefill(_draft(
        sheet_area={"epoxy_sf": 8000, "epoxy_sf_2": 386, "cove_lf": 120}))
    assert sheet["B42"] == 8386 and sheet["D42"] == 120
    # A draft saved before the sheet snapshot existed still gets its takeoff.
    old = isw.build_prefill(_draft(system_1_sf=5000, cove_1_lf=40))
    assert old["B42"] == 5000 and old["D42"] == 40


# Every real draft carries all seven priced tabs, all of them `kind: "base"`, with
# the five gyp variants sitting at zero square feet. Tests that invent a tidier
# shape pass while production does the wrong thing, so this mirrors what staging
# actually stores.
def _real_priced_tabs(*, epoxy_sf=0, cove_lf=0, polish_sf=0, polish_desc=""):
    gyp_zero = {"gyp_soft_sf": 0, "gyp_hard_sf": 0, "gyp_corridor_sf": 0}
    return [
        {"id": "Epoxy", "kind": "base", "role": "epoxy", "system_desc": "",
         "sf": {"epoxy_sf": epoxy_sf, "epoxy_sf_2": 0, "cove_lf": cove_lf, "cove_lf_2": 0}},
        {"id": "Polish", "kind": "base", "role": "polish", "system_desc": polish_desc,
         "sf": {"polish_sf": polish_sf}},
        {"id": 'Gyp (USG 1-8")', "kind": "base", "role": "gyp",
         "system_desc": 'N12 1/8"', "sf": dict(gyp_zero)},
        {"id": "Gyp (USG N12ULTRA)", "kind": "base", "role": "gyp",
         "system_desc": "N12 ULTRA", "sf": dict(gyp_zero)},
        {"id": "Gyp (FR)", "kind": "base", "role": "gyp",
         "system_desc": "Gyp (FR)", "sf": dict(gyp_zero)},
    ]


def test_a_combo_job_fills_the_sheets_second_system_block():
    """Epoxy + polish prices both a resin and a polish tab. Only one can be block
    one, so the other has to land in rows 45-49 or it never reaches the hand-off."""
    d = _draft(work_type="combo", system_name="Treadwell MACRO Flake",
               sheet_area={"epoxy_sf": 8000, "cove_lf": 120, "polish_sf": 4200},
               priced_tabs=_real_priced_tabs(epoxy_sf=8000, cove_lf=120, polish_sf=4200,
                                             polish_desc="Treadwell Polished Concrete"))
    pf = isw.build_prefill(d)
    assert (pf["B40"], pf["B42"], pf["D42"]) == ("Treadwell MACRO Flake", 8000, 120)
    assert (pf["B46"], pf["B48"]) == ("Treadwell Polished Concrete", 4200)
    assert "D48" not in pf                       # cove is not a polish quantity


def test_the_second_block_ignores_a_tab_the_customer_did_not_buy():
    """An option is an alternate that was quoted and declined; B57 covers the base
    scope only. Listing it as a second system would have ops order material and
    book crews for work nobody sold."""
    d = _draft(work_type="epoxy", base_tab_id="Epoxy",
               sheet_area={"epoxy_sf": 8000},
               priced_tabs=_real_priced_tabs(epoxy_sf=8000, polish_sf=4200,
                                             polish_desc="Treadwell Polished Concrete"),
               tab_opts={"Polish": {"is_option": True, "show": True,
                                    "price_mode": "total"}})
    pf = isw.build_prefill(d)
    for addr in ("B46", "B48", "D48"):
        assert addr not in pf, f"{addr} reported an option as sold scope"
    # Same draft without the option flag DOES report it — proves the guard is the
    # thing doing the work, not an unrelated skip.
    d2 = dict(d)
    d2["data"] = {**d["data"], "tab_opts": {}}
    assert isw.build_prefill(d2)["B48"] == 4200


def test_cove_is_not_reported_on_a_gypsum_handoff():
    """Gyp is a mobilization-based underlayment with no cove product, and its
    sheet snapshot has no cove cells — so any cove figure is a leftover intake
    number from a different scope."""
    pf = isw.build_prefill(_draft(work_type="gyp", base_tab_id='Gyp (USG 1-8")',
                                  gyp_soft_sf=20000, cove_1_lf=40))
    assert pf["B42"] == 20000 and "D42" not in pf


def test_the_second_block_skips_the_unpriced_gyp_variants():
    """The five gyp variants ride along on every draft at zero SF. Taking one by
    tab order would print 'N12 1/8"' on the sheet as if somebody had bid it."""
    d = _draft(work_type="epoxy", sheet_area={"epoxy_sf": 8000},
               priced_tabs=_real_priced_tabs(epoxy_sf=8000))   # polish at 0 too
    pf = isw.build_prefill(d)
    for addr in ("B46", "B48", "D48"):
        assert addr not in pf, f"{addr} came from a tab nobody priced"


def test_the_second_block_names_a_system_that_has_area_but_no_derived_name():
    """Real combo drafts carry `system_desc: ""` on the polish tab. An SF with no
    system beside it reads as a data-entry slip."""
    d = _draft(work_type="combo", sheet_area={"epoxy_sf": 8000, "polish_sf": 3000},
               priced_tabs=_real_priced_tabs(epoxy_sf=8000, polish_sf=3000))
    pf = isw.build_prefill(d)
    assert pf["B46"] == "Polished Concrete" and pf["B48"] == 3000


def test_a_single_system_job_leaves_the_second_block_empty():
    pf = isw.build_prefill(_draft(sheet_area={"epoxy_sf": 8000}))
    for addr in ("B46", "B48", "D48"):
        assert addr not in pf


def test_work_type_decides_the_base_when_no_tab_is_nominated():
    """`kind == "base"` does not discriminate — every priced tab has it. Reading
    the first one answered "epoxy" for every job, which sent a polish bid's SF
    looking for epoxy cells."""
    d = _draft(work_type="polish", base_tab_id=None,
               system_name="Polished Concrete - Cream",
               sheet_area={"polish_sf": 4070},
               priced_tabs=_real_priced_tabs(polish_sf=4070))
    pf = isw.build_prefill(d)
    assert pf["B42"] == 4070
    assert pf["B17"] == "Polish - Cream"


def test_cove_is_not_reported_against_a_polish_floor():
    """Cove base belongs to the resin systems; carrying an epoxy tab's LF onto a
    polish-only hand-off would have ops order material for work nobody bid."""
    pf = isw.build_prefill(_draft(work_type="polish",
                                  sheet_area={"polish_sf": 4200, "cove_lf": 120}))
    assert pf["B42"] == 4200 and "D42" not in pf


def test_costs_and_man_hours_come_from_the_snapshot_and_are_optional():
    filled = isw.build_prefill(_draft(cost_snapshot={"costs": 49614, "man_hours": 392}))
    assert filled["B58"] == 49614 and filled["I58"] == 392
    bare = isw.build_prefill(_draft())
    assert "B58" not in bare and "I58" not in bare


def test_deposit_answers_the_portal_not_the_draft():
    assert isw.build_prefill(_draft(), deposit_requested=True)["B59"] == "Y"
    assert isw.build_prefill(_draft())["B59"] == "N"


def test_an_unknown_lead_source_clears_the_templates_guess():
    """The template ships "Repeat Customer". Left alone it would be reported out
    of Foundation as though someone had chosen it."""
    assert isw.build_prefill(_draft(source="google_lead"))["B62"] == "Online"
    assert isw.build_prefill(_draft(source="email"))["B62"] == ""
    assert isw.build_prefill(_draft())["B62"] == ""


def test_nothing_pink_is_prefilled():
    """Market, payment terms, retainage, CCIP and the transition note are the
    estimator's call — Manufacturer and Color/Blend too, per Hanz."""
    pf = isw.build_prefill(_draft(cost_snapshot={"costs": 1, "man_hours": 1},
                                  cell_values={"Epoxy!B6": "Yes"}))
    for addr in ("B16", "B39", "B41", "B43", "B60", "B61", "B68", "B70"):
        assert addr not in pf, f"{addr} should be left blank"


# ── 3. Fill ───────────────────────────────────────────────────────────
def _filled(prefill, overrides=None):
    return openpyxl.load_workbook(
        io.BytesIO(isw.fill_info_sheet(prefill, overrides)))["Info Sheet"]


def test_what_the_estimator_typed_beats_the_prefill():
    ws = _filled({"B15": "Westport Commons"}, {"Info Sheet!B15": "Westport Commons Ph2"})
    assert ws["B15"].value == "Westport Commons Ph2"


def test_clearing_a_cell_clears_it():
    ws = _filled({"B62": "Online"}, {"Info Sheet!B62": ""})
    assert ws["B62"].value is None


def test_the_job_number_stays_text():
    """"26.100" cast to a float is 26.1, and B14 is what the Invoice tab and
    Foundation Import both print."""
    ws = _filled({"B14": "26.100"})
    assert ws["B14"].value == "26.100"


def test_typing_over_a_formula_or_a_label_replaces_it():
    """Hanz asked for the estimate grid's behaviour: any cell, including the
    derived ones and the labels. The formula-injection guard still applies."""
    ws = _filled({}, {"Info Sheet!B18": "Epoxy - Commercial",   # was derived
                      "Info Sheet!A15": "Job Name",             # was a label
                      "Info Sheet!B71": "N/A"})                 # was derived
    assert ws["B18"].value == "Epoxy - Commercial"
    assert ws["A15"].value == "Job Name"
    assert ws["B71"].value == "N/A"


def test_an_untouched_formula_is_left_alone():
    """Only cells the estimator actually sent are written."""
    ws = _filled({"B15": "Westport"})
    assert str(ws["B18"].value).startswith("=IF(")
    assert ws["A15"].value == "Project Name (Description):"


def test_a_typed_formula_trigger_is_neutralized():
    ws = _filled({}, {"Info Sheet!B15": "=cmd|'/c calc'!A0",
                      "Info Sheet!B14": "@SUM(1,1)"})
    assert str(ws["B15"].value).startswith("'=")
    assert str(ws["B14"].value).startswith("'@")


def test_every_prefilled_cell_actually_survives_into_the_workbook():
    """EDITABLE is the only gate deciding which prefilled values reach the file.
    A cell missing from it is dropped SILENTLY and still renders on screen, so the
    estimator would download a sheet whose contract amount or tax flag had
    vanished. Round-trips the whole prefill rather than the handful of cells the
    other fill tests poke at."""
    draft = _draft(job_number="26.153", estimator_name="Troy Holmes",
                   audience="Direct",
                   sheet_area={"epoxy_sf": 8000, "cove_lf": 120, "polish_sf": 4200},
                   priced_tabs=_real_priced_tabs(epoxy_sf=8000, cove_lf=120,
                                                 polish_sf=4200,
                                                 polish_desc="Treadwell Polished Concrete"),
                   work_type="combo", base_tab_id="Epoxy",
                   cost_snapshot={"costs": 49614, "man_hours": 392},
                   cell_values={"Epoxy!B6": "No", "Epoxy!D5": "Yes", "Epoxy!D6": "Yes"},
                   source="google_lead")
    prefill = isw.build_prefill(draft, deposit_requested=True)
    assert len(prefill) >= 25, "prefill got thinner — is this test still meaningful?"
    ws = _filled(prefill)
    missing = [a for a, v in prefill.items() if v != "" and ws[a].value in (None, "")]
    assert not missing, f"prefilled but absent from the .xlsx: {missing}"


def test_the_download_keeps_the_treadwell_logo():
    """openpyxl needs Pillow to write back images it read; without it every logo
    is dropped on save with no warning. The template carries the Treadwell mark on
    three tabs — one of them the Invoice — and staging shipped a 27 KB file with
    all three missing because Pillow was only ever a transitive local dependency.

    Asserted on the zip parts rather than on openpyxl's view, because that is what
    Excel actually opens.
    """
    import zipfile
    out = isw.fill_info_sheet({"B15": "Westport"})
    media = [n for n in zipfile.ZipFile(io.BytesIO(out)).namelist()
             if n.startswith("xl/media/")]
    assert len(media) == 3, f"lost the logo — is Pillow installed? parts: {media}"
    wb = openpyxl.load_workbook(io.BytesIO(out))
    for tab in ("Info Sheet", "Invoice", "Deposit"):
        assert wb[tab]._images, f"{tab} lost its logo"


def test_the_download_keeps_its_dropdowns_and_its_other_tabs():
    wb = openpyxl.load_workbook(io.BytesIO(isw.fill_info_sheet({"B15": "Westport"})))
    assert wb.sheetnames == ["Info Sheet", "SOV", "Foundation Import",
                             "Invoice", "Deposit", "Lists"]
    assert len(wb["Info Sheet"].data_validations.dataValidation) == 7
    assert "MarketList" in wb.defined_names


# ── The endpoints ─────────────────────────────────────────────────────
@pytest.fixture
def one_draft(monkeypatch):
    """A saved draft, a portal that says no deposit, and a capture of any
    write-back so the job-number mirror can be asserted."""
    saved = {}
    monkeypatch.setattr(main.drafts, "load_draft", lambda i: _draft() if i == "d1" else None)
    monkeypatch.setattr(main.drafts, "save_draft",
                        lambda i, d, **kw: saved.update(id=i, data=d))
    monkeypatch.setattr(main.drafts, "log_event", lambda *a, **kw: None)
    monkeypatch.setattr(main, "_portal", lambda *a, **kw: {"proposals": []})
    return saved


def test_get_returns_the_grid_and_the_prefill(one_draft):
    body = client.get("/api/info-sheet/d1").json()
    assert body["order"] == ["Info Sheet", "SOV", "Foundation Import", "Invoice", "Deposit"]
    assert body["sheets"]["Info Sheet"]["sheet"] == "Info Sheet"
    assert body["template_version"]
    # The prefill is merged into the cells, so the grid shows exactly what the
    # download will write — no second source of truth to reconcile.
    cells = {c["addr"]: c for c in body["sheets"]["Info Sheet"]["cells"]}
    assert cells["B15"]["value"] == "Westport Commons"
    assert cells["B15"]["role"] == "prefill"
    assert cells["B16"]["role"] == "decision"        # pink, left for a human


def test_get_404s_on_an_unknown_draft(one_draft):
    assert client.get("/api/info-sheet/nope").status_code == 404


def test_generate_returns_a_download_and_mirrors_the_job_number(one_draft, monkeypatch):
    """`job_number` is typed here for the first time anywhere in the tool, and
    the deposit invoice already reads it off the draft."""
    d = _draft(info_cell_values={"Info Sheet!B14": "26.153"})
    monkeypatch.setattr(main.drafts, "load_draft", lambda i: d)
    r = client.post("/api/info-sheet/generate", json={"draft_id": "d1"})
    assert r.status_code == 200
    url = r.json()["xlsx_download_url"]
    assert client.get(url).status_code == 200
    assert one_draft["data"]["job_number"] == "26.153"


def test_generate_uses_the_cells_sent_with_the_request(one_draft, monkeypatch):
    """The page autosaves onto the draft, but that PUT is debounced 2.5 s and every
    keystroke restarts the timer. Rebuilding from the saved draft handed over a
    workbook missing whatever was typed in the last few seconds — the market
    segment and job number — while the button still said "Downloaded". The cells
    now travel with the request."""
    stale = _draft(info_cell_values={"Info Sheet!B16": "Retail"})
    monkeypatch.setattr(main.drafts, "load_draft", lambda i: stale)
    r = client.post("/api/info-sheet/generate", json={
        "draft_id": "d1",
        "info_cell_values": {"Info Sheet!B16": "Religious", "Info Sheet!B14": "26.153"}})
    assert r.status_code == 200
    ws = openpyxl.load_workbook(
        io.BytesIO(client.get(r.json()["xlsx_download_url"]).content))["Info Sheet"]
    assert ws["B16"].value == "Religious"      # not the draft's stale "Retail"
    assert ws["B14"].value == "26.153"
    # The write-back carries the cells too, so the page's own later PUT of a
    # localStorage copy cannot silently revert the job number.
    assert one_draft["data"]["job_number"] == "26.153"
    assert one_draft["data"]["info_cell_values"]["Info Sheet!B16"] == "Religious"


def test_generate_still_works_for_a_page_that_sends_no_cells(one_draft, monkeypatch):
    """An older cached page build posts only draft_id; it must keep working off
    the saved draft rather than losing every edit."""
    monkeypatch.setattr(main.drafts, "load_draft",
                        lambda i: _draft(info_cell_values={"Info Sheet!B16": "Retail"}))
    r = client.post("/api/info-sheet/generate", json={"draft_id": "d1"})
    assert r.status_code == 200
    ws = openpyxl.load_workbook(
        io.BytesIO(client.get(r.json()["xlsx_download_url"]).content))["Info Sheet"]
    assert ws["B16"].value == "Retail"


def test_generate_survives_an_unreachable_portal(one_draft, monkeypatch):
    """Deposits live in the portal. A bad minute there must not stop the
    hand-off sheet — the estimator can set B59 themselves."""
    monkeypatch.setattr(main, "_portal", lambda *a, **kw: (_ for _ in ()).throw(RuntimeError("down")))
    assert client.post("/api/info-sheet/generate", json={"draft_id": "d1"}).status_code == 200


def test_a_requested_deposit_reaches_the_sheet(one_draft, monkeypatch):
    monkeypatch.setattr(main, "_portal", lambda *a, **kw: {"proposals": [
        {"proposal_id": "d1", "deposit_requested_at": "2026-07-27T10:00:00Z"}]})
    body = client.get("/api/info-sheet/d1").json()
    cells = {c["addr"]: c for c in body["sheets"]["Info Sheet"]["cells"]}
    assert cells["B59"]["value"] == "Y"


# ── The shared reader serves two workbooks ────────────────────────────
# estimate_writer's caches were keyed by (sheet name, mtime) with no path, so
# once a second workbook started reading through them one template's sheet
# could be handed back for the other's request. That failure looks like real
# data, which is why it needs a behavioural test and not a key-shape one.
@pytest.fixture
def _cold_caches():
    ew._WB_CACHE.clear(); ew._SHEET_GRID_CACHE.clear()
    yield
    ew._WB_CACHE.clear(); ew._SHEET_GRID_CACHE.clear()


def test_the_workbook_cache_keeps_the_two_templates_apart(_cold_caches):
    est = ew._load_template(data_only=False)
    info = ew._load_template(data_only=False, path=isw.TEMPLATE_PATH)
    assert est is not info
    assert "Epoxy" in est.sheetnames and "Epoxy" not in info.sheetnames
    assert "Info Sheet" in info.sheetnames
    assert ew._load_template(data_only=False) is est, "the estimate was evicted"


def test_the_grid_cache_cannot_serve_one_templates_sheet_out_of_another(_cold_caches):
    """Warm Epoxy, then ask the info workbook for it. A path-less cache hands
    back the warmed Epoxy grid; a correct one has never heard of the sheet."""
    ew.read_sheet_grid("Epoxy")
    with pytest.raises(KeyError):
        ew.read_sheet_grid("Epoxy", path=isw.TEMPLATE_PATH)
    ew.read_sheet_grid("Info Sheet", path=isw.TEMPLATE_PATH)
    with pytest.raises(KeyError):
        ew.read_sheet_grid("Info Sheet")


def test_a_defined_name_dropdown_resolves(_cold_caches):
    """Every picker on the Info Sheet points at a workbook-level name, because
    that is the only cross-sheet source form Excel and openpyxl both keep. The
    shared resolver could not follow one, so reusing it dropped all six."""
    wb = openpyxl.load_workbook(isw.TEMPLATE_PATH)
    opts = ew._resolve_range_to_options(wb, wb["Info Sheet"], "MarketList",
                                        path=isw.TEMPLATE_PATH)
    assert len(opts) == 19 and opts[0] == "-Select-" and "Religious" in opts


def test_the_estimates_own_dropdowns_still_resolve(_cold_caches):
    """The defined-name branch must not disturb literal ranges."""
    wb = ew._load_template(data_only=False)
    opts = ew._resolve_range_to_options(wb, wb["Epoxy"], "$B$161:$B$165")
    assert len(opts) == 5 and opts[0] == "Primer Options"


def test_a_row_insert_moves_the_info_sheets_dropdowns_with_their_cells(_cold_caches):
    """The headline case. Insert above the Y/N block and every picker below it
    has to come along, or the estimator gets free-text where a list should be
    and Foundation gets a category nobody can report on."""
    wb = openpyxl.load_workbook(isw.TEMPLATE_PATH)
    ew._apply_tab_structs(wb, ew._norm_structs(
        [{"sheet": "Info Sheet", "kind": "insert_rows", "at": 20, "count": 2}]))
    ws = wb["Info Sheet"]
    by_src = {d.formula1: str(d.sqref) for d in ws.data_validations.dataValidation}
    assert by_src["YNList"] == "B61 B63 B65:B66 B68:B70"   # was B59 B61 B63:B64 B66:B68
    assert by_src["TermsList"] == "B62"                    # was B60
    assert by_src["LeadSourceList"] == "B64"               # was B62
    assert by_src["MarketList"] == "B16"                   # above the insert, unmoved
    # And the payroll formula that slid into the old Y/N territory did not
    # inherit a Y/N picker.
    assert "B67" not in " ".join(by_src.values())
    assert str(ws["B67"].value).startswith("=IF(")


def test_a_row_insert_keeps_the_other_tabs_pointing_at_the_right_cells(_cold_caches):
    """Foundation Import, Invoice and Deposit read this sheet by address."""
    wb = openpyxl.load_workbook(isw.TEMPLATE_PATH)
    ew._apply_tab_structs(wb, ew._norm_structs(
        [{"sheet": "Info Sheet", "kind": "insert_rows", "at": 20, "count": 2}]))
    assert wb["Foundation Import"]["A1"].value == "='Info Sheet'!B14"   # above, unmoved
    assert wb["Invoice"]["C11"].value == "='Info Sheet'!B62"            # B60 pushed down 2
    assert wb["Invoice"]["C14"].value == "='Info Sheet'!B15"            # above, unmoved


# ── Five tabs, every cell writable ────────────────────────────────────
def test_the_visible_tabs_are_exactly_the_five_we_expect():
    """Derived from the workbook, so a master that unhides a tab or adds one
    fails here rather than quietly putting it on screen."""
    assert isw.visible_sheets() == ["Info Sheet", "SOV", "Foundation Import",
                                    "Invoice", "Deposit"]


def test_the_hidden_lists_tab_is_never_served():
    """It holds the dropdown source columns. Showing it invites someone to edit
    the options out from under every picker."""
    body = isw.read_workbook()
    assert "Lists" not in body["order"] and "Lists" not in body["sheets"]
    with pytest.raises(KeyError):
        isw.read_sheet("Lists")


def test_the_sov_grid_gets_the_border_symmetry_pass():
    """The reason the read path is shared rather than forked. SOV is a bordered
    table; Excel defines each wall on one side only, so without the mirror pass
    half its cells render with missing edges."""
    cells = isw.read_sheet("SOV")["cells"]
    assert sum(1 for c in cells if c.get("borders")) > 150


def test_every_visible_tab_accepts_a_write():
    ws = None
    wb = openpyxl.load_workbook(io.BytesIO(isw.fill_info_sheet({}, {
        "Info Sheet!B15": "Westport", "SOV!B7": "Submittals edit",
        "Foundation Import!C1": "CUST-9", "Invoice!C9": "26.153-01",
        "Deposit!C9": "26.153-02"})))
    assert wb["Info Sheet"]["B15"].value == "Westport"
    assert wb["SOV"]["B7"].value == "Submittals edit"
    assert wb["Foundation Import"]["C1"].value == "CUST-9"
    assert wb["Invoice"]["C9"].value == "26.153-01"
    assert wb["Deposit"]["C9"].value == "26.153-02"


def test_a_write_to_the_hidden_lists_tab_is_refused():
    wb = openpyxl.load_workbook(io.BytesIO(
        isw.fill_info_sheet({}, {"Lists!A4": "HACKED"})))
    assert wb["Lists"]["A4"].value == "-Select-"


def test_a_malformed_override_address_is_skipped_not_fatal():
    wb = openpyxl.load_workbook(io.BytesIO(isw.fill_info_sheet({}, {
        "B15": "no sheet prefix", "Info Sheet!B1:B9": "a range",
        "Info Sheet!MarketList": "a defined name", "Nope!B1": "unknown sheet",
        "Info Sheet!B15": "the good one"})))
    assert wb["Info Sheet"]["B15"].value == "the good one"


# ── Structural edits ──────────────────────────────────────────────────
_INS20 = [{"sheet": "Info Sheet", "kind": "insert_rows", "at": 20, "count": 2}]


def test_the_prefill_rides_a_structural_shift():
    """Prefill is authored in template coordinates, so it is written BEFORE the
    replay and moves with its cell — the same invariant fill_estimate keeps."""
    wb = openpyxl.load_workbook(io.BytesIO(
        isw.fill_info_sheet({"B57": 82496, "B15": "Westport"}, {}, tab_structs=_INS20)))
    ws = wb["Info Sheet"]
    assert ws["B59"].value == 82496          # B57 pushed down two
    assert ws["B15"].value == "Westport"     # above the insert, unmoved


def test_overrides_arrive_in_current_coordinates():
    """The estimator typed against the grid in front of them, which already had
    the inserted rows — so their addresses are written AFTER the replay, as-is."""
    wb = openpyxl.load_workbook(io.BytesIO(
        isw.fill_info_sheet({}, {"Info Sheet!B62": "Net 10"}, tab_structs=_INS20)))
    assert wb["Info Sheet"]["B62"].value == "Net 10"


def test_a_structural_edit_keeps_the_other_tabs_pointing_at_the_right_cells():
    wb = openpyxl.load_workbook(io.BytesIO(isw.fill_info_sheet({}, {}, tab_structs=_INS20)))
    assert wb["Invoice"]["C11"].value == "='Info Sheet'!B62"     # B60 moved
    assert wb["Foundation Import"]["A1"].value == "='Info Sheet'!B14"   # above, unmoved


def test_a_structural_edit_on_another_tab_replays_too():
    wb = openpyxl.load_workbook(io.BytesIO(isw.fill_info_sheet(
        {}, {}, tab_structs=[{"sheet": "SOV", "kind": "insert_rows", "at": 7, "count": 1}])))
    assert wb["Invoice"]["C21"].value == "=SOV!B8"               # was =SOV!B7


def test_a_structural_op_on_the_hidden_lists_tab_is_refused():
    """_apply_tab_structs only skips sheets absent from the workbook, and Lists
    is present — so without the visible-sheets gate this wipes the dropdown
    source and every picker silently goes free-text."""
    wb = openpyxl.load_workbook(io.BytesIO(isw.fill_info_sheet(
        {}, {}, tab_structs=[{"sheet": "Lists", "kind": "delete_rows",
                              "at": 4, "count": 20}])))
    market = [c[0].value for c in wb["Lists"]["A4":"A22"] if c[0].value]
    assert len(market) == 19 and market[0] == "-Select-"


def test_resolve_addr_follows_the_edits():
    assert isw.resolve_addr("B57", _INS20) == "B59"
    assert isw.resolve_addr("B14", _INS20) == "B14"          # above the insert
    assert isw.resolve_addr("B21", [{"sheet": "Info Sheet", "kind": "delete_rows",
                                     "at": 21, "count": 1}]) is None


# ── The colour key ────────────────────────────────────────────────────
def _fill_of(ws, addr):
    f = ws[addr].fill
    rgb = f.fgColor.rgb if f and f.patternType == "solid" else None
    return rgb if isinstance(rgb, str) else None


def test_prefilled_cells_come_out_chartreuse():
    ws = _filled({"B15": "Westport", "B57": 82496})
    assert _fill_of(ws, "B15") == isw.CHARTREUSE
    assert _fill_of(ws, "B57") == isw.CHARTREUSE


def test_a_field_we_knew_but_the_job_did_not_have_is_still_ticked():
    """On the marked-up FBC sheet D42 is chartreuse and blank — we knew the cove
    field, the job simply had none. Deriving the key from "cells that ended up
    non-empty" would lose that."""
    ws = _filled({"D42": None})
    assert ws["D42"].value is None
    assert _fill_of(ws, "D42") == isw.CHARTREUSE


def test_the_decisions_are_pink_even_when_a_human_filled_them():
    """Pink means "a person chose this", permanently. On FBC, B16 is pink and
    holds "Religious"."""
    ws = _filled({}, {"Info Sheet!B16": "Religious"})
    assert ws["B16"].value == "Religious"
    assert _fill_of(ws, "B16") == isw.PINK


def test_a_cell_the_tool_could_not_answer_is_not_ticked():
    ws = _filled({"B15": "Westport"})          # no cost_snapshot -> no B58 key
    assert _fill_of(ws, "B58") != isw.CHARTREUSE


def test_the_colour_key_leaves_the_labels_alone():
    ws = _filled({"B15": "Westport"})
    tpl = openpyxl.load_workbook(isw.TEMPLATE_PATH)["Info Sheet"]
    for addr in ("A15", "A57", "A20", "H5"):
        assert _fill_of(ws, addr) == _fill_of(tpl, addr), addr


def test_the_colour_key_keeps_the_number_format():
    ws = _filled({"B57": 82496, "B14": "26.100"})
    assert ws["B57"].number_format.startswith('"$"')
    assert ws["B14"].number_format == "@"


def test_the_colour_key_follows_a_structural_edit():
    wb = openpyxl.load_workbook(io.BytesIO(
        isw.fill_info_sheet({"B57": 82496}, {}, tab_structs=_INS20)))
    ws = wb["Info Sheet"]
    assert _fill_of(ws, "B59") == isw.CHARTREUSE     # prefill moved
    assert _fill_of(ws, "B62") == isw.PINK           # B60 pink moved


def test_the_prefill_and_the_pink_set_never_overlap():
    """Pink is painted last and would win. A test rather than a footnote."""
    draft = _draft(job_number="26.153", estimator_name="Troy",
                   sheet_area={"epoxy_sf": 8000, "cove_lf": 120},
                   cost_snapshot={"costs": 1, "man_hours": 1},
                   cell_values={"Epoxy!B6": "No", "Epoxy!D5": "Yes"},
                   source="google_lead")
    assert not (set(isw.build_prefill(draft, deposit_requested=True)) & isw.PINK_CELLS)
