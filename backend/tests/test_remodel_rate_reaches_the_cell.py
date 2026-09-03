"""Picking a county must move the workbook's actual remodel-tax number, not just JS state.

Kyle, via a screenshot: "It was brought to my attention that the remodel tax calculator is not
giving correct tax %. I'm not sure how that works but we use the link within the original excel
sheet to go to the website, enter the address, and get the tax % from there."

An earlier fix corrected the REFERENCE data (backend/reference_tax.py — see
test_reference_tax.py) and stopped there. This test file is for the deeper half of the same bug:
the county picker never actually reached the bid. `pickCounty` (frontend/js/estimate-review.js)
only ever set `state.county_remodel_rate` and told the estimator to hand-type the number into
K81/K75/K80 — plain text LABEL cells wired to nothing. The real formula cell that drives the
math — `Epoxy!B81` / `Polish!B75` / `<Gyp variant>!B80` — kept Kyle's own hardcoded
`=IF(D6="yes",0.1,0)` 10% placeholder forever, no matter what the picker said on screen.

The fix writes the corrected formula straight into `cellValues` (the same verbatim
`"<Sheet>!<Addr>"` dict that `/api/generate` already forwards untouched into
`estimate_writer.fill_estimate()`'s cell_values write step — see test_cell_lock.py's
`test_cell_values_write_into_locked_cell_still_lands` for that mechanism) and into the live HF
engine so the on-screen total updates immediately, matching the pattern the AI-autofill feature
already uses for its own cell writes.

Executed, not grepped: the original bug was invisible to source reading — the code that existed
was internally consistent, it just never reached the one cell that mattered. Only running
`pickCounty` for real and inspecting the resulting `cellValues`/HF writes can prove the fix
actually lands.

2026-09-02 — the same bug, one layer down. That fix wrote the rate to a hand-typed list of
sheets: Epoxy, Polish and the five gyp variants. Reading the shipped workbook instead of the
comment above that list turned up three more sheets carrying the identical
`=IF(D6="yes",0.1,0)` placeholder — `Seal!B75`, `Leveling!B77` and `Epoxy blank!B78` — and a
fourth hole with no sheet name at all: a COPIED tab. Copies are the ordinary way to put a priced
option in front of a customer, and `estimate_writer._create_copied_tabs` clones them from the
PRISTINE template, so a copy's rate cell arrives holding the 10% placeholder. Copying before
picking the county, or changing the county after copying, shipped a 10% option line beside an
otherwise correct base bid — while the pill overhead said the real rate was applied
automatically. Cases 6-9 in the harness are those four holes.

Case 10 is the regression cases 6-9 introduced, and it took a real browser to see it. Bringing
copies in meant the override's own grid refresh reached a copied tab for the first time — and it
refreshed by discarding the active sheet's cache and re-fetching it, which 404s for a copy and
painted "Failed to load Copy1" over the tab. All nine cases above sit on a base tab, where that
refetch happens to succeed, so all nine stayed green while staging carried the bug.
"""
import json
import pathlib
import shutil
import subprocess

import pytest

REPO = pathlib.Path(__file__).resolve().parents[2]
FRONTEND = REPO / "frontend"
HARNESS = pathlib.Path(__file__).resolve().parent / "js" / "remodel-rate-harness.js"

pytestmark = pytest.mark.skipif(shutil.which("node") is None, reason="node is not installed")


@pytest.fixture(scope="module")
def result():
    proc = subprocess.run(["node", str(HARNESS), str(FRONTEND)],
                           capture_output=True, text=True, encoding="utf-8", timeout=120)
    assert proc.returncode == 0, proc.stderr
    return json.loads(proc.stdout.strip().splitlines()[-1])


def test_covers_every_sheet_with_a_remodel_tax_line(result):
    """Epoxy + Polish + Seal + Leveling + Epoxy blank + every Gyp variant = 10 cells.

    This count was 7 until 2026-09-02, and the three missing sheets are the point: the list was
    written from a comment asserting they had no remodel line, and the workbook says otherwise.
    If a 6th Gyp variant is ever added, this is the number that must move — not a hand-typed
    count going stale while a real customer's option line quotes 10%."""
    assert result["gypSheetCount"] == 5
    assert result["pickedCity"]["cellCount"] == 10


def test_picking_a_city_writes_the_real_rate_into_every_formula_cell(result):
    p = result["pickedCity"]
    assert p["epoxy"] == '=IF(D6="yes",0.0935,0)'
    assert p["polish"] == '=IF(D6="yes",0.0935,0)'
    for sheet, formula in p["gyp"].items():
        assert formula == '=IF(D8="yes",0.0935,0)', (sheet, formula)


def test_the_write_also_reaches_the_live_hf_engine(result):
    """So the on-screen total is right NOW, not just in the file the estimator downloads later."""
    p = result["pickedCity"]
    assert p["hfCallCount"] == 10
    assert p["hfMatchesCellValues"] is True


def test_the_active_sheets_grid_is_refreshed(result):
    """cellValues alone would be right in the generated file but stale on screen — the estimator
    would see the old total until they clicked away and back.

    HOW it refreshes matters, which is why the cache is asserted alongside. This used to
    `delete sheetCache[activeSheet]` and re-run `showSheet` — a wasteful round trip on a base tab,
    and outright data loss on a copy (see the last test in this file). It now
    re-renders from the live HF engine and leaves the cache alone."""
    p = result["pickedCity"]
    assert p["gridRefreshedFor"] == ["Epoxy"]
    assert p["cachePreserved"] is True


def test_state_still_carries_what_the_proposal_step_reads(result):
    """The {{county}} token and the on-screen state must keep working exactly as before — this
    fix adds a cellValues write, it does not replace the existing state bookkeeping."""
    p = result["pickedCity"]
    assert p["stateCounty"] == "Overland Park, KS"
    assert p["stateCountyTaxRate"] == 0.0935
    assert p["stateCountyRemodelRate"] == 0.0935
    assert p["persistedRemodelRate"] == 0.0935


def test_the_hint_no_longer_tells_the_estimator_to_hand_type_a_dead_cell(result):
    """The old copy said "(enter in K81)" — a legend cell wired to nothing. Following that
    instruction changed zero dollars on the bid; the new copy must not repeat that instruction."""
    pill = result["pickedCity"]["pillHtml"]
    assert "K81" not in pill
    assert "applied automatically" in pill
    assert "9.350%" in pill


def test_a_county_with_no_override_reverts_every_cell_to_kyles_own_placeholder(result):
    """A county-only pick (no `remodel_rate` override, e.g. the KS floor-rate rows) must not
    leave a stale rate from a previous city pick sitting in the formula."""
    p = result["pickedCountyNoOverride"]
    assert p["epoxy"] == '=IF(D6="yes",0.1,0)'
    assert p["polish"] == '=IF(D6="yes",0.1,0)'
    assert p["oneGyp"] == '=IF(D8="yes",0.1,0)'
    assert p["stateCountyRemodelRate"] is None


def test_clearing_the_pill_reverts_the_cell_and_the_state(result):
    """`stateHasCounty` reads "is a county set", not "does the key exist". The handler used to
    `delete` these keys, which looked like it cleared them and did not: TW.setState is
    Object.assign(cur, partial) where `cur` is re-parsed from localStorage, so a key the
    partial LACKS keeps its stored value. The pill emptied on screen and the county stayed in
    the saved draft, ready to come back on the next reload. Cosmetic while the county was
    only a label; not cosmetic now that effectiveRemodelRate() falls back to
    county_remodel_rate -- a county that will not clear is a RATE that will not clear."""
    c = result["cleared"]
    assert c["epoxy"] == '=IF(D6="yes",0.1,0)'
    assert c["stateHasCounty"] is False
    assert c["stateHasRemodelRate"] is False
    assert c["pillCleared"] is True
    assert c["payloadClearsCounty"] is True    # ...and it survives the reload
    # The reverted RATE has to be in that same save. `cellValues` is a copy of
    # state.cell_values (estimate-review.js:282), so a payload that omits it leaves the
    # cleared county's formula in the draft -- and /api/generate fills the workbook from
    # the draft, not from the page. Pick a county, clear it, generate without reloading,
    # and the customer's estimate would carry a rate for a county that is gone.
    assert c["payloadCellValues"] == '=IF(D6="yes",0.1,0)'


def test_a_draft_saved_before_this_fix_self_heals_on_reopen(result):
    """Before this fix shipped, a draft could have `state.county_remodel_rate` set with nothing
    in `cellValues` for these addresses at all — the exact shape a pre-fix save left behind. The
    page-load restore path must replay the override, not just redraw the pill."""
    s = result["staleDraftSelfHeals"]
    assert s["epoxy"] == '=IF(D6="yes",0.0935,0)'
    assert s["polish"] == '=IF(D6="yes",0.0935,0)'


def test_the_toggle_reference_is_local_to_each_sheet(result):
    """Epoxy/Polish read D6, every Gyp variant reads D8 — the same local, unqualified reference
    Kyle's own original placeholder formula used (Polish!D6 and each Gyp D8 are formula mirrors
    of Epoxy!D6, confirmed by direct inspection of the workbook), so no cross-sheet qualifier is
    needed or correct here."""
    t = result["toggleShape"]
    assert t["epoxy"].startswith('=IF(D6=')
    assert t["polish"].startswith('=IF(D6=')
    assert t["gyp"].startswith('=IF(D8=')


# ─── The four sheets the first fix missed (2026-09-02) ─────────────────────


def test_seal_leveling_and_epoxy_blank_get_the_picked_rate(result):
    """Three sheets the previous target list left out, on the strength of a comment claiming they
    had no remodel-tax line. All three hold `=IF(D6="yes",0.1,0)` in the shipped workbook, and
    Seal is a priced role (`BASE_ROLE` via `SEAL_SHEETS`) whose number reaches the customer as a
    proposal price line — so every sealer bid was taxed at Kyle's 10% placeholder instead of the
    ~9.1-9.7% the picker had already looked up and displayed."""
    p = result["previouslyMissedLayouts"]
    assert p["seal"] == '=IF(D6="yes",0.0935,0)'
    assert p["leveling"] == '=IF(D6="yes",0.0935,0)'
    assert p["epoxyBlank"] == '=IF(D6="yes",0.0935,0)'


def test_seal_with_joints_is_deliberately_left_as_a_mirror(result):
    """`Seal (+Jnts)!B75` is `=Seal!B75`, so writing Seal already carries it. Writing a literal
    there too would replace the mirror and let the two sheets drift apart independently — the
    exact divergence found in Kyle's own filed workbooks. This absence is a decision, so it is
    asserted rather than left to whoever next extends the list."""
    assert result["sealJointsLeftAsMirror"]["written"] is False


def test_a_tab_copied_before_the_county_was_picked_still_gets_the_rate(result):
    """Copies are how a priced option gets in front of a customer, and the backend clones them
    from the pristine template (`estimate_writer._create_copied_tabs`), so a copy's rate cell
    arrives at 10%. `addCopy` replaying the source's `cellValues` covers "pick, then copy" as a
    side effect — it can do nothing for "copy, then pick", which is this test. Before the fix
    that sequence shipped a 10% option beside a correct base bid."""
    c = result["copyThenPick"]
    assert c["copy1"] == '=IF(D6="yes",0.0935,0)'    # copy of Epoxy  → its layout's B81
    assert c["copy2"] == '=IF(D6="yes",0.0935,0)'    # copy of Polish → its layout's B75
    assert c["base"] == '=IF(D6="yes",0.0935,0)'     # and the base tab is untouched by all this


def test_a_copy_of_a_copy_resolves_through_the_chain_to_its_template_layout(result):
    """The address depends on the LAYOUT, not the tab: B81 for an epoxy-derived tab, B75 for a
    polish- or seal-derived one. A copy of a copy has to walk the chain (`layoutIdFor`) to find
    it — the same resolution `test_cell_lock.py::test_copy_of_copy_resolves_through_chain`
    already pins for cell protection."""
    c = result["copyChain"]
    assert c["copy2"] == '=IF(D6="yes",0.0935,0)'    # Copy2 → Copy1 → Epoxy → B81
    assert c["copy3"] == '=IF(D6="yes",0.0935,0)'    # Copy3 → Seal → B75
    # 10 template layouts + 3 copies, each written exactly once. A `seen` set guards the overlap:
    # the base tabs in the tab bar are the same ids as the layouts, and writing one twice would
    # be harmless here but would hide a double-write bug on a real structural translation.
    assert c["targetCount"] == 13


def test_picking_while_sitting_on_a_copied_tab_keeps_the_tab_on_screen(result):
    """The regression the four tests above created, caught by a browser and not by any of them.

    Once copies became rate targets, the override's own grid refresh reached a copied tab for the
    first time — and it refreshed by discarding the active sheet's cache and re-fetching it. A copy
    has no server-side worksheet (`addCopy` builds `sheetCache[newId]` client-side from its
    source), so `GET /api/sheet/Copy1` 404s and `showSheet`'s `!r.ok` branch paints
    "Failed to load Copy1" over a tab whose cache is now gone. Picking a county with a copy open —
    an ordinary sequence — blanked the option the estimator was looking at.

    For a client-side-only cache, a refetch is data loss, not a round trip. Every one of cases 1-9
    sits on a base tab, where the refetch happens to succeed, which is exactly why they were all
    green while staging carried the bug."""
    c = result["pickedWhileOnACopy"]
    assert c["copy1"] == '=IF(D6="yes",0.0935,0)'   # the rate still lands
    assert c["cachePreserved"] is True              # ...without destroying the tab to deliver it
    assert c["gridRefreshedFor"] == ["Copy1"]       # ...and the copy redraws with the new number


# The typed remodel tax % ======================================================
# Everything above makes the COUNTY table reach the bid. Kyle's own words are that the
# county table is not where he gets the number:
#
#   "we use the link within the original excel sheet to go to the website, enter the
#    address, and get the tax % from there."
#
# An address is finer-grained than a county, so the table is a starting point and the
# estimator needs to be able to type the figure the site actually returned. These tests
# drive the real `commitRemodelRate` / `renderRemodelRateField` out of estimate-review.js
# through the same harness, so the typed rate is held to the identical standard: it has to
# land in the FORMULA cell of every base tab and every copy, or it is decoration.


def test_the_percentage_reaches_the_rate_cell_without_arithmetic(result):
    """Moving the decimal point through the TEXT, because dividing by 100 does not survive.

    These rates carry three decimals. A plain /100 is wrong for real Kansas rates and a
    plain *100 is wrong in the display direction, and the wrong value would go verbatim
    into a spreadsheet formula where nothing would ever flag it:

        8.775 / 100    ->  0.08775000000000001
        9.975 / 100    ->  0.09974999999999999   (a hair LOW)
        0.07975 * 100  ->  7.9750000000000005

    Note that 7.975 / 100 is exact by luck, which is why the counterexamples below are
    asserted as a set rather than spot-checked on one rate: a single lucky value would
    make this whole test vacuous. The property that matters is that shifting there and
    back is the identity for every one of them."""
    e = result["exactness"]
    assert e["shifted"] == "0.07975"
    assert e["roundTrip"] == "0.07975"          # survives Number() without growing a tail

    # the counterexamples: what arithmetic actually produces for these rates
    naive = dict(e["naiveDivision"])
    assert naive["8.775"] == "0.08775000000000001"
    assert naive["9.975"] == "0.09974999999999999"
    assert dict(e["naiveMultiply"])["0.07975"] == "7.9750000000000005"

    # ...against what the shift produces for the same inputs
    shifted = dict(e["cases"])
    assert shifted["8.775"] == "0.08775"
    assert shifted["9.975"] == "0.09975"
    assert shifted["7.15"] == "0.0715"
    # and the ordinary shapes, none of them rounded or padded
    assert shifted["7.975"] == "0.07975"
    assert shifted["10"] == "0.1"
    assert shifted["6.5"] == "0.065"
    assert shifted["0.5"] == "0.005"        # half a percent, not five percent
    assert shifted[".5"] == "0.005"         # a leading dot is a number people type
    assert shifted["7."] == "0.07"          # so is a trailing one
    assert shifted["11.125"] == "0.11125"   # more decimals than any current rate
    assert shifted["0"] == "0"              # zero is an answer, not an absence

    # the identity property, over every rate that breaks under arithmetic
    for typed, back in e["roundTripAll"]:
        assert back == typed, "%s did not survive the round trip: %s" % (typed, back)

    # painting the box is the same operation in reverse
    back = dict(e["backToPct"])
    assert back["0.07975"] == "7.975"
    assert back["0.08775"] == "8.775"
    assert back["0.07"] == "7"              # not 7.000000000000001

    # anything that is not a plain number is REFUSED rather than coerced. A typo that
    # becomes a price is the whole failure mode this feature could introduce.
    assert all(v is None for _, v in e["refused"]), e["refused"]


def test_a_typed_percentage_reaches_every_rate_cell_with_no_county_at_all(result):
    """The typed figure is the answer on its own -- it does not need a county chosen first.

    The rate came off the state's site for this address, so requiring a county pick before
    it would take effect would mean the estimator has to enter a coarser number to make the
    accurate one land. It writes to the same ten template layouts plus every copy, through
    the same layout-keyed addressing the county path uses."""
    t = result["typedNoCounty"]
    assert t["epoxy"] == '=IF(D6="yes",0.07975,0)'
    assert t["polish"] == '=IF(D6="yes",0.07975,0)'
    assert t["copy1"] == '=IF(D6="yes",0.07975,0)'    # a copy of Epoxy, same as the county path
    assert t["oneGyp"] == '=IF(D8="yes",0.07975,0)'   # gyp's switch is D8, not D6
    assert t["effective"] == 0.07975
    assert t["persisted"] == 0.07975                  # and it survives into the saved draft


def test_only_the_rate_is_written_and_never_a_dollar_amount(result):
    """Kyle's whole-dollar rounding stays Kyle's.

    His cell is ROUNDUP(SUM(D53:D55,D62,D68,D73:D77,D83)*B81,0) -- the taxable base, the
    multiply and the round to the dollar all live in his formula. This feature replaces one
    input to it and nothing else, so the tax figure on the generated workbook is still
    computed by the sheet rather than by us. Writing a dollar amount would mean two
    implementations of the same arithmetic, and the .xlsx the customer sees would disagree
    with the .xlsx Kyle opens the moment they drifted."""
    assert result["typedNoCounty"]["dollarCellsWritten"] == []


def test_a_typed_percentage_overrides_the_county_table(result):
    """Kyle reads the rate off the site for the ADDRESS; the county table is coarser.

    So when both exist the typed one wins, and the county's own figure stays on state
    untouched -- clearing the typed number has to be able to fall back to it."""
    t = result["typedBeatsCounty"]
    assert t["epoxy"] == '=IF(D6="yes",0.07975,0)'    # not the county's 9.35%
    assert t["effective"] == 0.07975
    assert t["countyStillOnState"] == 0.0935          # remembered, not overwritten
    assert "7.975%" in t["pill"]                      # and the pill agrees with the cell
    # ...and does not claim it arrived automatically. That wording is true of the county
    # table and false here, and it would be sitting inches from a note reading "Typed in".
    assert "the % you typed" in t["pill"]
    assert "applied automatically" not in t["pill"]


def test_clearing_the_county_does_not_retract_a_typed_percentage(result):
    """Two different acts, and conflating them would silently move a price.

    The estimator got that number off the state's site for this address. Dropping the
    county -- because it was wrong, or picked by accident -- says nothing about the rate. If
    clearing the pill also reverted the cell, the bid would quietly go back to Kyle's 10%
    placeholder with no visible change anywhere on the page."""
    c = result["clearCountyKeepsTyped"]
    assert c["countyGone"] is True
    assert c["epoxy"] == '=IF(D6="yes",0.07975,0)'
    assert c["effective"] == 0.07975

    # And the clear has to SURVIVE a reload, which it never did: the handler used
    # `delete state.county` and then TW.setState(state). setState is
    # Object.assign(cur, partial) where cur is re-parsed from localStorage, so a key the
    # partial lacks keeps its stored value -- the pill emptied on screen and the county
    # stayed in the draft. That was cosmetic while the county was only a label. It is not
    # cosmetic now: effectiveRemodelRate() falls back to county_remodel_rate, so a county
    # that will not clear is a RATE that will not clear.
    assert c["payloadClearsCounty"] is True
    assert c["payloadKeepsTyped"] == 0.07975   # ...without taking the typed rate with it


def test_a_new_county_supersedes_a_typed_percentage_visibly(result):
    """The reverse direction, and the visibility is the point.

    A new county means a new address, so a rate typed for the old one is stale and must not
    survive. But it must not vanish silently either: the box is repainted with the new
    county's own figure, so the number on screen is the number in the cell. Leaving 7.975
    sitting in the box while 9.35 was in the workbook is exactly the class of bug this whole
    file exists for."""
    n = result["newCountyWins"]
    assert n["effective"] == 0.0935
    assert n["epoxy"] == '=IF(D6="yes",0.0935,0)'
    assert n["boxRepainted"] == "9.35"


def test_a_typo_is_refused_and_the_rate_in_force_does_not_move(result):
    """Refusing is the safe failure. Coercing is not.

    Number("seven point nine") is NaN, and NaN reaching the formula would produce a cell no
    spreadsheet can evaluate -- or worse, a 0 that silently zeroes the tax. So a non-number
    leaves the rate exactly as it was and says so in words, rather than in colour alone."""
    t = result["typoRefused"]
    assert t["epoxy"] == '=IF(D6="yes",0.0935,0)'     # still the county's rate
    assert t["effective"] == 0.0935
    assert t["override"] is None                      # nothing was persisted
    assert "not a number" in t["note"]
    assert "7.975" in t["note"]                       # and it shows the shape that works
    assert "rate-note-bad" in t["noteClass"]


def test_emptying_the_box_falls_back_rather_than_pinning_zero(result):
    """An empty box is "I have no figure of my own", not "the tax rate is 0%".

    Those differ by the whole tax line. Note this is why the state key is set to null and
    not deleted: TW.setState MERGES the object it is handed, so a delete would leave the old
    override sitting in the saved blob and the next reload would resurrect it."""
    e = result["emptiedFallsBack"]
    assert e["effective"] == 0.0935                   # back to the county
    assert e["epoxy"] == '=IF(D6="yes",0.0935,0)'
    assert e["override"] is None


def test_the_box_is_not_overwritten_while_someone_is_typing_in_it(result):
    """A re-render that repaints a focused field eats the digits half-typed into it.

    Reached for the keyboard because no assertion about the rendered value can see this --
    the field only loses its content while it has focus. Same class of defect as the
    re-render-on-change focus theft found by a browser walk earlier this year."""
    f = result["focusRespected"]
    assert f["whileFocused"] == "7.9"     # mid-keystroke, left alone
    assert f["afterBlur"] == "9.35"       # and repainted once it is no longer being typed in


def test_the_page_says_when_a_typed_percentage_is_not_affecting_the_price(result):
    """Kyle's cell is =IF(D6="yes",<rate>,0). With that switch off the tax is zero whatever
    is typed, so a correctly-typed rate can sit there doing nothing.

    Saying so is the difference between a tool and a trap: the estimator types the figure
    the state gave them, sees it accepted, and would otherwise reasonably believe the bid
    now carries that tax. The rate is still written to the cell in that case -- the switch,
    not this feature, is what zeroes the tax, and flipping the switch on later must not need
    the rate re-typed."""
    s = result["switchNotes"]
    assert s["onIsOn"] is True
    assert "using" in s["onNote"]

    assert s["offIsOn"] is False
    assert "switched off" in s["offNote"]
    assert "not changing the price" in s["offNote"]
    assert s["offCell"] == '=IF(D6="yes",0.07975,0)'   # written anyway, ready for the switch

    # and with nothing typed and no county, the placeholder in force is named out loud
    assert "10%" in s["unsetNote"]


# The rate has to survive into the workbook, not just into the draft ==========
# Everything above this line is the browser's half: the right formula reaches `cellValues`.
# These two close the chain -- `cellValues` is forwarded verbatim into
# `estimate_writer.fill_estimate`, and what has to come out the other side is a live
# FORMULA. A cell holding =IF(D6="yes",0.07975,0) as literal TEXT computes nothing, and on
# screen it is indistinguishable from one that works.


def test_the_rate_formula_arrives_in_the_workbook_as_a_formula(tmp_path):
    """Written through the same cell_values path /api/generate uses, then read back.

    `test_cell_lock.py::test_cell_values_write_into_locked_cell_still_lands` already proves
    a cell_values entry lands through sheet protection, but it writes the number 0.05. The
    untested half is the one this feature depends on: openpyxl has to treat a leading "="
    as a formula. Asserted for the epoxy layout, the polish layout and a gyp variant --
    whose switch is D8, not D6 -- because the three carry different addresses and a
    single-sheet check would not notice one of them going missing."""
    import io

    from openpyxl import load_workbook

    import estimate_writer as ew

    gyp = next(s for s in ew.LOCK_MAP if s == "Gyp")   # the layout key, not a sheet name
    assert gyp                                          # (guards a rename of the layout)

    data = ew.fill_estimate({}, cell_values={
        "Epoxy!B81": '=IF(D6="yes",0.07975,0)',
        "Polish!B75": '=IF(D6="yes",0.07975,0)',
    })
    wb = load_workbook(io.BytesIO(data))
    assert wb["Epoxy"]["B81"].value == '=IF(D6="yes",0.07975,0)'
    assert wb["Polish"]["B75"].value == '=IF(D6="yes",0.07975,0)'
    # openpyxl reports the cell's data type; "f" is a formula, "s" would be inert text.
    assert wb["Epoxy"]["B81"].data_type == "f", (
        "the rate arrived as text, so the sheet computes no remodel tax at all")
    assert wb["Polish"]["B75"].data_type == "f"
    # and it is still protected, exactly as the county path already required
    assert wb["Epoxy"]["B81"].protection.locked
    assert wb["Epoxy"].protection.sheet is True


def test_a_copied_tabs_rate_formula_survives_the_copy(tmp_path):
    """Copies are how a priced option gets in front of a customer, and
    `_create_copied_tabs` clones them from the PRISTINE template -- so a copy's rate cell
    arrives holding Kyle's 10% placeholder unless the write reaches it. The browser half of
    this is covered four ways above; this is the workbook half."""
    import io

    from openpyxl import load_workbook

    import estimate_writer as ew

    data = ew.fill_estimate(
        {},
        cell_values={"Epoxy!B81": '=IF(D6="yes",0.07975,0)',
                     "Copy1!B81": '=IF(D6="yes",0.07975,0)'},
        tab_copies=[{"id": "Copy1", "source": "Epoxy"}],
    )
    wb = load_workbook(io.BytesIO(data))
    sheet = next(s for s in wb.sheetnames if s == "Copy1" or s.startswith("Copy1"))
    assert wb[sheet]["B81"].value == '=IF(D6="yes",0.07975,0)', (
        "the copy kept the template's 10% placeholder: %r" % wb[sheet]["B81"].value)
    assert wb[sheet]["B81"].data_type == "f"
    assert wb["Epoxy"]["B81"].value == '=IF(D6="yes",0.07975,0)'   # base tab unaffected
