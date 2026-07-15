"""Gyp (Gypsum Underlayment) work type — estimate-sheet fill + JS↔PY parity.

The 5 gyp variant sheets share one layout with column-E totals (vs Epoxy/Polish
column-D) and NO phase cell (gyp is mobilization-based). Intake seeds the three
SF buckets onto ALL five variants (G9/I9/K9) and the project info onto the base
sheet only (B2/B3/… — NOT =Epoxy! mirrors). This module pins the writer's gyp
behavior and asserts the frontend estimate-review.js constants match the Python
maps so the on-screen total bar reads the same cells the .xlsx carries.
"""
import io
import pathlib
import re

from openpyxl import load_workbook

import estimate_writer as ew

_JS = (pathlib.Path(__file__).resolve().parents[2]
       / "frontend" / "js" / "estimate-review.js").read_text(encoding="utf-8")


def _wb(data):
    return load_workbook(io.BytesIO(data))


# ── fill: SF buckets flow to all five gyp sheets; project info base-only ──
def test_fill_gyp_sf_lands_on_all_five_sheets():
    data = ew.fill_estimate({
        "work_type": "gyp", "project_name": "Branson Meadows",
        "gyp_soft_sf": 27825, "gyp_hard_sf": 11795, "gyp_corridor_sf": 5655,
    })
    wb = _wb(data)
    for name in ew.GYP_SHEETS:
        ws = wb[name]
        assert ws["G9"].value == 27825, name
        assert ws["I9"].value == 11795, name
        assert ws["K9"].value == 5655, name
    # Project info lands on the gyp BASE sheet only.
    assert wb[ew.GYP_SHEET]["B2"].value == "Branson Meadows"
    # A non-base gyp variant does NOT get the project name (base-only map).
    assert wb["Gyp (USG N12ULTRA)"]["B2"].value != "Branson Meadows"
    # The Epoxy SF input is untouched by a gyp job.
    assert wb["Epoxy"]["E20"].value != 27825


def test_fill_gyp_blank_sf_keeps_takeoff_formula():
    # No gyp SF supplied -> the G9/I9/K9 =P28/=Q28/=R28 takeoff formulas survive.
    wb = _wb(ew.fill_estimate({"work_type": "gyp"}))
    ws = wb[ew.GYP_SHEET]
    assert ws["G9"].value == "=P28"
    assert ws["I9"].value == "=Q28"
    assert ws["K9"].value == "=R28"


def test_fill_gyp_cell_values_write_lands():
    # A direct cell edit on a gyp sheet flows through the cell_values path.
    data = ew.fill_estimate({}, cell_values={f"{ew.GYP_SHEET}!B20": 12345})
    assert _wb(data)[ew.GYP_SHEET]["B20"].value == 12345


def test_read_totals_has_gyp_keys():
    out = ew.read_totals(ew.fill_estimate({"work_type": "gyp"}))
    assert "gyp" in out
    assert set(out["gyp"].keys()) == set(ew.GYP_TOTALS.keys())


# ── xlsx template pin (the gyp layout the maps depend on) ─────────────────
def test_gyp_template_cell_pin():
    tmpl = pathlib.Path(__file__).resolve().parent.parent / "templates" / "estimate_sheet_5.7.xlsx"
    wb = load_workbook(tmpl)
    ws = wb[ew.GYP_SHEET]
    assert ws["B18"].value == "=E87"                 # "Total Base Bid:" display
    assert ws["E87"].value == "=SUM(E69,E72:E76,E81,E84)"
    assert ws["G9"].value == "=P28"                  # SF input (soft)
    assert ws["B16"].value == 'N12 1/8"'             # system name


# ── JS ↔ PY parity ────────────────────────────────────────────────────────
def _js_str_list(name):
    m = re.search(rf"const {name} = \[(.*?)\];", _JS, re.S)
    assert m, f"{name} not found in estimate-review.js"
    return m.group(1)


def _js_obj_pairs(block):
    # key: "VALUE" pairs from a JS object-literal body.
    return dict(re.findall(r'(\w+)\s*:\s*"([A-Z0-9]+)"', block))


def test_gyp_base_and_sheets_parity():
    m = re.search(r"const GYP_BASE = '([^']+)';", _JS)
    assert m and m.group(1) == ew.GYP_SHEET
    sheets_src = _js_str_list("GYP_SHEETS")
    # GYP_SHEETS[0] is the GYP_BASE variable; the other 4 are string literals.
    for name in ew.GYP_SHEETS[1:]:
        assert f"'{name}'" in sheets_src, f"{name} missing from JS GYP_SHEETS"
    # Same count of entries (1 var + 4 literals = 5 commas-separated items).
    assert len([s for s in sheets_src.split(",") if s.strip()]) == len(ew.GYP_SHEETS)


def test_gyp_sf_cells_parity():
    m = re.search(r"const GYP_SF_CELLS = \{(.*?)\};", _JS, re.S)
    assert m, "GYP_SF_CELLS not found in estimate-review.js"
    assert _js_obj_pairs(m.group(1)) == dict(ew.GYP_SF_MAP)


def test_total_cells_gyp_parity_and_no_phase_key():
    m = re.search(r"Gyp:\s*\{(.*?)\}", _JS, re.S)
    assert m, "TOTAL_CELLS.Gyp not found in estimate-review.js"
    gyp = _js_obj_pairs(m.group(1))
    assert gyp == {
        "total": "E87", "psf": "E18", "material": "E41", "labor": "E52",
        "tooling": "E61", "sales_tax": "E79", "remodel": "E80",
    }
    # Gyp is mobilization-based — it must NOT carry a phase cell (unlike Epoxy/Polish).
    assert "phase" not in gyp
    # And these map to the writer's totals cells (same E-column layout).
    assert gyp["total"] == ew.GYP_TOTALS["lump_sum"]
    assert gyp["material"] == ew.GYP_TOTALS["material_total"]
    assert gyp["labor"] == ew.GYP_TOTALS["labor_install"]
    assert gyp["tooling"] == ew.GYP_TOTALS["tooling_total"]
