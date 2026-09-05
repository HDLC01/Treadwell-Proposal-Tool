"""A tax-exempt job must be tax-exempt on every sheet it is priced from.

Kyle, mid-estimate: he copied a tab to make a priced option on a TAX-EXEMPT job and the copy
charged 9.475% sales tax while its own Taxable box read "No".

WHAT THE WORKBOOK ACTUALLY DOES. The sales-tax rate cell is SHEET-relative on all eleven priced
sheets -- ``=IF($B$6="no",0,0.09475)`` on Epoxy/Polish/Seal/Seal (+Jnts)/Epoxy blank/Leveling and
``=IF($B$8="no",0,0.09475)`` on all five gyp variants. ``$B$6`` is column- and row-absolute but
sheet-relative, so **each sheet reads its own flag cell**. Seven of those flag cells are mirrors
(``=Epoxy!B6``, ``=Polish!B6``, ``='Gyp (USG 1-8")'!B8``); four are independent literals:

    Epoxy!B6              'Yes'
    Leveling!B6           'Yes'
    'Gyp (USG 1-8")'!B8   'Yes'
    'Gyp (FR)'!B8         'Yes'    <- does NOT mirror the gyp base, unlike the other three variants

The tool wrote the estimator's answer to exactly one of them, ``Epoxy!B6``. So the reported bug
(a copied tab) was the visible third of it: **every tax-exempt gypsum and Leveling bid has been
quoted carrying 9.475% it should not have, with no copy involved at all.**

WHY THIS FILE READS THE WORKBOOK RATHER THAN RESTATING IT. The addresses above are not a
convention -- Kyle's gyp block sits one row lower than the epoxy block (B6 there is "Miles Away"),
and 'Gyp (FR)' breaks the pattern the other four gyp variants follow. A test that hard-codes what
this file's author believed would have agreed with the bug. So every address, every mirror and
every rate formula below comes out of ``estimate_sheet_5.7.xlsx`` live, and the fix's own maps are
checked AGAINST it.

WHY DOLLARS, NOT STRUCTURE. Asserting "B6 says No" passes while the rate cell is separately wrong.
openpyxl cannot evaluate a workbook, and neither LibreOffice nor HyperFormula is available here --
so the chain is walked with the workbook's OWN formula text at every step: resolve the flag
through its mirrors, evaluate Kyle's own ``IF(...)`` to get the RATE, then confirm the tax dollar
cell still multiplies by that rate cell and that the TOTAL still sums it. Rate 0 -> that row is
$0 -> the total drops by exactly the tax that used to be in it. Nothing here re-implements the
pricing; ``pricing.py`` is deliberately not involved (it respects ``taxable`` but is ~1% off on
quartz and ~30% wrong on polish, so routing the printed figure through it would swap one wrong
number for another).

THE THREE SURFACES. A fix that makes two of them agree is this bug again in a new disguise -- the
mirror-formula version of it would have had the screen and the .docx saying tax-free while the
downloaded .xlsx charged 9.475%, because ``_coerce`` turns a quoted-sheet-name formula into text.
So: the .xlsx is asserted here through the real ``fill_estimate``; the on-screen chip/total and
the proposal figure both come off the live HyperFormula engine, and ``taxable-flag-harness.js``
asserts the same literal reached it, cell for cell, out of the shipped ``copyTab``.
"""
import io
import json
import pathlib
import re
import shutil
import subprocess
import warnings

import pytest

REPO = pathlib.Path(__file__).resolve().parents[2]
FRONTEND = REPO / "frontend"
HARNESS = pathlib.Path(__file__).resolve().parent / "js" / "taxable-flag-harness.js"
TEMPLATE = pathlib.Path(__file__).resolve().parents[1] / "templates" / "estimate_sheet_5.7.xlsx"

needs_node = pytest.mark.skipif(shutil.which("node") is None, reason="node is not installed")

# The sheets that carry a bid. Everything else in the workbook (Takeoff, Stnd Alts,
# Specs+Dwgs+Addn, validation, Unit Layouts) has no TAXES & FEES block at all, which the
# label walk below proves rather than assumes.
PRICED_SHEETS = ["Epoxy", "Polish", "Seal", "Seal (+Jnts)", "Epoxy blank", "Leveling",
                 'Gyp (USG 1-8")', "Gyp (USG N12ULTRA)", 'Gyp (USG N25 1-4")',
                 "Gyp (GWorx SC190)", "Gyp (FR)"]


@pytest.fixture(scope="module")
def wb():
    """Kyle's shipped template, formulas not values. Read-only; never saved."""
    openpyxl = pytest.importorskip("openpyxl")
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")     # the x14 dataValidation extension openpyxl drops
        return openpyxl.load_workbook(TEMPLATE, data_only=False)


@pytest.fixture(scope="module")
def result():
    if shutil.which("node") is None:
        pytest.skip("node is not installed")
    import estimate_writer as ew
    proc = subprocess.run(
        ["node", str(HARNESS), str(FRONTEND), json.dumps(ew.list_sheet_names())],
        capture_output=True, text=True, encoding="utf-8", timeout=120)
    assert proc.returncode == 0, (
        "the harness itself failed -- read this before assuming a product bug:\n" + proc.stderr)
    return json.loads(proc.stdout.strip().splitlines()[-1])


# ── reading the workbook's own formulas ──────────────────────────────────────

_MIRROR = re.compile(r"^=(?:'([^']+)'|([A-Za-z0-9 ()+\-]+))!\$?([A-Z]+)\$?(\d+)$")
_RATE = re.compile(r'^=IF\(\$?([A-Z]+)\$?(\d+)="([a-z]+)",([0-9.]+),([0-9.]+)\)$', re.I)


def _resolve_flag(wb, sheet, addr, _depth=0):
    """Follow a flag cell through its mirror chain to the cell that actually holds the answer.

    Returns (sheet, addr, value). Raises rather than guessing: an unrecognised formula in this
    block means the template changed shape and every address in the fix needs re-deriving."""
    assert _depth < 8, "mirror loop at %s!%s" % (sheet, addr)
    v = wb[sheet][addr].value
    if isinstance(v, str) and v.startswith("="):
        m = _MIRROR.match(v.strip())
        assert m, "%s!%s is a formula this walk does not understand: %r" % (sheet, addr, v)
        return _resolve_flag(wb, m.group(1) or m.group(2), m.group(3) + m.group(4), _depth + 1)
    return sheet, addr, v


def _rate_from(formula, flag_value):
    """Evaluate Kyle's own ``=IF($B$6="no",0,0.09475)`` for a given flag value.

    Deliberately strict: any other shape raises, because a silently-unparsed rate cell is how a
    green test would certify a bid that still charges tax."""
    m = _RATE.match(str(formula).strip())
    assert m, "the sales-tax rate cell is no longer an IF this walk understands: %r" % (formula,)
    _col, _row, needle, when_true, when_false = m.groups()
    hit = str(flag_value or "").strip().lower() == needle.strip().lower()
    return float(when_true if hit else when_false)


def _tax_row(wb, sheet):
    """(rate_addr, dollar_addr, flag_addr) for a sheet's Sales Tax line, found by its LABEL.

    By label, not by a typed address: the row differs per layout (Epoxy 80, Polish/Seal 74,
    Leveling 76, gyp 79) and the totals column differs too (D on epoxy/polish, E on gyp)."""
    ws = wb[sheet]
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value or "").strip().lower() != "sales tax":
            continue
        rate_addr = "B%d" % row
        formula = ws[rate_addr].value
        m = _RATE.match(str(formula).strip())
        assert m, "%s!%s is not the IF this walk understands: %r" % (sheet, rate_addr, formula)
        flag_addr = m.group(1) + m.group(2)
        dollar = next(("%s%d" % (c, row) for c in ("D", "E")
                       if str(ws["%s%d" % (c, row)].value or "").startswith("=ROUNDUP")), None)
        assert dollar, "%s has a Sales Tax rate with no dollar cell beside it" % sheet
        return rate_addr, dollar, flag_addr
    raise AssertionError("no 'Sales Tax' label in column A of %s" % sheet)


# ── 1. the workbook invariant that would have caught this ────────────────────


def test_every_priced_sheet_has_a_sales_tax_line_that_reads_a_flag_the_fix_writes(wb):
    """The walk whose absence let four sheets drift.

    For every sheet with a "Sales Tax" label, follow the rate formula's own reference through the
    mirror chain and assert it lands on a cell ``TAXABLE_FLAG_CELLS`` writes. That is the whole
    claim, stated over the shipped workbook rather than over a list somebody typed: add a sixth
    gyp variant, or make one of today's mirrors independent, and this fails until the map moves.

    Before 2026-09-05 the equivalent assertion existed only in test_intake_conditions.py and only
    looked at Polish -- the one sheet whose mirror was already correct."""
    import estimate_writer as ew

    written = {"%s!%s" % (s, a) for s, a in ew.TAXABLE_FLAG_CELLS.items()}
    assert written, "the fix writes nothing at all"
    seen = []
    for sheet in PRICED_SHEETS:
        rate_addr, _dollar, flag_addr = _tax_row(wb, sheet)
        src_sheet, src_addr, _v = _resolve_flag(wb, sheet, flag_addr)
        key = "%s!%s" % (src_sheet, src_addr)
        assert key in written, (
            "%s's sales tax reads %s!%s, which nothing writes -- a tax-exempt bid on that sheet "
            "keeps the template's 'Yes' and bills 9.475%%" % (sheet, src_sheet, src_addr))
        seen.append((sheet, rate_addr, key))
    # ...and every cell the fix writes is actually reached by some sheet's rate formula, so the
    # map cannot grow a target that changes nothing.
    assert {k for _s, _r, k in seen} == written, sorted(seen)


def test_the_sheets_with_no_bid_on_them_have_no_sales_tax_line(wb):
    """PRICED_SHEETS above is a list, and a list can go stale. This is what stops it: any other
    worksheet growing a Sales Tax row would be a priced sheet nobody added to the walk."""
    for name in wb.sheetnames:
        if name in PRICED_SHEETS:
            continue
        labels = {str(wb[name].cell(row=r, column=1).value or "").strip().lower()
                  for r in range(1, min(wb[name].max_row, 200) + 1)}
        assert "sales tax" not in labels, (
            "%s has a Sales Tax line and is not in PRICED_SHEETS" % name)


def test_the_gyp_block_sits_one_row_lower_and_gyp_fr_is_its_own_answer(wb):
    """Two facts that a generalised "write B6" would get wrong, in opposite directions.

    On a gyp layout B6 is *Miles Away* -- writing the tax answer there puts a word in a mileage
    cell and leaves the tax alone. And 'Gyp (FR)' does NOT mirror the gyp base the way the other
    three variants do, so a fix covering "the gyp sheets" as a group ships with it still broken."""
    gyp_base = 'Gyp (USG 1-8")'
    assert str(wb[gyp_base]["A6"].value).strip().lower().startswith("miles")
    assert str(wb[gyp_base]["A8"].value).strip().lower().startswith("taxable")
    assert str(wb["Epoxy"]["A6"].value).strip().lower().startswith("taxable")

    assert wb[gyp_base]["B8"].value == "Yes", "the gyp base's Taxable is a literal"
    assert wb["Gyp (FR)"]["B8"].value == "Yes", (
        "'Gyp (FR)'!B8 is no longer an independent literal -- re-derive TAXABLE_FLAG_CELLS")
    for variant in ("Gyp (USG N12ULTRA)", 'Gyp (USG N25 1-4")', "Gyp (GWorx SC190)"):
        assert wb[variant]["B8"].value == "='%s'!B8" % gyp_base, (
            "%s stopped mirroring the gyp base -- it now needs writing too" % variant)


def test_the_fixs_literal_map_matches_the_workbook_cell_for_cell(wb):
    """Every cell the fix writes really is an independent literal, and every mirror really is a
    mirror. A literal written into a mirror replaces a live reference and forks the two sheets
    apart for good -- the divergence found in Kyle's own filed workbooks, and the one PR #432
    refused to introduce for the remodel rate."""
    import estimate_writer as ew

    for sheet, addr in ew.TAXABLE_FLAG_CELLS.items():
        v = wb[sheet][addr].value
        assert isinstance(v, str) and not v.startswith("="), (
            "%s!%s is a formula (%r) -- writing a literal there forks a working mirror" %
            (sheet, addr, v))
        assert v.strip().lower() in ("yes", "no")
    mirrors = {"Polish": "B6", "Seal": "B6", "Seal (+Jnts)": "B6", "Epoxy blank": "B6",
               "Gyp (USG N12ULTRA)": "B8", 'Gyp (USG N25 1-4")': "B8", "Gyp (GWorx SC190)": "B8"}
    for sheet, addr in mirrors.items():
        assert sheet not in ew.TAXABLE_FLAG_CELLS
        assert str(wb[sheet][addr].value).startswith("="), (
            "%s!%s stopped being a mirror -- it now needs writing like the four literals" %
            (sheet, addr))


def test_remodel_tax_has_exactly_one_literal_and_it_is_epoxy(wb):
    """Kyle's twin defect, and the reason it ships in the same change rather than after it.

    Epoxy!D6 is the ONLY literal remodel toggle -- Leveling!D6, both gyp D8 and every other
    sheet's are ``=Epoxy!D6``. So the base tabs were always right and only a COPY of the Epoxy
    layout froze it, at the template's 'No'. That direction UNDERBIDS: it drops a remodel tax the
    estimator switched on, which is the error that costs Treadwell rather than the customer."""
    assert wb["Epoxy"]["D6"].value == "No"
    for sheet, addr in (("Polish", "D6"), ("Seal", "D6"), ("Seal (+Jnts)", "D6"),
                        ("Epoxy blank", "D6"), ("Leveling", "D6"),
                        ('Gyp (USG 1-8")', "D8"), ("Gyp (FR)", "D8"),
                        ("Gyp (USG N12ULTRA)", "D8")):
        v = str(wb[sheet][addr].value)
        assert v.startswith("="), "%s!%s is no longer a mirror: %r" % (sheet, addr, v)


# ── the flag block as a WHOLE, so this file cannot certify what it did not fix ──


def _flag_labels(wb, sheet):
    """Every Yes/No question in the sheet's A1:D10 block, as {label: addr}.

    "Yes/No" is decided by what the cell RESOLVES to, not by the label: the block also holds
    "< 70 miles?" pointing at a Google Earth hint and "Miles Away" holding a number, and neither
    is an answer this walk has anything to say about."""
    ws, out = wb[sheet], {}
    for row in range(1, 11):
        for label_col, value_col in (("A", "B"), ("C", "D")):
            label = str(ws["%s%d" % (label_col, row)].value or "").strip()
            if not label.endswith("?"):
                continue
            addr = "%s%d" % (value_col, row)
            if wb[sheet][addr].value is None:
                continue
            _s, _a, resolved = _resolve_flag(wb, sheet, addr)
            if str(resolved or "").strip().lower() not in ("yes", "no"):
                continue
            out[label.rstrip("?").strip().lower()] = addr
    return out


FIXED_FLAGS = {"taxable", "remodel tax"}
# Local? and Hard Bid? are frozen on copied tabs by the IDENTICAL mechanism and are NOT in this
# change: they drive markup tiers, the hard-bid discount, gyp soft costs and travel/lodging, none
# of them a flat percentage, and nobody has yet quantified how far off a real bid they put it.
# Raised as Issue 5, deliberately not fixed here.
KNOWN_UNFIXED_FLAGS = {"local", "hard bid"}


@pytest.mark.parametrize("sheet", PRICED_SHEETS)
def test_every_yes_no_flag_in_the_block_is_either_written_or_a_mirror(wb, sheet):
    """THE WALK IS DELIBERATELY WIDER THAN THE FIX, and that is the point of it.

    A walk scoped to "Taxable?" and "Remodel Tax?" would go green while Local? and Hard Bid? stay
    frozen by the same mechanism -- a vacuous invariant certifying the rest of the block as fine.
    So this looks at EVERY Yes/No question in A1:D10 and requires each one to be a mirror (it
    follows the master) or a cell the tool writes. The two that are neither are named out loud
    below and xfail against Issue 5, so a green run says "two known holes", not "no holes".

    Prevailing Wage is a third case and needs no entry: Epoxy!D5 is its only literal and the tool
    has always written it."""
    import estimate_writer as ew

    written = {"%s!%s" % (s, a) for s, a in ew.TAXABLE_FLAG_CELLS.items()}
    written |= {"Epoxy!D6", "Epoxy!D5", "Epoxy!B4", "Epoxy!B5", "Polish!B4", "Polish!B5"}
    unfixed = []
    for label, addr in _flag_labels(wb, sheet).items():
        src_sheet, src_addr, _v = _resolve_flag(wb, sheet, addr)
        if "%s!%s" % (src_sheet, src_addr) in written:
            continue
        unfixed.append((label, "%s!%s" % (src_sheet, src_addr)))
    stray = [u for u in unfixed if u[0] not in KNOWN_UNFIXED_FLAGS]
    assert not stray, "a Yes/No flag nothing writes and nobody has flagged: %r" % (stray,)
    if unfixed:
        pytest.xfail("Issue 5 -- Local?/Hard Bid? are frozen the same way and are not in this "
                     "change: %s carries %r" % (sheet, sorted(u[1] for u in unfixed)))


# ── 2. the .xlsx the estimator downloads ─────────────────────────────────────


def _fill(taxable, tab_copies=None, cell_values=None):
    import estimate_writer as ew
    from openpyxl import load_workbook
    cv = dict(cell_values or {})
    data = ew.fill_estimate({"taxable": taxable}, cell_values=cv, tab_copies=tab_copies)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        return load_workbook(io.BytesIO(data), data_only=False)


# Filling this workbook costs ~5s, so the parametrised cases below share four of them rather
# than generating one per sheet. One copy PER priced layout in a single file also matches what
# an estimator with several options actually downloads.
COPY_OF = {"Copy%d" % (i + 1): s for i, s in enumerate(PRICED_SHEETS)}


def _copy_flag_values(answer):
    """What the browser's fan-out puts in cell_values for a bid with one copy of every layout."""
    import estimate_writer as ew
    out = {}
    for cid, src in COPY_OF.items():
        if src in ew.TAXABLE_FLAG_CELLS:
            out["%s!%s" % (cid, ew.TAXABLE_FLAG_CELLS[src])] = answer
    return out


@pytest.fixture(scope="module")
def exempt_wb():
    return _fill("No")


@pytest.fixture(scope="module")
def taxable_wb():
    return _fill("Yes")


@pytest.fixture(scope="module")
def exempt_copies_wb():
    return _fill("No", tab_copies=[{"id": c, "source": s} for c, s in COPY_OF.items()],
                 cell_values=_copy_flag_values("No"))


@pytest.fixture(scope="module")
def taxable_copies_wb():
    return _fill("Yes", tab_copies=[{"id": c, "source": s} for c, s in COPY_OF.items()],
                 cell_values=_copy_flag_values("Yes"))


def _sales_tax_rate(out_wb, sheet, source_wb=None):
    """The RATE the generated workbook's own Sales Tax formula produces on that sheet.

    Walks: label -> rate formula -> the flag address it references -> through the mirror chain ->
    the value actually sitting there -> Kyle's IF, evaluated. Every step reads the file that was
    just generated, so a write that lands on the wrong cell shows up as a rate, not as a silence."""
    rate_addr, dollar_addr, flag_addr = _tax_row(out_wb, sheet)
    _s, _a, value = _resolve_flag(out_wb, sheet, flag_addr)
    rate = _rate_from(out_wb[sheet][rate_addr].value, value)
    # ...and the money still hangs off that rate cell. A rate of 0 is only worth anything if the
    # dollar row still multiplies by it and the TOTAL still sums the row.
    dollar = str(out_wb[sheet][dollar_addr].value)
    assert re.search(r"\*%s\b" % rate_addr, dollar), (
        "%s's tax dollar cell stopped multiplying by %s: %r" % (sheet, rate_addr, dollar))
    return rate


@pytest.mark.parametrize("sheet", PRICED_SHEETS)
def test_a_tax_exempt_job_charges_no_sales_tax_on_any_priced_sheet(exempt_wb, sheet):
    """The bug, stated as money, one case per priced layout.

    Rate 0 makes the row ``ROUNDUP(SUM(<base>)*0,0)`` = $0 whatever the base is, so the TOTAL is
    the same total with the sales-tax row at zero -- which is what "tax exempt" means. Before this
    change Leveling and the two literal gyp sheets came back 0.09475 here."""
    assert _sales_tax_rate(exempt_wb, sheet) == 0.0, (
        "%s still bills 9.475%% on a tax-exempt job" % sheet)


@pytest.mark.parametrize("sheet", PRICED_SHEETS)
def test_a_taxable_job_still_charges_sales_tax_on_every_priced_sheet(taxable_wb, sheet):
    """The other direction, and it is not a formality: the cheapest way to make the test above
    pass is to switch sales tax off, which would quietly under-bill every ordinary job."""
    assert _sales_tax_rate(taxable_wb, sheet) == 0.09475, (
        "%s stopped charging sales tax on a TAXABLE job" % sheet)


def test_the_totals_chain_from_the_sales_tax_row_is_intact(wb):
    """Zeroing the rate only reaches the customer if the TOTAL still sums the row it zeroes.

    Epoxy: D80 (sales tax) -> D82 (Total Taxes, =SUM(D80:D81)) -> D88 (TOTAL, includes D82). The
    equivalent chain is asserted on every layout, out of the workbook's own formula text, so
    "rate 0" above is a statement about the printed price and not just about a cell."""
    for sheet in PRICED_SHEETS:
        _rate_addr, dollar_addr, _flag = _tax_row(wb, sheet)
        col, row = re.match(r"([A-Z]+)(\d+)", dollar_addr).groups()
        subtotal = "%s%d" % (col, int(row) + 2)          # "Total Taxes", two rows down
        assert dollar_addr in str(wb[sheet][subtotal].value), (
            "%s's Total Taxes (%s) no longer sums its sales-tax row" % (sheet, subtotal))
        total = next((("%s%d" % (col, r)) for r in range(int(row) + 3, int(row) + 12)
                      if str(wb[sheet]["A%d" % r].value or "").strip().upper() == "TOTAL"), None)
        assert total, "%s has no TOTAL row under its taxes block" % sheet
        assert subtotal in str(wb[sheet][total].value), (
            "%s's TOTAL (%s) no longer includes its Total Taxes cell" % (sheet, total))


@pytest.mark.parametrize("copy_id", sorted(COPY_OF, key=lambda c: int(c[4:])))
def test_a_copied_tab_on_a_tax_exempt_job_charges_no_sales_tax(exempt_copies_wb, exempt_wb,
                                                               copy_id):
    """Kyle's actual report, one case per copyable source.

    A copy is the ordinary way to put a priced option in front of a customer, and
    ``_create_copied_tabs`` clones it from the PRISTINE template -- so a copy of Epoxy, Leveling,
    the gyp base or 'Gyp (FR)' arrives holding the template's literal 'Yes'. ``copyTab``'s edit
    replay deliberately skips A1:D10, and ``canonicalTarget`` redirects any edit there to the
    master, so neither the copy nor the estimator could reach the cell. It read "No" and billed
    9.475%.

    The copy's own cell_values entry is what the browser fan-out produces; the mirrors need none,
    because ``copy_worksheet`` keeps ``=Epoxy!B6`` pointing at Epoxy."""
    source = COPY_OF[copy_id]
    assert copy_id in exempt_copies_wb.sheetnames
    assert _sales_tax_rate(exempt_copies_wb, copy_id) == 0.0, (
        "a copy of %s bills 9.475%% on a tax-exempt job" % source)
    # ...and the source sheet is unaffected by the copy existing
    assert _sales_tax_rate(exempt_copies_wb, source) == 0.0
    assert _sales_tax_rate(exempt_wb, source) == 0.0


@pytest.mark.parametrize("copy_id", sorted(COPY_OF, key=lambda c: int(c[4:])))
def test_a_copied_tab_on_a_taxable_job_still_charges_sales_tax(taxable_copies_wb, copy_id):
    """The other direction on a copy: put the template's 'Yes' back and the rate has to return to
    9.475%. Without it the case above would pass on a "fix" that simply switched sales tax off for
    every copied tab -- which is the same size of error, pointed at Treadwell."""
    assert _sales_tax_rate(taxable_copies_wb, copy_id) == 0.09475, COPY_OF[copy_id]


def test_the_mirrors_are_still_mirrors_in_the_generated_file(wb, exempt_wb):
    """Written into the .xlsx and read back, because that is where a fork would be permanent.

    Seven of the eleven flag cells are live references. Replacing one with a literal decouples the
    sheets for good: from then on a tax-exempt job would be exempt on the sheet that was written
    and taxable on the one that used to follow it, with nothing on screen to show it."""
    out = exempt_wb
    for sheet, addr in (("Polish", "B6"), ("Seal", "B6"), ("Seal (+Jnts)", "B6"),
                        ("Epoxy blank", "B6"), ("Gyp (USG N12ULTRA)", "B8"),
                        ('Gyp (USG N25 1-4")', "B8"), ("Gyp (GWorx SC190)", "B8")):
        assert out[sheet][addr].value == wb[sheet][addr].value, (
            "%s!%s was forked into a literal" % (sheet, addr))
        assert out[sheet][addr].data_type == "f"


def test_the_answer_arrives_as_a_word_and_never_as_a_formula():
    """``_coerce``'s whitelist is why this fix stamps literals instead of writing mirrors.

    Handed ``='Gyp (USG 1-8")'!B8`` it returns an apostrophe-escaped TEXT literal: the whitelist
    regex reads the quoted sheet-name prefix as a call to an un-whitelisted function ``Gyp``.
    HyperFormula would still get the working formula, so the SCREEN and the .docx would say
    tax-free while the downloaded .xlsx charged 9.475% -- three artefacts, two answers. Pinned
    here so nobody reintroduces it, along with the two forms that DO survive."""
    import estimate_writer as ew

    for word in ("Yes", "No"):
        assert ew._coerce(word) == word
    assert ew._coerce("=Epoxy!B6") == "=Epoxy!B6"
    assert ew._coerce("=Epoxy!D6") == "=Epoxy!D6"
    mangled = ew._coerce("='Gyp (USG 1-8\")'!B8")
    assert not str(mangled).startswith("='Gyp"), (
        "_coerce now passes quoted-sheet-name formulas through -- mirrors became an option "
        "again, but check info_sheet_writer._flag before taking it")


# ── 3. the hand-off sheet accounting reads ───────────────────────────────────


def test_the_info_sheet_still_reads_the_answer_as_yes_or_no():
    """``info_sheet_writer._flag`` is the second reason this fix writes literals.

    It expects a word. A mirror string is non-empty and not truthy, so ``_yn`` would return 'N'
    and B66 would print **"Tax Exempt? Y" on every gypsum job** -- plus the request-a-certificate
    instruction to Foundation, which is verbatim the failure that function exists to prevent,
    sign-flipped. The fan-out's own output is fed in here, not a hand-typed 'No'."""
    import info_sheet_writer as isw

    gyp_base = 'Gyp (USG 1-8")'
    draft = lambda cv: {"owner_email": "kyle@wetreadwell.com",
                        "data": {"project_name": "Westport Commons", "work_type": "gyp",
                                 "base_tab_id": gyp_base, "cell_values": cv}}
    exempt = isw.build_prefill(draft({"Epoxy!B6": "No", "Leveling!B6": "No",
                                      gyp_base + "!B8": "No", "Gyp (FR)!B8": "No"}))
    assert exempt["B66"] == "Y", "a tax-exempt gyp job must print Tax Exempt = Y"
    taxed = isw.build_prefill(draft({"Epoxy!B6": "Yes", "Leveling!B6": "Yes",
                                     gyp_base + "!B8": "Yes", "Gyp (FR)!B8": "Yes"}))
    assert taxed["B66"] == "N", (
        "a TAXABLE gyp job would tell Foundation to chase an exemption certificate")


def test_a_mirror_formula_in_the_gyp_flag_cell_would_be_read_as_taxable():
    """The counterexample that makes the test above bite.

    This is what the rejected mirror-formula fix would have put in that cell. Asserted so the
    consequence is on the record: it does not raise, it does not look wrong, it just prints the
    opposite answer on the sheet accounting works from."""
    import info_sheet_writer as isw

    gyp_base = 'Gyp (USG 1-8")'
    d = {"owner_email": "kyle@wetreadwell.com",
         "data": {"work_type": "gyp", "base_tab_id": gyp_base,
                  "cell_values": {gyp_base + "!B8": "='%s'!B8" % gyp_base}}}
    assert isw.build_prefill(d)["B66"] == "Y", (
        "a mirror string no longer reads as 'not taxable' -- if _yn changed, re-check whether "
        "mirrors are safe again")


# ── 4. the intake, which is where a base tab's answer comes from ─────────────


@needs_node
def test_the_intake_writes_the_answer_to_all_four_literal_cells(wb):
    """``frontend/js/index.js`` wrote ``Epoxy!B6`` alone, which is the whole of the base-tab half
    of this bug. The list is checked against the workbook rather than against itself."""
    import estimate_writer as ew

    src = (FRONTEND / "js" / "index.js").read_text(encoding="utf-8")
    m = re.search(r'\{ key: "taxable".*?cells: \[(.*?)\]', src, re.S)
    assert m, "the taxable condition moved -- re-derive this test"
    cells = [c.strip().strip("'\"") for c in m.group(1).split(",")]
    assert cells[0] == "Epoxy!B6", "hydrateConditions reads cells[0] to paint the switch"
    assert set(cells) == {"%s!%s" % (s, a) for s, a in ew.TAXABLE_FLAG_CELLS.items()}
    for c in cells:
        sheet, addr = c.split("!")
        v = wb[sheet][addr].value
        assert isinstance(v, str) and not v.startswith("="), (
            "the intake writes %s, which is a formula in the workbook" % c)


# ── 5. the browser half, executed ────────────────────────────────────────────


@needs_node
def test_the_target_list_is_four_cells_for_taxable_and_one_for_remodel(result):
    """Four, and the fifth would be a forked mirror. One, because Epoxy!D6 is the remodel
    toggle's only literal. If a sixth gyp variant is ever added, these are the numbers that
    must move -- not a count going stale while an option line bills tax that was switched off."""
    t = result["baseTargets"]["taxable"]
    assert [tuple(x) for x in t] == [("Epoxy", "B6"), ("Leveling", "B6"),
                                     ('Gyp (USG 1-8")', "B8"), ("Gyp (FR)", "B8")]
    assert [tuple(x) for x in result["baseTargets"]["remodel"]] == [("Epoxy", "D6")]


@needs_node
def test_the_answer_reaches_every_literal_sheet_and_no_mirror(result):
    """One answer, four cells, and the same values pushed into the live engine -- so the screen,
    the downloaded .xlsx and the proposal's snapshotted total cannot disagree about it."""
    f = result["baseFanout"]
    assert f["written"] == {"Epoxy!B6": "No", "Leveling!B6": "No",
                            'Gyp (USG 1-8")!B8': "No", "Gyp (FR)!B8": "No"}
    assert f["hfCallCount"] == 4 and f["hfMatches"] is True
    assert result["taxableStaysTaxable"] == {"Epoxy!B6": "Yes", "Leveling!B6": "Yes",
                                             'Gyp (USG 1-8")!B8': "Yes", "Gyp (FR)!B8": "Yes"}


@needs_node
def test_an_untouched_draft_collects_nothing(result):
    """Every one of these cells already holds the template's own default, so writing them back
    would only grow a blob that is PUT whole on every save. Same rule the remodel-rate self-heal
    applies to itself."""
    assert result["untouched"] == {"changed": 0, "keys": []}


@needs_node
def test_copying_a_tab_on_a_tax_exempt_job_carries_the_answer_onto_the_copy(result):
    """THE REPORT, through the shipped ``copyTab``.

    Executed rather than read: the old code is internally consistent and simply never reaches the
    cell. ``taxInEngine`` is the value behind the on-screen chip and total -- 'Yes' there beside
    a box reading "No" is exactly what Kyle saw."""
    for source, c in result["copies"].items():
        gyp = source.startswith("Gyp")
        literal = source in ("Epoxy", "Leveling", 'Gyp (USG 1-8")', "Gyp (FR)")
        assert c["taxInEngine"] == ("No" if literal else
                                    ("='%s'!B8" % 'Gyp (USG 1-8")' if gyp else "=Epoxy!B6")), source
        assert c["taxWritten"] is literal, (
            "%s: a copy of a mirror layout must NOT be forked into a literal" % source)
        assert c["cacheAlive"] is True, "%s: the fan-out's refresh destroyed the copy" % source
        assert c["remodelRateApplied"] == 1, "%s: copyTab stopped stamping the remodel rate" % source
    # the remodel twin, on the one layout that freezes it — and it freezes OFF, which underbids
    assert result["copies"]["Epoxy"]["remodelCellValue"] == "Yes"
    assert result["copies"]["Epoxy"]["remodelInEngine"] == "Yes"
    assert result["copies"]["Polish"]["remodelWritten"] is False


@needs_node
def test_a_copy_of_a_copy_gets_it_too_and_so_does_a_copy_made_first(result):
    """Both orders, because they fail differently. Copying first means the answer has to find a
    tab that did not exist when it was given; a copy of a copy has to resolve its LAYOUT through
    the chain, since 'Copy1' is in no map."""
    assert result["copyChain"] == {"c1": "No", "c2": "No", "c1Engine": "No", "c2Engine": "No"}
    a = result["answerAfterCopy"]
    assert a["copy1"] == a["copy2"] == a["copy3"] == "No"
    assert a["copy3NotB6"] is True, (
        "the gyp copy's answer went into B6, which on a gyp layout is Miles Away")


@needs_node
def test_typing_the_answer_anywhere_reaches_every_cell_that_holds_it(result):
    """The keystroke, through the grid's real edit listener.

    Four places the estimator can type it -- the master, a mirror tab, a copy, a gyp tab (whose
    flag is B8 and whose canonical is the gyp base, not Epoxy) -- and all four have to produce
    the same four-cell answer. Typing into the copy is what Kyle tried; ``canonicalTarget`` sent
    it to the master, which already said "No", so nothing happened at all."""
    expect = {"Epoxy!B6": "No", "Leveling!B6": "No",
              'Gyp (USG 1-8")!B8': "No", "Gyp (FR)!B8": "No"}
    t = result["typed"]
    assert t["master"] == expect
    assert t["mirrorTab"] == expect
    assert t["gypTab"] == expect
    assert t["onACopy"] == dict(expect, **{"Copy1!B6": "No"})
    # Retyping the ORIGINAL answer has to reach every cell too. The single-cell path deletes the
    # key on a revert, which would leave the copies holding the answer just retracted.
    assert t["revert"] == {"epoxy": "Yes", "copy": "Yes"}
    # ...and in every one of the four the OPEN tab is redrawn from the engine. Three of them are
    # sitting on a sheet that is not written -- Polish, a copy, a mirroring gyp variant -- whose
    # own flag cell is a live reference and whose totals recompute from it. Refreshing only when
    # a written sheet happens to be open would leave the estimator looking at the old price
    # after typing the answer, which is a quieter version of the bug being fixed. The caches
    # survive it, for the reason remodel-rate case 10 exists: a copy has no server-side
    # worksheet, so a refetch is data loss rather than a round trip.
    assert t["gridRefreshed"] == {"master": ["Epoxy"], "mirrorTab": ["Polish"],
                                  "onACopy": ["Copy1"], "gypTab": ['Gyp (USG N25 1-4")']}
    assert t["cachesAlive"] is True
    # ...and the remodel toggle behaves identically one row across
    assert t["remodelOnACopy"] == {"epoxy": "Yes", "copy": "Yes", "copyEngine": "Yes"}


@needs_node
def test_an_ordinary_cell_edit_is_untouched_by_the_fan_out(result):
    """The fan-out fires for two addresses per layout and for nothing else -- including the cells
    sitting beside them in the same block. ``Gyp!B6`` is *Miles Away*, and a "B6 is the tax flag"
    rule would have turned a mileage into a four-cell tax answer."""
    assert result["ordinaryEdits"] == {"Epoxy!E20": "5000", "Epoxy!B4": "No", "Epoxy!B5": "Yes",
                                       "Epoxy!D5": "Yes", 'Gyp (USG 1-8")!B6': "12"}
    k = result["kinds"]
    assert k["epoxyB6"] == "taxable" and k["epoxyD6"] == "remodel"
    assert k["gypB8"] == "taxable" and k["gypD8"] == "remodel"
    assert k["gypB6"] is None, "Miles Away was mistaken for the Taxable flag"
    assert k["epoxyB4"] is None, "Local? is Issue 5, not this change"
    assert k["takeoffB6"] is None, "a sheet with no flag block must never match"
    assert k["copyOfGypB8"] == ["taxable", None], "a copy resolves through its LAYOUT"


@needs_node
def test_the_shared_project_info_block_is_not_forked_by_any_of_this(result):
    """The A1:D10 redirect exists to share project name / bid date / address across every tab, and
    it is right about those. This fix changes where two answers go, not the block they live in."""
    p = result["projectInfoStillShared"]
    assert p["b1"] == {"sheet": "Epoxy", "addr": "B1"}
    assert p["b2"] == {"sheet": "Epoxy", "addr": "B2"}
    assert p["b3"] == {"sheet": "Epoxy", "addr": "B3"}
    assert p["gypB2"] == {"sheet": 'Gyp (USG 1-8")', "addr": "B2"}
    assert p["typedB1"] == {"onMaster": "New Name", "forkedOntoTheCopy": False}


@needs_node
def test_a_structurally_edited_tab_is_written_at_its_real_address_or_not_at_all(result):
    """Rows inserted above the block move the flag; deleting its row removes it. Both go through
    ``txAddr``, and a deleted cell is SKIPPED rather than written at a stale address -- which
    would put a word into whatever moved up into row 6."""
    s = result["structural"]
    assert [tuple(x) for x in s["movedTargets"]] == [("Copy1", "B8")]
    assert [tuple(x) for x in s["movedRemodelTargets"]] == [("Copy1", "D8")]
    assert s["movedWritten"] == "No" and s["movedNotAtStale"] is True
    assert s["goneTargets"] == [] and s["goneWrittenAnywhere"] == []


@needs_node
def test_the_base_tabs_family_wins_when_two_old_answers_disagree(result):
    """``canonicalSheetFor`` splits project info by family, so there are two canonical stores:
    ``Epoxy!B6`` for the epoxy family and the gyp base's ``B8`` for gyp. After this fix they
    cannot disagree -- every write fans out to both -- but a draft made BEFORE it can, in exactly
    one shape: the intake wrote Epoxy!B6 and the estimator then typed the real answer into the
    gyp tab's own box, which is the workaround being handed out for this very bug.

    Reading Epoxy first would throw that keystroke away and put the tax back on. So the BASE
    TAB's family wins: the base tab is the sheet the bid is priced from."""
    t = result["twoStores"]
    assert t["gypJobAnswer"] == "No" and t["gypJobGyp"] == "No", (
        "a gyp job's own Taxable answer was overwritten by the intake's stale epoxy one")
    assert t["epoxyJobAnswer"] == "No" and t["epoxyJobEpoxy"] == "No"
    assert t["onlyGypReaches"] == "No", "an answer given only on the gyp tab must reach Epoxy too"
    assert t["blankAnswer"] is None
    assert t["emptyStringIsNoAnswer"] is None, (
        "a cleared box is 'no answer', which is what fill_estimate does with an empty value")


@needs_node
def test_a_remodel_toggle_an_older_build_already_forked_is_kept_in_step(result):
    """The gyp base's D8 is ``=Epoxy!D6``, but ``canonicalTarget`` has always routed a gyp tab's
    Remodel keystroke onto it -- so drafts exist carrying a literal there. Left alone it outranks
    the master and goes on charging a remodel tax the estimator switched off. We do not create
    that fork and we do not silently un-fork it; we keep it in step."""
    f = result["forkedRemodel"]
    assert f["keptInStep"] == "No" and f["master"] == "No"
    assert [tuple(x) for x in f["cleanTargets"]] == [("Epoxy", "D6")]
    assert f["cleanKeys"] == ["Epoxy!D6"], "a fork was created where there was none"


# ── 6. the two call sites this harness cannot reach ─────────────────────────
#
# Everything above executes. These two do not, and the reason is recorded rather than papered
# over: `init()` is a 120-line async function that fetches /api/sheets, /api/named-expressions
# and all sixteen worksheets before it gets here, and the autofill call site is an anonymous
# async click handler wrapped round a fetch. Executing either means stubbing the network, which
# would prove nothing about the ordering that actually matters.
#
# So these assert the ORDER, out of the shipped source: WHERE the call sits is the whole of what
# can go wrong with it, and both positions are load-bearing. Ordering, not presence -- a plain
# "is the string there" check would pass with the call in the wrong half of the function.


def test_the_self_heal_runs_after_the_copies_exist_and_before_the_first_paint():
    """`applyJobFlags` has to see the whole bid.

    Run before the copied tabs are rehydrated and it writes to base sheets only, leaving exactly
    the copy Kyle reported still frozen. Run after `showSheet` and the first paint shows the old
    answer. So it belongs between the cellValues replay and the opening tab, and the save has to
    come after the totals are re-rendered -- `/api/generate` fills the workbook from the STORED
    draft, so a correction that never reached the save still downloads the tax."""
    src = (FRONTEND / "js" / "estimate-review.js").read_text(encoding="utf-8")
    body = src[src.index("async function init()"):src.index("\nfunction renderTabs()")]
    replay = body.index("// Apply saved overrides")
    heal = body.index("applyJobFlags()")
    paint = body.index("showSheet(initialSheet)")
    save = body.index("if (_flagsHealed) persistTabState()")
    assert replay < heal < paint < save, (
        "the load-time self-heal moved: replay=%d heal=%d paint=%d save=%d"
        % (replay, heal, paint, save))
    assert body.index("tab_copies.filter") < heal, (
        "the self-heal now runs before the copied tabs are rehydrated, so it cannot reach one")


def test_the_autofill_fans_its_flags_out_before_the_page_re_renders():
    """The AI writes all seven flags to hardcoded `Epoxy!...` keys whatever the work type, so on
    a gyp or Leveling bid its Taxable answer lands on a sheet that bid is not priced from. The
    fan-out has to run inside the same apply, before the grid is redrawn off those values."""
    src = (FRONTEND / "js" / "estimate-review.js").read_text(encoding="utf-8")
    body = src[src.index('document.getElementById("autofill-btn")'):]
    body = body[:body.index("function snapshotLumpSumsToState()")]
    write = body.index("cellValues[k] = v;")
    fan = body.index("applyJobFlags()")
    redraw = body.index("await showSheet(activeSheet)")
    assert write < fan < redraw, (
        "the autofill fan-out moved: write=%d fan=%d redraw=%d" % (write, fan, redraw))


@needs_node
def test_an_in_flight_draft_heals_itself_on_reopen_and_only_once(result):
    """The shape Kyle's job is in right now: ``Epoxy!B6="No"``, a copy made, nothing else ever
    written. Reopening has to fix all of it -- and report that it changed something, because the
    correction has to be SAVED (``/api/generate`` fills the workbook from the stored draft, not
    from the page). Opening an already-correct draft must change nothing, or every page load
    writes a save."""
    s = result["selfHeal"]
    assert s["changed"] == 4
    assert s["changedOnSecondOpen"] == 0
    assert s["written"] == {"Epoxy!B6": "No", "Leveling!B6": "No", 'Gyp (USG 1-8")!B8': "No",
                            "Gyp (FR)!B8": "No", "Copy1!B6": "No"}
    assert s["copyEngine"] == "No"
