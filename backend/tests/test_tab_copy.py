"""Worksheet copy + rename (the "literally a separate sheet" feature).

Covers the estimate_writer side:
- fill_estimate(tab_copies=…) duplicates a worksheet with ALL its contents
  (template cells + formulas), and the copy's "<id>!<addr>" edits land.
- fill_estimate(tab_labels=…) retitles worksheets to their display labels AND
  rewrites every cross-sheet `=Epoxy!` formula reference so the downloaded .xlsx
  still calculates (no dangling refs to a renamed sheet).
- the _rewrite_formula_refs helper (quoting, escaping, non-matches).
"""
import io
import re

from openpyxl import load_workbook

import estimate_writer as ew


def _wb(data):
    return load_workbook(io.BytesIO(data))


def _all_sheet_refs(wb):
    """Every distinct `Sheet!`/'Sheet Name'! reference found in any formula."""
    refs = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    for m in ew._SHEET_REF_RE.finditer(v):
                        tok = m.group(1)
                        name = tok[1:-1].replace("''", "'") if tok.startswith("'") else tok
                        refs.add(name)
    return refs


# ── helper: cross-sheet formula reference rewrite ───────────────────────
def test_rewrite_formula_refs_quoting_and_escaping():
    f = "=Epoxy!B1+SUM(Epoxy!D70:D77)+Polish!E18"
    out = ew._rewrite_formula_refs(f, {"Epoxy": "Grooming Room"})
    # space → quoted; Polish untouched (not in the map)
    assert out == "='Grooming Room'!B1+SUM('Grooming Room'!D70:D77)+Polish!E18"


def test_rewrite_formula_refs_simple_name_not_quoted():
    out = ew._rewrite_formula_refs("=Epoxy!B1", {"Epoxy": "Lobby"})
    assert out == "=Lobby!B1"          # simple token → no quotes needed


def test_rewrite_formula_refs_leaves_substring_names_alone():
    # "MyEpoxy!" is a different whole token than "Epoxy" → must NOT be rewritten.
    out = ew._rewrite_formula_refs("=MyEpoxy!B1+Epoxy!B2", {"Epoxy": "X"})
    assert out == "=MyEpoxy!B1+X!B2"


# ── fill_estimate: true copy keeps all contents ─────────────────────────
def test_copy_preserves_template_contents():
    data = ew.fill_estimate({}, tab_copies=[{"id": "Copy1", "source": "Epoxy"}])
    wb = _wb(data)
    src, cp = wb["Epoxy"], wb["Copy1"]
    # A formula cell and a static label cell both come across verbatim.
    assert str(cp["D88"].value or "").startswith("=")
    assert cp["A88"].value == src["A88"].value      # static label preserved


# ── fill_estimate: tab_labels retitle + formula rewrite ─────────────────
def test_label_retitles_and_keeps_formulas_valid():
    # Rename the base Epoxy tab → its =Epoxy! mirrors on other tabs must follow.
    data = ew.fill_estimate({"project_name": "P"},
                            tab_labels={"Epoxy": "Grooming"})
    wb = _wb(data)
    assert "Grooming" in wb.sheetnames
    assert "Epoxy" not in wb.sheetnames
    # No formula anywhere still points at the old "Epoxy" sheet…
    assert "Epoxy" not in _all_sheet_refs(wb)
    # …and every sheet reference that remains resolves to a real worksheet.
    names = set(wb.sheetnames)
    assert _all_sheet_refs(wb) <= names


def test_label_on_copy_and_base_together():
    data = ew.fill_estimate(
        {"project_name": "P"},
        cell_values={"Copy1!E20": 5000},
        tab_copies=[{"id": "Copy1", "source": "Epoxy"}],
        tab_labels={"Epoxy": "Grooming", "Copy1": "Exam Room"},
    )
    wb = _wb(data)
    assert {"Grooming", "Exam Room"} <= set(wb.sheetnames)
    assert "Epoxy" not in wb.sheetnames and "Copy1" not in wb.sheetnames
    # The copy kept its own edit and its =Epoxy! mirror now points at "Grooming".
    assert wb["Exam Room"]["E20"].value == 5000
    refs = _all_sheet_refs(wb)
    assert "Epoxy" not in refs and "Copy1" not in refs
    assert refs <= set(wb.sheetnames)


def test_no_labels_leaves_titles_unchanged():
    data = ew.fill_estimate({"project_name": "P"}, tab_labels={})
    wb = _wb(data)
    assert "Epoxy" in wb.sheetnames and "Polish" in wb.sheetnames


def test_label_equal_to_id_is_noop():
    # A label identical to the id must not trigger a (pointless) rename/rewrite.
    data = ew.fill_estimate({"project_name": "P"}, tab_labels={"Epoxy": "Epoxy"})
    wb = _wb(data)
    assert "Epoxy" in wb.sheetnames


# ── hardening: quoting, collisions, temp-name safety, reorder ───────────
def test_emit_ref_quotes_cell_ref_and_reserved_labels():
    assert ew._emit_ref("A1") == "'A1'!"          # cell-address shaped → must quote
    assert ew._emit_ref("XFD1") == "'XFD1'!"
    assert ew._emit_ref("TRUE") == "'TRUE'!"
    assert ew._emit_ref("R") == "'R'!" and ew._emit_ref("C1") == "'C1'!"
    assert ew._emit_ref("Lobby") == "Lobby!"      # normal name → no quotes
    assert ew._emit_ref("Grooming Room") == "'Grooming Room'!"


def test_cell_ref_shaped_label_is_quoted_in_xlsx():
    # Renaming a tab to "A1" must produce ='A1'!B1 (quoted), never =A1!B1.
    data = ew.fill_estimate({"project_name": "P"}, tab_labels={"Epoxy": "A1"})
    wb = _wb(data)
    assert "A1" in wb.sheetnames
    pol = wb["Polish"]["B1"].value
    if isinstance(pol, str) and pol.startswith("="):
        assert "=A1!" not in pol            # never an unquoted cell-shaped ref
    assert "Epoxy" not in _all_sheet_refs(wb)
    assert _all_sheet_refs(wb) <= set(wb.sheetnames)


def test_label_colliding_with_kept_sheet_is_deduped():
    # Direct-API edge: rename Epoxy -> "Polish" (a kept sheet). Must NOT collide;
    # both sheets keep distinct titles and every ref resolves.
    data = ew.fill_estimate({"project_name": "P"}, tab_labels={"Epoxy": "Polish"})
    wb = _wb(data)
    assert len(wb.sheetnames) == len(set(wb.sheetnames))   # no duplicate titles
    assert "Polish" in wb.sheetnames                       # the real Polish kept its name
    assert _all_sheet_refs(wb) <= set(wb.sheetnames)       # nothing dangling


def test_temp_name_collision_with_copy_id_is_safe():
    # A copy whose id is literally the temp prefix must not crash the rename pass.
    data = ew.fill_estimate(
        {"project_name": "P"},
        tab_copies=[{"id": "__twtmp0__", "source": "Epoxy"}],
        tab_labels={"Epoxy": "Grooming", "__twtmp0__": "Exam"},
    )
    wb = _wb(data)
    assert {"Grooming", "Exam"} <= set(wb.sheetnames)
    assert len(wb.sheetnames) == len(set(wb.sheetnames))
    assert _all_sheet_refs(wb) <= set(wb.sheetnames)


def test_reorder_tabs():
    data = ew.fill_estimate({"project_name": "P"}, tab_order=["Polish", "Epoxy"])
    wb = _wb(data)
    assert wb.sheetnames[:2] == ["Polish", "Epoxy"]      # requested order first


def test_reorder_then_rename_together():
    data = ew.fill_estimate(
        {"project_name": "P"},
        tab_order=["Polish", "Epoxy"],
        tab_labels={"Epoxy": "Grooming"},
    )
    wb = _wb(data)
    assert wb.sheetnames[:2] == ["Polish", "Grooming"]   # reorder by id, then retitle
    assert _all_sheet_refs(wb) <= set(wb.sheetnames)
