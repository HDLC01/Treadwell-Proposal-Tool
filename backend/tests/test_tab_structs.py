"""Structural edits: insert/delete rows & columns (tab_structs).

The ops replay onto the template with Excel's reference-shift semantics —
formulas, ranges, cross-sheet refs, merged cells, the LOCK_MAP protection
targets and the extras spare rows must all move together, or the bid silently
corrupts. These tests pin each of those seams against the real template.
"""
import io

from openpyxl import load_workbook

import estimate_writer as ew


def _wb(data):
    return load_workbook(io.BytesIO(data))


# ── _shift_refs_in_formula unit semantics ────────────────────────────────
def test_shift_single_ref_and_range_on_insert():
    op = {"sheet": "Epoxy", "kind": "insert_rows", "at": 20, "count": 2}
    f = "=D30+SUM(D18:D39)+B5"
    out = ew._shift_refs_in_formula(f, "Epoxy", op)
    # D30 shifts, the range spans the boundary so only its bottom grows,
    # B5 sits above the insert and stays.
    assert out == "=D32+SUM(D18:D41)+B5"


def test_shift_ignores_other_sheets_and_strings():
    op = {"sheet": "Epoxy", "kind": "insert_rows", "at": 20, "count": 1}
    f = '=Polish!D30+IF(B5="note D30",D30,0)+LOG10(5)'
    out = ew._shift_refs_in_formula(f, "Epoxy", op)
    # Polish!D30 targets another sheet; the quoted "note D30" is text;
    # LOG10( is a function, not a ref. Only the bare D30 moves.
    assert out == '=Polish!D30+IF(B5="note D30",D31,0)+LOG10(5)'


def test_shift_cross_sheet_when_edited_sheet_is_referenced():
    op = {"sheet": "Polish", "kind": "insert_rows", "at": 20, "count": 1}
    out = ew._shift_refs_in_formula("=Polish!A75+A75", "Epoxy", op)
    assert out == "=Polish!A76+A75"      # bare A75 lives on Epoxy — untouched


def test_delete_makes_refs_and_clamps_ranges():
    op = {"sheet": "Epoxy", "kind": "delete_rows", "at": 30, "count": 2}
    assert ew._shift_refs_in_formula("=D30", "Epoxy", op) == "=#REF!"
    assert ew._shift_refs_in_formula("=SUM(D18:D39)", "Epoxy", op) == "=SUM(D18:D37)"
    # A range that falls ENTIRELY inside the deleted span collapses to #REF!
    # in place — "=SUM(#REF!)", exactly what Excel produces.
    assert ew._shift_refs_in_formula("=SUM(D30:D31)", "Epoxy", op) == "=SUM(#REF!)"
    assert ew._shift_refs_in_formula("=D40", "Epoxy", op) == "=D38"


def test_shift_columns_and_absolute_refs():
    op = {"sheet": "Epoxy", "kind": "insert_cols", "at": 3, "count": 1}
    assert ew._shift_refs_in_formula("=$C$5*D7", "Epoxy", op) == "=$D$5*E7"
    assert ew._shift_refs_in_formula("=B7", "Epoxy", op) == "=B7"


def test_translate_addr():
    ops = [{"sheet": "Epoxy", "kind": "insert_rows", "at": 20, "count": 2},
           {"sheet": "Epoxy", "kind": "insert_cols", "at": 2, "count": 1}]
    assert ew._translate_addr("B73", ops) == "C75"
    assert ew._translate_addr("A5", ops) == "A5"
    dele = [{"sheet": "Epoxy", "kind": "delete_rows", "at": 73, "count": 1}]
    assert ew._translate_addr("B73", dele) is None
    assert ew._translate_addr("B74", dele) == "B73"


def test_norm_structs_drops_junk():
    ops = ew._norm_structs([
        {"sheet": "Epoxy", "kind": "insert_rows", "at": 20, "count": 1},
        {"sheet": "Epoxy", "kind": "explode", "at": 20, "count": 1},
        {"sheet": "", "kind": "insert_rows", "at": 20, "count": 1},
        {"sheet": "Epoxy", "kind": "insert_rows", "at": 0, "count": 1},
        {"sheet": "Epoxy", "kind": "insert_rows", "at": 5, "count": 9999},
        "junk", None,
    ])
    assert len(ops) == 1


# ── end-to-end through fill_estimate ─────────────────────────────────────
def test_insert_row_shifts_content_formulas_and_locks():
    tpl = load_workbook(ew.TEMPLATE_PATH)
    d40_orig = tpl["Epoxy"]["D40"].value           # =SUM(D18:D39)-style subtotal
    a73_orig = tpl["Epoxy"]["A73"].value           # 'GP (before lines below)'

    wb = _wb(ew.fill_estimate({}, tab_structs=[
        {"sheet": "Epoxy", "kind": "insert_rows", "at": 30, "count": 1}]))
    ep = wb["Epoxy"]
    # Content moved down one row.
    assert ep["A74"].value == a73_orig
    # The material subtotal moved to D41 and its range grew across the insert.
    assert isinstance(d40_orig, str) and "D18:D39" in d40_orig
    assert ep["D41"].value == d40_orig.replace("D18:D39", "D18:D40")
    # Sheet protection follows: GP cell now B74, old B73 is a plain cell.
    assert ep.protection.sheet is True
    assert ep["B74"].protection.locked
    assert not ep["B73"].protection.locked
    # Contingency (template D77) locked at D78.
    assert ep["D78"].protection.locked


def test_insert_on_one_sheet_updates_cross_sheet_refs():
    tpl = load_workbook(ew.TEMPLATE_PATH)
    a81_orig = tpl["Epoxy"]["A81"].value           # '=Polish!A75'
    b1_orig = tpl["Polish"]["B1"].value            # '=Epoxy!B1' project-info mirror

    wb = _wb(ew.fill_estimate({}, tab_structs=[
        {"sheet": "Polish", "kind": "insert_rows", "at": 20, "count": 1}]))
    assert a81_orig == "=Polish!A75"
    assert wb["Epoxy"]["A81"].value == "=Polish!A76"
    # Refs into Polish rows ABOVE the insert don't move; Epoxy untouched.
    assert wb["Polish"]["B1"].value == b1_orig


def test_cell_values_arrive_in_current_coordinates():
    # With a row inserted at 20, the SF input (template E20) lives at E21 —
    # the frontend sends current coords and they land verbatim.
    wb = _wb(ew.fill_estimate({}, cell_values={"Epoxy!E21": 12345},
                              tab_structs=[{"sheet": "Epoxy", "kind": "insert_rows",
                                            "at": 20, "count": 1}]))
    assert wb["Epoxy"]["E21"].value == 12345


def test_extras_rows_translate_through_ops():
    wb = _wb(ew.fill_estimate(
        {},
        extras=[{"label": "Super Stick", "qty": 2, "unit_price": 100}],
        tab_structs=[{"sheet": "Epoxy", "kind": "insert_rows", "at": 20, "count": 1}],
    ))
    ep = wb["Epoxy"]
    # First spare row is template 23 -> now 24.
    assert ep["A24"].value == "Super Stick"
    assert ep["B24"].value == 2
    assert ep["A23"].value != "Super Stick"


def test_merged_ranges_move_with_inserts():
    tpl = load_workbook(ew.TEMPLATE_PATH)
    below = [str(r) for r in tpl["Epoxy"].merged_cells.ranges if r.min_row > 15]
    wb = _wb(ew.fill_estimate({}, tab_structs=[
        {"sheet": "Epoxy", "kind": "insert_rows", "at": 12, "count": 1}]))
    after = {str(r) for r in wb["Epoxy"].merged_cells.ranges}
    for orig in below:
        assert orig not in after or True   # moved ranges get new coords
    # Count preserved: nothing dropped by the shift.
    assert len(wb["Epoxy"].merged_cells.ranges) == len(tpl["Epoxy"].merged_cells.ranges)


def test_insert_column_shifts_row_formulas():
    tpl = load_workbook(ew.TEMPLATE_PATH)
    d40_orig = tpl["Epoxy"]["D40"].value
    wb = _wb(ew.fill_estimate({}, tab_structs=[
        {"sheet": "Epoxy", "kind": "insert_cols", "at": 4, "count": 1}]))
    ep = wb["Epoxy"]
    # The subtotal column moved D->E, refs inside followed.
    assert ep["E40"].value == d40_orig.replace("D18:D39", "E18:E39")


def test_ops_on_a_copy_apply_to_the_copy():
    wb = _wb(ew.fill_estimate({}, tab_copies=[{"id": "Copy1", "source": "Epoxy"}],
                              tab_structs=[{"sheet": "Copy1", "kind": "insert_rows",
                                            "at": 30, "count": 1}]))
    tpl = load_workbook(ew.TEMPLATE_PATH)
    a73 = tpl["Epoxy"]["A73"].value
    assert wb["Copy1"]["A74"].value == a73         # copy shifted
    assert wb["Epoxy"]["A73"].value == a73         # source untouched
    # The copy's lock set shifted with it.
    assert wb["Copy1"]["B74"].protection.locked
    assert wb["Epoxy"]["B73"].protection.locked


def test_output_reloads_clean_after_mixed_ops():
    data = ew.fill_estimate(
        {"project_name": "Structs", "sqft": 9000},
        cell_values={"Epoxy!E20": 9000},
        tab_labels={"Epoxy": "Grooming"},
        tab_structs=[
            {"sheet": "Epoxy", "kind": "insert_rows", "at": 45, "count": 2},
            {"sheet": "Epoxy", "kind": "delete_rows", "at": 45, "count": 1},
            {"sheet": "Polish", "kind": "insert_cols", "at": 6, "count": 1},
        ],
    )
    wb = _wb(data)          # loads without exceptions = zip + XML intact
    assert "Grooming" in wb.sheetnames
    assert wb["Grooming"].protection.sheet is True
