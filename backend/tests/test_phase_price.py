"""'Add for additional phase' price ($4,500 default) → proposal NOTES bullet.

Covers: the backend _notes_for backstop (literal $xxxx → phase price), the
_phase_price_or_default coercion, the /api/generate end-to-end docx bullet, the
template pin (the reproducibility artifact for the hand/zip-surgery template
edit) incl. the x14-dropdown canary, and the generated-xlsx cell flow-through.
"""
import io
import re
import zipfile
from pathlib import Path

from docx import Document
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import main
import estimate_writer as ew

client = TestClient(main.app)

_MC = "{http://schemas.openxmlformats.org/markup-compatibility/2006}"
_TEMPLATE = Path(__file__).resolve().parent.parent / "templates" / "estimate_sheet_5.7.xlsx"
_PHASE_RE = re.compile(r"Add \$([\d,]+(?:\.\d+)?) for each additional phase beyond the above stated schedule\.")


def _rendered(docx_bytes):
    d = Document(io.BytesIO(docx_bytes))
    out = []
    for p in d.element.xpath("//w:p"):
        if any(True for _ in p.iterancestors(f"{_MC}Fallback")):
            continue
        t = "".join(x.text or "" for x in p.xpath(".//w:t")).strip()
        if t:
            out.append(t)
    return "\n".join(out)


_VALS = {
    "job_name": "J", "city_state": "C", "bid_date_formatted": "6/19/26",
    "base_bid_formatted": "$58,523.00", "material_tax_formatted": "$2,639.00",
    "state_name": "Kansas", "total_formatted": "$36,763.00",
    "system_name": "MACRO", "texture": "OP", "epoxy_sf": "12,000",
    "scope_notes": "demo", "schedule_notes": "~5 days", "work_description": "w",
    "site_visit_date": "6/19", "disposal": "d",
    "site_visit_phrase": "per site visit on 6/19/26",
    "base_tax_phrase": "(material sales tax INCLUDED)", "exclusions": "std",
}


# ── _notes_for backstop ────────────────────────────────────────────────────
def test_notes_for_substitutes_phase_price_in_defaults():
    for wt in ("epoxy", "polish", "combo"):
        texts = [n["text"] for n in main._notes_for(wt, [], 6000)]
        blob = "\n".join(texts)
        assert "Add $6,000 for each additional phase beyond the above stated schedule." in blob
        assert "$xxxx" not in blob


def test_notes_for_defaults_to_4500_when_no_price():
    blob = "\n".join(n["text"] for n in main._notes_for("epoxy", [], None))
    assert "Add $4,500 for each additional phase beyond the above stated schedule." in blob
    assert "$xxxx" not in blob


def test_notes_for_substitutes_literal_xxxx_in_user_notes():
    user = ["Add $xxxx for each additional phase beyond the above stated schedule.", "Custom note."]
    texts = [n["text"] for n in main._notes_for("epoxy", user, 5200)]
    assert texts[0] == "Add $5,200 for each additional phase beyond the above stated schedule."
    assert texts[1] == "Custom note."


def test_notes_for_leaves_hand_typed_amount_untouched():
    # A numeric amount the estimator hand-set is NOT a literal $xxxx → backend keeps it.
    user = ["Add $9,999 for each additional phase beyond the above stated schedule."]
    texts = [n["text"] for n in main._notes_for("epoxy", user, 5200)]
    assert texts[0] == "Add $9,999 for each additional phase beyond the above stated schedule."


def test_notes_for_does_not_mutate_default_notes():
    before = list(main._DEFAULT_NOTES.get("epoxy") or [])
    main._notes_for("epoxy", [], 6000)
    assert main._DEFAULT_NOTES.get("epoxy") == before
    assert any("$xxxx" in ln for ln in before)  # the anchor is preserved in the source


# ── _phase_price_or_default coercion (never raises) ─────────────────────────
def test_phase_price_coercion():
    f = main._phase_price_or_default
    assert f(6000) == 6000.0
    assert f("6,000") == 6000.0
    assert f("$5,200") == 5200.0
    assert f("abc") == 4500.0
    assert f(-5) == 4500.0
    assert f(0) == 4500.0
    assert f(10_000_000) == 4500.0        # out of range
    assert f(None) == 4500.0
    assert f({"x": 1}) == 4500.0
    assert f([1, 2]) == 4500.0


def test_generate_survives_malformed_phase_price():
    body = {"work_type": "epoxy", "audience": "Direct",
            "values": {**_VALS, "phase_price": "abc"}, "cell_values": {}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text


# ── end-to-end docx bullet ──────────────────────────────────────────────────
def test_generate_docx_phase_bullet_from_values():
    body = {"work_type": "epoxy", "audience": "Direct",
            "values": {**_VALS, "phase_price": 6000}, "cell_values": {}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    m = _PHASE_RE.search(blob)
    assert m and m.group(1) == "6,000", blob
    assert "$xxxx" not in blob


def test_generate_docx_phase_bullet_default_when_absent():
    body = {"work_type": "epoxy", "audience": "Direct", "values": dict(_VALS), "cell_values": {}}
    r = client.post("/api/generate", json=body)
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    m = _PHASE_RE.search(blob)
    assert m and m.group(1) == "4,500", blob


# ── GC proposals: additional-phase clause is static template text, driven by the
#    cell ONLY when edited off the $4,500 default (else each GC template keeps its
#    native default — Resinous $5,000, Polish/Sealer $2,300). ──────────────────
_GC_PHASE_RE = re.compile(
    r"Add \$([\d,]+) for each additional required phase beyond above stated schedule\.")


def _gc_amount(work_type, phase_price):
    """Generate a GC proposal, return the phase amount rendered in its docx."""
    values = dict(_VALS)
    if phase_price is not None:
        values["phase_price"] = phase_price
    r = client.post("/api/generate", json={
        "work_type": work_type, "audience": "GC", "values": values, "cell_values": {}})
    assert r.status_code == 200, r.text
    blob = _rendered(client.get(r.json()["docx_download_url"]).content)
    m = _GC_PHASE_RE.search(blob)
    assert m, "GC phase clause not found in docx:\n" + blob
    return m.group(1)


def test_gc_epoxy_keeps_native_5000_when_unedited():
    # No phase_price at all, and the sentinel default 4,500, both mean "unedited"
    # → the Resinous template's native $5,000 stays.
    assert _gc_amount("epoxy", None) == "5,000"
    assert _gc_amount("epoxy", 4500) == "5,000"


def test_gc_epoxy_follows_cell_when_edited():
    assert _gc_amount("epoxy", 6000) == "6,000"
    assert _gc_amount("epoxy", "5,200") == "5,200"


def test_gc_polish_keeps_native_2300_when_unedited():
    assert _gc_amount("polish", None) == "2,300"
    assert _gc_amount("polish", 4500) == "2,300"


def test_gc_polish_follows_cell_when_edited():
    assert _gc_amount("polish", 6000) == "6,000"


def test_phase_price_override_amount_coercion():
    f = main._phase_price_override_amount
    assert f(None) is None
    assert f("") is None
    assert f("abc") is None
    assert f(4500) is None          # the cell default = "unedited"
    assert f("4,500") is None
    assert f(0) is None
    assert f(10_000_000) is None    # out of range
    assert f(6000) == "6,000"
    assert f("5,200") == "5,200"
    assert f("$5,200") == "5,200"


def test_apply_gc_phase_override_is_idempotent_and_safe():
    import io as _io
    from docx import Document as _Doc
    import proposal_writer as pw
    tmpl = pw.pick_template("epoxy", "GC")
    d = _Doc(str(tmpl))
    # blank amount is a no-op
    assert pw._apply_gc_phase_override(d, "") == 0
    # first apply rewrites the clause; re-applying the SAME value rewrites nothing
    first = pw._apply_gc_phase_override(d, "6,000")
    assert first >= 1
    assert pw._apply_gc_phase_override(d, "6,000") == 0
    # and the doc now reads $6,000, not the native $5,000
    buf = _io.BytesIO(); d.save(buf)
    blob = _rendered(buf.getvalue())
    assert "Add $6,000 for each additional required phase" in blob
    assert "Add $5,000 for each additional required phase" not in blob


# ── template pin (reproducibility artifact for the template edit) ───────────
def test_template_has_phase_cells():
    wb = load_workbook(_TEMPLATE, data_only=False)
    for sheet, label_addr, val_addr in (("Epoxy", "B91", "C91"), ("Polish", "B85", "C85")):
        ws = wb[sheet]
        assert ws[label_addr].value == "Add for additional phase", f"{sheet}!{label_addr}"
        assert ws[val_addr].value == 4500, f"{sheet}!{val_addr}"
        assert "$" in ws[val_addr].number_format, f"{sheet}!{val_addr} fmt"
        assert ws[val_addr].protection.locked is False, f"{sheet}!{val_addr} must stay editable"


def test_template_x14_dropdowns_preserved():
    # Canary: the zip-surgery template edit must NOT strip the x14 data
    # validations (an openpyxl round-trip would) — most grid dropdowns need them.
    z = zipfile.ZipFile(_TEMPLATE)
    for part in ("xl/worksheets/sheet1.xml", "xl/worksheets/sheet2.xml"):
        xml = z.read(part).decode("utf-8", "ignore")
        assert "<x14:dataValidation " in xml, f"{part} lost its x14 dropdowns"


# ── generated xlsx carries the cell + user override ─────────────────────────
def test_generated_xlsx_phase_cell_flow_through():
    wb = load_workbook(io.BytesIO(ew.fill_estimate({"project_name": "P"})), data_only=False)
    assert wb["Epoxy"]["C91"].value == 4500
    assert wb["Polish"]["C85"].value == 4500


def test_generated_xlsx_phase_cell_user_override():
    wb = load_workbook(
        io.BytesIO(ew.fill_estimate({"project_name": "P"}, cell_values={"Epoxy!C91": 5200})),
        data_only=False,
    )
    assert wb["Epoxy"]["C91"].value == 5200
