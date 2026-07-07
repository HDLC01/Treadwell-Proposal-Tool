"""
Estimate-sheet writer.

Takes a dict of form values from the frontend, opens Kyle's actual
`estimate sheet - 5.7.xlsx` template (in `templates/`), writes the
typed values into the right input cells, returns the filled workbook
as bytes ready for download.

Kyle's template is the source of truth — we never modify it on disk.
Every call clones it into memory first.

Cell map (input cells per tab) was discovered by inspecting the
workbook with openpyxl. The map is intentionally narrow — only the
cells Troy actually types into. Computed cells (formulas) are left
alone; Excel re-evaluates them when the file is opened.

The mapping is data-driven (one dict per tab) so adding new fields
later means adding rows to the dict, not rewriting logic.
"""
from __future__ import annotations

import io
import logging
import re
from pathlib import Path
from typing import Any, Dict, Mapping

log = logging.getLogger("proposal_tool.estimate_writer")

from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.workbook import Workbook


TEMPLATE_PATH = (
    Path(__file__).parent / "templates" / "estimate_sheet_5.7.xlsx"
)

# Hand-curated dropdowns the UI exposes on the canonical Epoxy sheet —
# the source xlsx leaves these as plain text cells, but Treadwell uses
# them as Yes/No / New-Reno toggles. Maps cell address → option list.
PROJECT_INFO_DROPDOWNS: Dict[str, list] = {
    "B4":  ["Yes", "No"],          # Local?
    "B5":  ["Yes", "No"],          # Hard Bid?
    "D5":  ["Yes", "No"],          # Prevailing Wage?
    "B6":  ["Yes", "No"],          # Taxable?
    "D6":  ["Yes", "No"],          # Remodel Tax (tax-exempt)?
    "B10": ["New", "Reno"],        # New construction or Renovation
}


def _parse_x14_data_validations(sheet_name: str) -> list[tuple[list[str], list]]:
    """Parse Excel-extension (x14 namespace) data validations that
    openpyxl drops with a warning. Returns a list of (cell_addresses,
    options) tuples.

    The xlsx zip contains a per-sheet XML file (word/worksheet1.xml or
    similar) with an <extLst> at the bottom holding x14 validations:

        <extLst>
          <ext uri="{CCE6A557-...}">
            <x14:dataValidations>
              <x14:dataValidation type="list">
                <x14:formula1><xm:f>"Yes,No"</xm:f></x14:formula1>
                <xm:sqref>B4 B5</xm:sqref>
              </x14:dataValidation>
              ...
    """
    import re as _re
    import zipfile
    from xml.etree import ElementTree as _ET
    from openpyxl import load_workbook as _load_workbook

    # Find the worksheet's relative path in the xlsx
    wb = _load_template(data_only=False)
    if sheet_name not in wb.sheetnames:
        return []
    sheet_index = wb.sheetnames.index(sheet_name)
    sheet_xml_name = f"xl/worksheets/sheet{sheet_index + 1}.xml"

    with zipfile.ZipFile(TEMPLATE_PATH, "r") as z:
        if sheet_xml_name not in z.namelist():
            return []
        xml = z.read(sheet_xml_name).decode("utf-8", errors="replace")

    # Pull out the x14:dataValidations block via regex (simpler than parsing
    # the whole 5MB xml tree). Then parse only that fragment with ET.
    m = _re.search(r"<x14:dataValidations\b.*?</x14:dataValidations>", xml, _re.DOTALL)
    if not m:
        return []

    NS = {
        "x14": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main",
        "xm":  "http://schemas.microsoft.com/office/excel/2006/main",
    }
    # Wrap in a root so we can parse. Include every namespace Excel might
    # reference inside the data-validation block (uid, revision IDs, etc.).
    fragment = (
        '<root '
        'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main" '
        'xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main" '
        'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision" '
        'xmlns:xr2="http://schemas.microsoft.com/office/spreadsheetml/2015/revision2" '
        'xmlns:xr3="http://schemas.microsoft.com/office/spreadsheetml/2016/revision3">'
        + m.group(0) + "</root>"
    )
    try:
        root = _ET.fromstring(fragment)
    except _ET.ParseError:
        return []

    out: list[tuple[list[str], list]] = []
    for dv in root.iter("{http://schemas.microsoft.com/office/spreadsheetml/2009/9/main}dataValidation"):
        if dv.get("type") != "list":
            continue
        f1 = dv.find(".//x14:formula1/xm:f", NS)
        sqref_el = dv.find("xm:sqref", NS)
        if f1 is None or f1.text is None or sqref_el is None or sqref_el.text is None:
            continue
        formula = f1.text.strip()
        # Inline list
        if formula.startswith('"') and formula.endswith('"'):
            opts = [s.strip() for s in formula.strip('"').split(",") if s.strip()]
        else:
            opts = _resolve_range_to_options(wb, wb[sheet_name], formula)
        if not opts:
            continue

        # Expand the sqref into individual cell addresses.
        # sqref is space-separated, each token is an addr or range
        addrs: list[str] = []
        for token in sqref_el.text.split():
            if ":" in token:
                # Range — enumerate cells inside
                try:
                    rng = wb[sheet_name][token]
                    if not isinstance(rng, tuple):
                        rng = ((rng,),)
                    elif rng and not isinstance(rng[0], tuple):
                        rng = (rng,)
                    for row in rng:
                        for cell in row:
                            addrs.append(cell.coordinate)
                except Exception:
                    pass
            else:
                addrs.append(token)
        if addrs:
            out.append((addrs, opts))
    return out


def _resolve_range_to_options(wb, current_ws, formula: str) -> list:
    """Resolve a data-validation range reference (like `$B$161:$B$165`
    or `'Stnd Alts'!$A$1:$A$10`) into a list of option strings.

    Option cells are frequently FORMULAS that mirror the real names stored
    elsewhere on the sheet (e.g. R184 holds `=A192`, whose value is
    "MACRO Flake Single Broadcast"). So we read each option cell's CACHED
    value from the data_only workbook — that's the resolved name Excel
    shows in its own dropdown. Reading the formula workbook instead surfaces
    the formula text ("=A192") as the option label — which was the bug where
    dropdowns showed "=A171, =A177, …" instead of system names.

    Falls back to the literal formula-workbook value for plain-text option
    cells that happen to have no cached value, so we never drop a real
    option. Returns [] if the formula can't be parsed."""
    import re as _re

    f = formula.lstrip("=").strip()
    # Optional sheet prefix — either quoted ('Sheet Name'!range) or bare
    # (validation!$A$1:$A$10).
    sheet_match = _re.match(r"^'([^']+)'!(.+)$", f) or _re.match(r"^([A-Za-z_][\w\s\(\)\.\-]*)!(.+)$", f)
    if sheet_match:
        sheet_name = sheet_match.group(1)
        cell_range = sheet_match.group(2)
    else:
        sheet_name = current_ws.title
        cell_range = f
    if sheet_name not in wb.sheetnames:
        return []

    # Strip $ signs
    cell_range = cell_range.replace("$", "")
    if ":" not in cell_range:
        cell_range = f"{cell_range}:{cell_range}"

    try:
        cells_f = wb[sheet_name][cell_range]                       # formulas
        cells_v = _load_template(data_only=True)[sheet_name][cell_range]  # cached values
    except Exception:
        return []

    def _grid(cells):
        # ws[range] may return a single Cell, a row tuple, or a tuple-of-rows.
        if not hasattr(cells, "__iter__"):
            return [[cells]]
        return [list(row) if hasattr(row, "__iter__") else [row] for row in cells]

    grid_f, grid_v = _grid(cells_f), _grid(cells_v)
    out: list = []
    for r_idx, row in enumerate(grid_v):
        for c_idx, cell_v in enumerate(row):
            v = cell_v.value
            # Prefer the cached (resolved) value. If it's missing or itself a
            # formula string, fall back to the literal formula-workbook value
            # so plain-text options (e.g. "None") aren't lost.
            if v is None or (isinstance(v, str) and v.startswith("=")):
                try:
                    v = grid_f[r_idx][c_idx].value
                except IndexError:
                    v = None
            if v is None:
                continue
            s = str(v).strip()
            if not s or s.startswith("="):
                continue
            if s not in out:
                out.append(s)
    return out


# ─── Workbook caches ───────────────────────────────────────────────────
# Loading the 797 KB template costs ~300ms per call. We read it twice per
# /api/sheet call (formulas + cached values), and the user may switch
# between 16 tabs back and forth. Cache both reads as module globals,
# invalidated by file mtime so a manual swap of the template still works.

_WB_CACHE: Dict[str, tuple[float, Any]] = {}
# Per-sheet JSON-response cache, keyed by (sheet_name, file_mtime).
# Building one Epoxy response takes ~80ms over 5K cells; caching brings
# repeat tab visits to ~1ms.
_SHEET_GRID_CACHE: Dict[tuple[str, float], Dict[str, Any]] = {}


def _load_template(*, data_only: bool):
    """Load + cache the template workbook. Re-loads if the file mtime
    on disk changed (so swapping a new template still works without
    restarting the server)."""
    key = f"data_only={data_only}"
    mtime = TEMPLATE_PATH.stat().st_mtime
    if key in _WB_CACHE:
        cached_mtime, wb = _WB_CACHE[key]
        if cached_mtime == mtime:
            return wb
    wb = load_workbook(
        TEMPLATE_PATH,
        keep_vba=False,
        data_only=data_only,
    )
    _WB_CACHE[key] = (mtime, wb)
    return wb


# ─── Cell maps per tab ─────────────────────────────────────────────────
# Each entry: form_field_name → (cell_coordinate, value_transformer?)
# Value transformer is optional; identity by default.

EPOXY_CELL_MAP: Dict[str, str] = {
    # Project-level metadata
    "project_name":      "B1",     # next to "Project" label in row 1
    "bid_date":          "B2",
    "address":           "B3",
    "city_state":        "C3",     # sits next to the address
    "local":             "B4",     # "Yes" / "No"
    "prevailing_wage":   "D5",     # "Yes" / "No"
    "taxable":           "B6",     # "Yes" / "No"
    "remodel_tax":       "D6",     # "Yes" / "No"
    "approx_start_date": "B7",
    "architect":         "B8",
    "drawings_dated":    "B9",
    "new_or_reno":       "B10",    # "New" / "Reno"

    # System 1
    "system_1_cost_per_sf": "C18",
    "system_1_sf":          "E20",

    # System 2 (optional add-on epoxy run)
    "system_2_sf":          "E24",

    # Cove
    "cove_1_lf":   "E34",
    "cove_2_lf":   "E37",

    # Discounts / margin
    "discount_overage":   "B41",

    # Labor — Crew 1
    "labor_crew_size":    "A47",
    "labor_days":         "B47",
    "labor_rate":         "C47",

    # Travel
    "travel_hours":       "B52",
    "travel_rate":        "C52",
    "labor_burden_pct":   "C55",

    # Tooling
    "moisture_test_cost":    "C58",
    "tooling_consumables":   "C59",
    "demo_tooling":          "C60",

    # Travel — lodging / food
    "travel_lodging_per_day":  "C65",
    "travel_food_per_day":     "C66",

    # Markup
    "superintendent_pto_pct":  "B75",
    "soft_costs_pct":          "B76",
    "contingency":             "D77",
    "bond":                    "B84",
}

POLISH_CELL_MAP: Dict[str, str] = {
    "local":            "B4",
    "patch_material_sf": "E18",
    "floor_material_lf": "E19",
    "densifier_cost":    "C20",
    "sealer_cost":       "C21",
    "grout_compound_cost": "C22",
    "dye_cost_1":        "C25",
    "apply_dye":         "E25",
    "dye_cost_2":        "C26",
    "joint_filler_qty":  "C29",
    "apply_joint_filler": "E29",
    "remove_existing_jf": "F29",
    "shipping":          "B32",
    "polish_labor_crew_size": "A37",
    "polish_labor_rate":      "C37",
    "mockup_crew":            "A40",
    "mockup_days":            "B40",
    "joint_filler_crew":      "A44",
    "polish_labor_burden_pct": "C47",
    "polish_demo_tooling":   "C52",
    "slurry_hardener_qty":   "C53",
    "plastic_per_lf":        "C54",
    "polish_travel_lodging": "C58",
    "polish_travel_food":    "C59",
    "polish_superintendent_pto_pct": "B69",
    "polish_soft_costs_pct":         "B70",
    "polish_contingency":            "D71",
    "polish_bond":                   "B78",
}

# Computed total cells we surface back to the frontend (read-only).
# Used by `read_totals` to read back cached totals from a saved workbook.
EPOXY_TOTALS: Dict[str, str] = {
    "material_total":  "D43",
    "labor_install":   "D53",
    "tooling_total":   "D62",
    "travel_total":    "D68",
    "subtotal":        "D70",
    "lump_sum":        "D88",   # final TOTAL after markup + taxes
    "price_per_sf":    "D16",
}

POLISH_TOTALS: Dict[str, str] = {
    "material_total":  "D33",
    "labor_total":     "D45",
    "tooling_total":   "D55",
    "travel_total":    "D61",
    "subtotal":        "D64",
    "lump_sum":        "D82",
    "price_per_sf":    "D15",
}


# ─── Rate / markup / tax cell-lock map ─────────────────────────────────
# Cells that get Excel sheet-protection LOCKED in the generated .xlsx so the
# GP %, Hard Bid Discount %, Superintendent, Soft Costs, Contingency,
# Sales Tax %, Kansas Remodel Tax %, and Bond inputs can't be fat-fingered
# when Troy/Kyle open the file. Every OTHER cell stays editable.
#
# Keyed by the worksheet's TEMPLATE LAYOUT (its stable base id / source), not
# its final display title — sheet titles change during fill (the alternate tab
# is renamed, copies get user ids, display labels are applied at the end), so we
# resolve each worksheet to its layout while ids are still stable and apply the
# protection at the very end (see _resolve_ws_layouts / _apply_cell_protection).
#
# Contingency is col D on most layouts but col E on the Gyp sheets — reflected
# below. No password is set, so the sheet can still be unprotected in Excel
# (Review ▸ Unprotect Sheet) if Troy/Kyle need to.
LOCK_MAP: Dict[str, list[str]] = {
    "Epoxy":        ["B73", "B74", "B75", "B76", "D77", "B80", "B81", "B84"],
    "Polish":       ["B67", "B68", "B69", "B70", "D71", "B74", "B75", "B78"],
    "Seal":         ["B67", "B68", "B69", "B70", "D71", "B74", "B75", "B78"],
    "Seal (+Jnts)": ["B67", "B68", "B69", "B70", "D71", "B74", "B75"],
    "Leveling":     ["B69", "B70", "B71", "B72", "D73", "B76", "B77", "B80"],
    "Epoxy blank":  ["B70", "B71", "B72", "B73", "D74", "B77", "B78", "B81"],
    "Gyp":          ["B72", "B74", "B75", "E76", "B79", "B80", "B83"],
}


# ─── Public API ────────────────────────────────────────────────────────
# Epoxy material section spare manual rows — the "=B*C" lines that fall
# inside D40's SUM(D18:D39) range (Super Stick / Floor Graphic live here).
# A = label, B = qty, C = unit price; D is the =B*C formula we leave intact.
EPOXY_EXTRA_ROWS: list[int] = [23, 27, 28, 32, 33, 39]

# ── Alternate (recommended) system tab ──────────────────────────────────
# The spare "Epoxy blank" tab is repurposed to show Kyle's recommended
# alternate. Its project-info + labor cells are FORMULA MIRRORS to `Epoxy!`
# (don't overwrite). We drive it with the engine's alternate material SUBTOTAL
# (pre-shipping) injected into a "MATERIAL - Extras" =B*C row + the alt SF, and
# let the tab's own formulas recompute its Total Base Bid (D85). We never
# hard-write D85 (openpyxl can't evaluate formulas — Excel recomputes on open).
ALT_TAB_NAME = "Epoxy blank"
ALT_TAB_RENAME = "Alternate System"
ALT_SF_CELL = "E20"
ALT_MATERIAL_ROW = 29          # first "MATERIAL - Extras" =B*C row on the blank tab


def fill_estimate(
    values: Mapping[str, Any],
    cell_values: Mapping[str, Any] | None = None,
    extras: list[Mapping[str, Any]] | None = None,
    alternate: Mapping[str, Any] | None = None,
    tab_copies: list[Mapping[str, Any]] | None = None,
    tab_labels: Mapping[str, Any] | None = None,
    tab_order: list | None = None,
) -> bytes:
    """Open the template, write input cells, return filled workbook bytes.

    Three ways to specify what to write:

    1. `values` — flat dict from named fields. Keys map to the
       EPOXY_CELL_MAP / POLISH_CELL_MAP lookups above. Kept for
       backward compatibility.

    2. `cell_values` — dict keyed by `"<SheetName>!<CellAddress>"`,
       e.g. `{"Epoxy!E20": 18000, "Polish!C20": 0.07}`. Bypasses the
       named-field maps entirely so the UI can write to ANY cell
       without us having to hand-curate every one.

    3. `extras` — custom material lines (label/qty/unit_price), written
       into the Epoxy spare "=B*C" rows so they roll into D40 Material
       Sub Total. Beyond the 6 native rows the overflow is lumped into a
       single "Misc materials" line so the .xlsx formulas stay intact.

    `tab_copies` — duplicated worksheets: each {id, source} clones the `source`
    worksheet into a new sheet titled `id` (created BEFORE the `cell_values`
    pass so "<id>!<addr>" writes land). copy_worksheet keeps intra-tab formulas
    self-referential and `=Epoxy!` project-info mirrors intact.

    `tab_labels` — {internal_id: display_label}. Worksheets keep a stable internal
    id while edited (so the hardcoded Epoxy!/Polish! cell maps stay valid); the
    display labels are applied to the worksheet TITLES at the very end, with all
    cross-sheet `=id!` formula references rewritten so the .xlsx still calculates.

    All are applied; `cell_values` wins on conflicts.
    """
    wb = load_workbook(TEMPLATE_PATH, keep_vba=False)

    # 1. Named-field writes (legacy / typed-form path)
    epoxy = wb["Epoxy"]
    for field, coord in EPOXY_CELL_MAP.items():
        if field in values and values[field] not in (None, ""):
            epoxy[coord] = _coerce(values[field])

    polish = wb["Polish"]
    for field, coord in POLISH_CELL_MAP.items():
        if field in values and values[field] not in (None, ""):
            polish[coord] = _coerce(values[field])

    # 1.5 Duplicated worksheets: clone each {id, source} BEFORE the cell_values
    # loop (which skips sheets not yet in wb.sheetnames), so the copy's
    # "<id>!<addr>" writes land on the freshly-created sheet.
    _create_copied_tabs(wb, tab_copies)

    # 1.6 Resolve each worksheet OBJECT to its rate-cell lock layout NOW, while
    # every sheet still carries its stable id (before the alternate-rename and
    # display-label passes change titles). Worksheet objects survive those
    # renames, so we hold these references and apply the actual protection at the
    # very end, just before saving.
    ws_layouts = _resolve_ws_layouts(wb, tab_copies)

    # 2. Direct-cell writes (verbatim cell-for-cell editor path)
    for sheet_addr, val in (cell_values or {}).items():
        if val in (None, ""):
            continue
        if "!" not in sheet_addr:
            continue
        sheet_name, addr = sheet_addr.split("!", 1)
        if sheet_name not in wb.sheetnames:
            continue
        addr = addr.strip()
        # Only write single-cell coordinates (A1 .. XFD1048576). Rejects ranges,
        # defined names, and malformed input before handing them to openpyxl.
        if not _CELL_SHAPE_RE.fullmatch(addr):
            log.warning("estimate_writer: skipping non-cell address %r", sheet_addr)
            continue
        try:
            wb[sheet_name][addr] = _coerce(val)
        except Exception as exc:  # noqa: BLE001 — log the skip instead of swallowing it
            log.warning("estimate_writer: failed to write %s: %s", sheet_addr, exc)

    # 3. Extra material lines -> spare "=B*C" rows on the Epoxy tab.
    _write_extra_materials(epoxy, extras)

    # 4. Alternate (recommended) system -> the spare "Epoxy blank" tab.
    if alternate:
        _write_alternate_tab(wb, alternate)

    # 5. Reorder worksheets (by stable id) to match the user's tab order, THEN
    #    apply display labels + rewrite cross-sheet refs, after every write has
    #    landed on the stable ids.
    _reorder_tabs(wb, tab_order)
    _apply_tab_labels(wb, tab_labels)

    # 6. Excel sheet-protection: lock the rate/markup/tax cells so they can't be
    #    fat-fingered in Excel. Done LAST — after every write, extra, alternate
    #    tab, copy, reorder and rename — using the ws→layout map captured while
    #    titles were still stable ids.
    _apply_cell_protection(ws_layouts)

    # Stream to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_alternate_tab(wb, alternate: Mapping[str, Any]) -> None:
    """Populate the spare 'Epoxy blank' tab as the recommended alternate.

    `alternate` = {"sf", "material_sub", "label"}. We set the SF input and drop
    the engine's alternate material subtotal into a "MATERIAL - Extras" =B*C row
    so it flows D37 -> D40 -> ... -> D85 (the tab recomputes its own Total Base
    Bid on open). The PROPOSAL uses the engine's alternate total as authoritative;
    this tab is the supporting worksheet.
    """
    if ALT_TAB_NAME not in wb.sheetnames:
        return

    def _f(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return 0.0

    ws = wb[ALT_TAB_NAME]
    sf = _f(alternate.get("sf"))
    if sf:
        ws[ALT_SF_CELL] = sf
    material_sub = _f(alternate.get("material_sub"))
    if material_sub:
        r = ALT_MATERIAL_ROW
        ws[f"A{r}"] = (alternate.get("label") or "Alternate system").strip()
        ws[f"B{r}"] = 1
        ws[f"C{r}"] = material_sub      # D{r} stays =B{r}*C{r} -> flows into D37/D40/D85
    # Rename so the tab reads as the alternate (safe: its formulas reference
    # `Epoxy!`, not its own title, and no other tab references it by name).
    try:
        ws.title = ALT_TAB_RENAME
    except Exception:  # noqa: BLE001 — duplicate name or odd char; keep original
        pass


def _create_copied_tabs(wb, tab_copies: list[Mapping[str, Any]] | None) -> None:
    """Duplicate worksheets the user copied in the editor (true "copy worksheet").

    Each entry is {id, source}: clone the `source` worksheet into a new sheet
    titled `id` (the stable internal name the copy's "<id>!<addr>" cell_values are
    keyed to). openpyxl `copy_worksheet` copies cell values + formula STRINGS
    verbatim, so the copy's intra-tab formulas (e.g. `D88=SUM(D70,D73:D77,D82,D85)`)
    self-refer to the new sheet (its own bid) and its `=Epoxy!B1` project-info cells
    keep mirroring the master. The copy's "<id>!<addr>" edits are written by the
    caller's loop afterwards. Ids are clamped to Excel's 31-char limit; collisions /
    unknown sources / blanks are skipped. Sources are resolved in order so a copy of
    a copy works if its source was created earlier in the list.
    """
    for c in (tab_copies or []):
        if not isinstance(c, dict):
            continue
        new_id = str(c.get("id") or "").strip()[:31]
        src = str(c.get("source") or "Epoxy").strip() or "Epoxy"
        if not new_id or new_id in wb.sheetnames or src not in wb.sheetnames:
            continue
        try:
            ws = wb.copy_worksheet(wb[src])
            ws.title = new_id
        except Exception:  # noqa: BLE001 — bad title / odd char; skip this copy
            pass


def _lock_layout_for(base_id: str) -> list[str] | None:
    """LOCK_MAP addresses for a stable base sheet id, or None when the layout
    has no locked cells. All Gyp variants share the one "Gyp" set (their rate
    rows sit at identical coordinates); every other layout matches by exact id."""
    addrs = LOCK_MAP.get(base_id)
    if addrs is None and base_id.lower().startswith("gyp"):
        addrs = LOCK_MAP["Gyp"]
    return addrs


def _resolve_ws_layouts(wb, tab_copies: list[Mapping[str, Any]] | None) -> list[tuple[Any, list[str]]]:
    """Map each worksheet OBJECT to its rate-cell lock addresses.

    Called while every sheet still carries its stable id (after _create_copied_tabs,
    before the alternate-rename and tab-label passes), so titles resolve reliably:
    copies follow their {id, source} chain back to a template layout, Gyp variants
    share the Gyp set, and sheets with no locked cells (Takeoff, validation, …) are
    left out entirely — they get no protection. The worksheet objects survive the
    later retitles, so the returned pairs stay valid until save."""
    src_by_id: dict[str, str] = {}
    for c in (tab_copies or []):
        if not isinstance(c, dict):
            continue
        cid = str(c.get("id") or "").strip()[:31]
        if cid:
            src_by_id[cid] = str(c.get("source") or "Epoxy").strip() or "Epoxy"

    out: list[tuple[Any, list[str]]] = []
    for ws in wb.worksheets:
        base, guard = ws.title, 0
        while base in src_by_id and guard < 20:      # copy-of-a-copy chains
            base = src_by_id[base]
            guard += 1
        addrs = _lock_layout_for(base)
        if addrs:
            out.append((ws, addrs))
    return out


def _apply_cell_protection(ws_layouts: list[tuple[Any, list[str]]]) -> None:
    """Enable Excel sheet-protection so only the rate/markup/tax cells are locked.

    Excel's default is every cell locked once protection turns on, so everything
    else must be explicitly marked editable — in TWO tiers, because iterating the
    full dimension window would MATERIALIZE ~76k empty cells per sheet (measured:
    5x generate latency, 3x file size). Instead:

    1. cells that exist get locked=False directly (their styling is untouched —
       protection is one facet of the style, not a replacement);
    2. virgin cells (never created, so nothing to stamp) inherit the workbook's
       "Normal" named style — flipping ITS locked flag to False keeps blank rows
       and columns typeable under protection without serializing a single extra
       cell. Same trick as unchecking Format Cells ▸ Protection ▸ Locked on the
       Normal style in Excel.

    Then the LOCK_MAP addresses are re-locked and protection enabled with NO
    password — Troy/Kyle can still Review ▸ Unprotect Sheet when they genuinely
    need to change a rate. Formatting stays allowed (cell/column/row) so
    column-width tweaks aren't blocked; structural edits (insert/delete rows)
    stay protected. Failures are logged and skipped — protection is a guard
    rail, never worth failing a generation over."""
    if not ws_layouts:
        return
    unlocked = Protection(locked=False)
    locked = Protection(locked=True)

    # Tier 2 first (workbook-wide, once): Normal style → unlocked, so virgin
    # cells on the protected sheets stay editable. Harmless on unprotected
    # sheets — the flag only matters when a sheet's protection is on.
    try:
        wb = ws_layouts[0][0].parent
        for style in wb._named_styles:
            if style.name == "Normal":
                style.protection = Protection(locked=False)
    except Exception as exc:  # noqa: BLE001 — guard rail, not worth failing generation
        log.warning("estimate_writer: Normal-style unlock skipped: %s", exc)

    for ws, addrs in ws_layouts:
        try:
            # Tier 1: only cells that already exist — list() because assigning
            # a style can grow the dict via style interning side effects.
            for cell in list(ws._cells.values()):
                cell.protection = unlocked
            for addr in addrs:
                ws[addr].protection = locked
            ws.protection.sheet = True
            ws.protection.formatCells = False
            ws.protection.formatColumns = False
            ws.protection.formatRows = False
            ws.protection.enable()
        except Exception as exc:  # noqa: BLE001 — guard rail, not worth failing generation
            log.warning("estimate_writer: cell protection skipped on %r: %s", ws.title, exc)


# Match a sheet reference token in a formula: optional single-quoted name (with
# '' escaping) OR a bare name, immediately followed by '!'. Used to rewrite
# cross-sheet references when a worksheet is retitled. The '#' in the negative
# lookbehind keeps error literals like #REF!/#NULL! from matching as a "sheet".
_SHEET_REF_RE = re.compile(r"(?<![A-Za-z0-9_.#])('(?:[^']|'')+'|[A-Za-z_][A-Za-z0-9_.]*)!")


_CELL_SHAPE_RE = re.compile(r"(?i)[A-Z]{1,3}[0-9]{1,7}")      # A1, XFD1048576
_R1C1_RE = re.compile(r"(?i)R\d*C\d*|[RC]")                   # R, C, R1C1, RC


def _needs_quoting(title: str) -> bool:
    """A sheet name needs single quotes in a formula unless it's a simple
    [A-Za-z_][A-Za-z0-9_]* token that is NOT shaped like a cell reference (A1),
    an R1C1 token, or a boolean literal. When unsure we quote — over-quoting a
    name is always valid in Excel, under-quoting (e.g. =A1!B1) breaks the file."""
    t = title or ""
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", t):
        return True
    if t.upper() in ("TRUE", "FALSE"):
        return True
    if _CELL_SHAPE_RE.fullmatch(t) or _R1C1_RE.fullmatch(t):
        return True
    return False


def _emit_ref(title: str) -> str:
    """Render `<title>!` for a formula, quoting/escaping when required."""
    if _needs_quoting(title):
        return "'" + title.replace("'", "''") + "'!"
    return title + "!"


def _rewrite_formula_refs(formula: str, rename: Mapping[str, str]) -> str:
    """Rewrite every `OldSheet!`/`'Old Sheet'!` reference in a formula to the
    sheet's new title (with correct quoting). Names not in `rename` are left as-is.
    """
    def repl(m: "re.Match[str]") -> str:
        tok = m.group(1)
        name = tok[1:-1].replace("''", "'") if tok.startswith("'") else tok
        new = rename.get(name)
        return _emit_ref(new) if new is not None else m.group(0)

    return _SHEET_REF_RE.sub(repl, formula)


_INVALID_TITLE_RE = re.compile(r"[\\/*?:\[\]]")


def _sanitize_title(label: Any) -> str:
    """Make a user label safe as an Excel worksheet title: strip illegal chars
    (\\ / * ? : [ ]), trim surrounding apostrophes, clamp to 31 chars."""
    t = _INVALID_TITLE_RE.sub(" ", str(label or "")).strip().strip("'").strip()
    return t[:31].strip()


def _apply_tab_labels(wb, tab_labels: Mapping[str, Any] | None) -> None:
    """Retitle worksheets to their display labels and rewrite cross-sheet formula
    references so the downloaded .xlsx still calculates.

    `tab_labels` maps {internal_id: display_label}. Worksheets are edited under
    their stable ids (Epoxy/Polish/Copy1/…) so the hardcoded cell maps and the
    `cell_values` writes all land; only here, at the very end, do we rename them.
    openpyxl does NOT rewrite `=Epoxy!B1` on `ws.title=`, so we resolve a
    DETERMINISTIC, unique final title per sheet (sanitized + de-duplicated, never
    letting openpyxl silently suffix a clamped/colliding title), rewrite every
    formula reference to those exact titles, then retitle through collision-proof
    temp names. This keeps formula refs and actual titles in lockstep even when a
    direct API caller sends colliding/odd labels.
    """
    # Requested, sanitized labels for sheets that exist and actually change.
    requested: dict[str, str] = {}
    for sid, raw in (tab_labels or {}).items():
        sid = str(sid)
        label = _sanitize_title(raw)
        if label and sid in wb.sheetnames and label != sid:
            requested[sid] = label
    if not requested:
        return

    # Resolve unique final titles, reserving the suffix room; names of sheets NOT
    # being renamed are off-limits so we never collide with a kept sheet.
    taken = {n.lower() for n in wb.sheetnames if n not in requested}
    final: dict[str, str] = {}
    for sid, label in requested.items():
        t, i = label, 2
        while t.lower() in taken:
            suffix = f" {i}"
            t = label[: 31 - len(suffix)] + suffix
            i += 1
        if t != sid:
            final[sid] = t
        taken.add(t.lower())
    if not final:
        return

    # 1. Rewrite all cross-sheet refs in cell formulas + defined names to the
    #    deterministic final titles. Only EXISTING cells — formulas can't live in
    #    virgin cells, and iter_rows() would materialize the full dimension window
    #    (~76k empty cells per sheet; measured as the main generate slowdown).
    for ws in wb.worksheets:
        for cell in list(ws._cells.values()):
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                new_v = _rewrite_formula_refs(v, final)
                if new_v != v:
                    cell.value = new_v
    try:
        for dn in wb.defined_names.values():
            if isinstance(dn.value, str) and "!" in dn.value:
                dn.value = _rewrite_formula_refs(dn.value, final)
    except Exception:  # noqa: BLE001 — defined-name shapes vary; non-fatal
        pass

    # 2. Retitle via temp names unique vs ALL current titles, then to the finals
    #    (each final is unique by construction, so openpyxl won't mangle it).
    seen = {n.lower() for n in wb.sheetnames}
    tmp, counter = {}, 0
    for sid in final:
        while f"__twtmp{counter}__".lower() in seen:
            counter += 1
        t = f"__twtmp{counter}__"
        seen.add(t.lower())
        counter += 1
        wb[sid].title = t
        tmp[t] = final[sid]
    for t, label in tmp.items():
        wb[t].title = label


def _reorder_tabs(wb, tab_order: list | None) -> None:
    """Reorder worksheets to match the user's drag-to-reorder order (by stable id;
    called BEFORE _apply_tab_labels while titles are still ids). Ids not present
    are skipped; sheets not listed keep their relative order at the end."""
    if not tab_order:
        return
    desired = [sid for sid in tab_order if isinstance(sid, str) and sid in wb.sheetnames]
    if not desired:
        return
    rest = [s for s in wb.sheetnames if s not in desired]
    order = desired + rest
    try:
        wb._sheets.sort(key=lambda ws: order.index(ws.title))
    except Exception:  # noqa: BLE001 — never fail generation over sheet order
        pass


def _write_extra_materials(epoxy, extras: list[Mapping[str, Any]] | None) -> None:
    """Write custom material lines into the Epoxy spare rows.

    Each spare row's D is a `=B*C` formula, so we only set A/B/C and let
    Excel compute the amount. Up to 6 lines map one-to-one; if there are
    more, the first 5 are itemized and the rest are summed into a single
    "Misc materials" line in the last spare row — keeping every formula
    below (D40 SUM, the bid chain) valid without inserting rows.
    """
    items = []
    for e in (extras or []):
        label = str(e.get("label") or "").strip()
        if not label:
            continue
        try:
            qty = float(e.get("qty") or 0)
            up = float(e.get("unit_price") or 0)
        except (TypeError, ValueError):
            continue
        amount = e.get("amount")
        try:
            amount = float(amount) if amount is not None else qty * up
        except (TypeError, ValueError):
            amount = qty * up
        if not amount:
            continue
        items.append({"label": label, "qty": qty, "up": up, "amount": amount})

    if not items:
        return

    rows = EPOXY_EXTRA_ROWS
    cap = len(rows)
    if len(items) <= cap:
        placed = [(r, it) for r, it in zip(rows, items)]
    else:
        # Itemize the first cap-1, lump the remainder into the last row.
        placed = [(r, it) for r, it in zip(rows[:cap - 1], items[:cap - 1])]
        overflow = items[cap - 1:]
        lump_amt = round(sum(it["amount"] for it in overflow), 2)
        placed.append((rows[cap - 1], {
            "label": f"Misc materials ({len(overflow)} lines)",
            "qty": 1, "up": lump_amt, "amount": lump_amt,
        }))

    for r, it in placed:
        epoxy[f"A{r}"] = it["label"]
        epoxy[f"B{r}"] = _coerce(it["qty"])
        epoxy[f"C{r}"] = _coerce(it["up"])
        # D{r} stays as the template's =B{r}*C{r} formula.


# ── Typed-formula validator (Google-Sheets-style live formulas) ─────────
# The estimator can TYPE a formula into any grid cell ("=C60+C59") and it
# must persist to the downloaded .xlsx as a real formula. Cell text arrives
# over the API though, so "=…" strings are only let through when they look
# like a plain spreadsheet calculation: a conservative charset (cell refs,
# ranges, numbers, operators, comparisons, text literals, quoted sheet
# refs) plus a function-name whitelist. Exfiltration/DDE shapes fail one or
# the other — =WEBSERVICE/=HYPERLINK/=IMPORTXML aren't whitelisted, =cmd|…
# has "|" outside the charset, external [Book1] refs have "[" — and keep
# the apostrophe-escape below, exactly as before.
# Bounded quantifiers everywhere + a hard input cap in _is_safe_formula:
# this regex runs on API-supplied text, and an unbounded name class made it
# polynomial (CodeQL ReDoS: a megabyte of "AAA…" costs O(n^2) backtracking).
# Excel function names max out well under 31 chars.
_FORMULA_MAX_LEN = 512
_FORMULA_FUNC_RE = re.compile(r"([A-Za-z_][A-Za-z0-9_.]{0,30})[ \t]{0,4}\(")
_FORMULA_CHARSET_RE = re.compile(r'^=[A-Za-z0-9_ .,:;!$"\'()+\-*/^%&<>=]+$')
_SAFE_FORMULA_FUNCS = {
    "SUM", "IF", "MIN", "MAX", "ROUND", "ROUNDUP", "ROUNDDOWN", "AVERAGE",
    "COUNT", "COUNTA", "COUNTIF", "COUNTIFS", "SUMIF", "SUMIFS", "SUMPRODUCT",
    "ABS", "AND", "OR", "NOT", "IFERROR", "VLOOKUP", "HLOOKUP", "INDEX",
    "MATCH", "CONCATENATE", "TEXT", "CEILING", "FLOOR", "MOD", "POWER",
    "SQRT", "PI", "AVERAGEIF", "PRODUCT", "TRUNC",
}


def _is_safe_formula(s: str) -> bool:
    if len(s) > _FORMULA_MAX_LEN:      # nobody hand-types a 512+ char formula
        return False
    if not _FORMULA_CHARSET_RE.fullmatch(s):
        return False
    return all(f.upper() in _SAFE_FORMULA_FUNCS for f in _FORMULA_FUNC_RE.findall(s))


def _coerce(v: Any) -> Any:
    """Cast strings to numbers where the user typed a number, leave
    everything else alone. Keeps the workbook's number-format intact."""
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return v
        try:
            if "." in s or "e" in s or "E" in s:
                return float(s)
            return int(s)
        except ValueError:
            # A genuine typed formula persists as a formula (Excel recomputes
            # it on open, same as the template's own formulas).
            if s.startswith("=") and _is_safe_formula(s):
                return s
            # Neutralize Excel formula/DDE injection: a string that starts with
            # a formula trigger (= + - @) or a tab/CR gets a leading apostrophe
            # so Excel treats the whole cell as literal text.
            if v[:1] in ("=", "+", "-", "@", "\t", "\r"):
                return "'" + v
            return v
    return v


# ─── Cell-for-cell sheet reader ───────────────────────────────────────
def _serialize_cell(cell, *, display_value, is_formula, fill_hex, font_color) -> Dict[str, Any]:
    """Build a COMPACT cell dict for the grid API.

    Only `addr`/`row`/`col` are always present; every other field is included
    only when it differs from its default. The frontend reads each field
    defensively (`cell.bold`, `cell.fmt || ""`, `c.formula ?? c.value`, …), so
    a missing field == its default. Omitting the ~12 mostly-default fields per
    cell cuts the raw payload and the client's JSON.parse time roughly in half
    (on top of gzip, which the transport layer adds).
    """
    out: Dict[str, Any] = {"addr": cell.coordinate, "row": cell.row, "col": cell.column}
    if display_value is not None:
        out["value"] = display_value
    if is_formula:
        out["isFormula"] = True
        out["formula"] = cell.value          # HyperFormula needs the formula text
    if fill_hex:
        out["fill"] = fill_hex
    if font_color:
        out["fontColor"] = font_color
    font = cell.font
    if font.bold:
        out["bold"] = True
    if font.italic:
        out["italic"] = True
    if font.underline:
        out["underline"] = True
    if font.size:
        out["fontSize"] = float(font.size)
    fmt = cell.number_format or ""
    if fmt and fmt != "General":
        out["fmt"] = fmt
    al = cell.alignment
    if al.horizontal:
        out["align"] = al.horizontal
    if al.vertical:
        out["valign"] = al.vertical
    if al.wrap_text:
        out["wrap"] = True
    borders = _cell_borders(cell)
    if borders:
        out["borders"] = borders
    return out


def _normalize_cell_value(v):
    """Make openpyxl's date/time objects render sensibly in the grid.

    openpyxl (data_only) hands back `datetime`/`date`/`time` for date-formatted
    cells. We want:
      - a real date  → a readable "M/D/YYYY" string (not an ISO datetime)
      - the Excel epoch (serial 0) → BLANK. A formula like `=Epoxy!B2` over an
        EMPTY B2 computes to 0, which a date-formatted cell stores/caches as
        midnight 1899 → openpyxl reads it back as time(0,0) → "00:00:00" (or an
        1899/1900 date). That's a non-value and should show blank.
      - a bare time-of-day → BLANK (no estimate field is a clock time; it's the
        same epoch-zero artifact).
    """
    import datetime as _dt
    if isinstance(v, _dt.datetime):
        return None if v.year <= 1900 else f"{v.month}/{v.day}/{v.year}"
    if isinstance(v, _dt.date):          # plain date (datetime already handled)
        return None if v.year <= 1900 else f"{v.month}/{v.day}/{v.year}"
    if isinstance(v, _dt.time):
        return None
    return v


def read_sheet_grid(sheet_name: str) -> Dict[str, Any]:
    """Return every used cell on `sheet_name` as a flat list.

    Used by `GET /api/sheet/{name}` to power the cell-for-cell UI. We
    walk every cell up to (max_row, max_column), capturing value,
    formula, fill colour, font weight, number format, and merge info.

    Loads the workbook TWICE:
      - `data_only=False` to get the formula text + styling
      - `data_only=True`  to get the cached value Excel computed on
        its last save (so `=D88` shows as $25,135 instead of "=D88").
    Both modes are needed because openpyxl chooses one or the other
    at load time and can't expose both for the same cell.

    Cells that have neither value nor distinguishing formatting are
    omitted (keeps the payload manageable for huge sheets like Takeoff).
    """
    mtime = TEMPLATE_PATH.stat().st_mtime
    cache_key = (sheet_name, mtime)
    if cache_key in _SHEET_GRID_CACHE:
        return _SHEET_GRID_CACHE[cache_key]

    wb = _load_template(data_only=False)
    if sheet_name not in wb.sheetnames:
        raise KeyError(sheet_name)
    ws = wb[sheet_name]

    # Second workbook just for the cached values
    wb_vals = _load_template(data_only=True)
    ws_vals = wb_vals[sheet_name]

    # Merged ranges — structured for the frontend:
    #   { anchor: 'A1', minRow, maxRow, minCol, maxCol, rowSpan, colSpan }
    merged = []
    merged_inner: set[str] = set()
    for mr in ws.merged_cells.ranges:
        anchor = mr.coord.split(":")[0]
        merged.append({
            "anchor":  anchor,
            "range":   mr.coord,
            "minRow":  mr.min_row,
            "maxRow":  mr.max_row,
            "minCol":  mr.min_col,
            "maxCol":  mr.max_col,
            "rowSpan": mr.max_row - mr.min_row + 1,
            "colSpan": mr.max_col - mr.min_col + 1,
        })
        for row in ws[mr.coord]:
            for cell in row:
                if cell.coordinate != anchor:
                    merged_inner.add(cell.coordinate)

    cells: list[Dict[str, Any]] = []
    max_row, max_col = ws.max_row, ws.max_column

    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.coordinate in merged_inner:
                continue

            value = cell.value
            fill_hex = _fill_hex(cell)
            font_color = _font_color(cell)
            is_formula = isinstance(value, str) and value.startswith("=")

            # Skip truly empty cells (no value, no fill, not a merge anchor)
            if value is None and not fill_hex and not cell.font.bold:
                continue

            # For formula cells, pull the cached computed value Excel left
            # in the file — that's what users see in Excel itself.
            # We deliberately DO NOT expose the formula text over the
            # API; formulas live server-side only. The user sees the
            # computed result, the backend preserves the formula verbatim.
            if is_formula:
                cached = ws_vals[cell.coordinate].value
                display_value = (
                    cached if cached is not None and not (
                        isinstance(cached, str) and cached.startswith("=")
                    )
                    else None
                )
            else:
                display_value = value

            display_value = _normalize_cell_value(display_value)
            cells.append(_serialize_cell(
                cell, display_value=display_value, is_formula=is_formula,
                fill_hex=fill_hex, font_color=font_color,
            ))

    # Column widths (in Excel's char-width units; multiply by ~7px for px approx)
    col_widths: Dict[str, float] = {}
    for letter, dim in ws.column_dimensions.items():
        if dim.width:
            col_widths[letter] = dim.width

    # Row heights (Excel points; multiply by ~1.33 for px)
    row_heights: Dict[int, float] = {}
    for idx, dim in ws.row_dimensions.items():
        if dim.height:
            row_heights[int(idx)] = dim.height

    # Cells with data validations (e.g. dropdowns).
    # Two sources combine here:
    #   1. Data validations defined in the xlsx (inline lists + range refs)
    #   2. Hand-added project-info dropdowns Treadwell needs (Yes/No, New/Reno)
    #      — Kyle's xlsx leaves these as plain text cells but the tool
    #      treats them as dropdowns for usability.
    dropdowns: Dict[str, list[str]] = {}

    # (1) xlsx-defined validations
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list" or not dv.formula1:
            continue
        formula = dv.formula1.strip()
        opts: list[str] = []
        # Inline list like '"Yes,No"'
        if formula.startswith('"') and formula.endswith('"'):
            opts = [s.strip() for s in formula.strip('"').split(",")]
        else:
            # Range reference, e.g. '$B$161:$B$165' or "'Stnd Alts'!$A$1:$A$10"
            # Strip leading '=' and resolve via openpyxl
            opts = _resolve_range_to_options(wb, ws, formula)
        if not opts:
            continue
        for cell_range in dv.sqref.ranges:
            cells_in_range = ws[cell_range.coord]
            # ws[range] returns either a single Cell, a row tuple,
            # or a tuple-of-rows depending on the range shape. Normalize.
            if not isinstance(cells_in_range, tuple):
                cells_in_range = ((cells_in_range,),)
            elif cells_in_range and not isinstance(cells_in_range[0], tuple):
                cells_in_range = (cells_in_range,)
            for row in cells_in_range:
                for cell in row:
                    dropdowns[cell.coordinate] = opts

    # (1b) Excel "extension" data validations (x14 namespace, Excel 2010+).
    # openpyxl strips these and logs a warning — we parse them ourselves
    # from the raw XML. These are where most of Kyle's dropdowns live.
    try:
        x14_dvs = _parse_x14_data_validations(sheet_name)
        for cells_addrs, opts in x14_dvs:
            for addr in cells_addrs:
                if addr not in dropdowns:
                    dropdowns[addr] = opts
    except Exception as exc:
        # Non-fatal — log and continue with the regular dropdowns
        import logging
        logging.getLogger("proposal_tool.estimate_writer").warning(
            "x14 data validation parse failed for %s: %s", sheet_name, exc
        )

    # (2) Project-info hard-coded dropdowns — appear on every sheet so
    # the user can toggle them from any tab. Limited to the canonical
    # source sheet (Epoxy) — the mirror cells on other tabs are formula
    # references and editing them is canonicalised by the frontend.
    if sheet_name == "Epoxy":
        for addr, options in PROJECT_INFO_DROPDOWNS.items():
            dropdowns.setdefault(addr, options)

    # Border-symmetry pass: if cell A has `right` defined and cell to
    # its right (B) has no `left`, mirror A.right → B.left (and same
    # for top/bottom). Excel often defines borders on only one side of
    # a shared wall, but the user wants the line visible whichever way
    # they look at it. Without this, adjacent un-defined cells appear
    # to "lose" the wall because they have no override on that side and
    # fall back to the default thin gridline.
    by_addr: Dict[str, Dict[str, Any]] = {c["addr"]: c for c in cells}
    OPPOSITES = {"right": "left", "left": "right", "top": "bottom", "bottom": "top"}
    NEIGHBOR_OFFSET = {
        "right":  (0, 1),   # right of A is left of (A.row, A.col+1)
        "left":   (0, -1),
        "bottom": (1, 0),
        "top":    (-1, 0),
    }
    from openpyxl.utils import get_column_letter
    for c in cells:
        b = c.get("borders")
        if not b:
            continue
        for side, info in list(b.items()):
            if not info or not info.get("style"):
                continue
            dr, dc = NEIGHBOR_OFFSET[side]
            nr = c["row"] + dr
            nc = c["col"] + dc
            if nr < 1 or nc < 1:
                continue
            naddr = get_column_letter(nc) + str(nr)
            neighbor = by_addr.get(naddr)
            if not neighbor:
                continue
            opp = OPPOSITES[side]
            nb = neighbor.get("borders") or {}
            if not nb.get(opp):
                nb[opp] = info
                neighbor["borders"] = nb

    result = {
        "sheet":      sheet_name,
        "max_row":    max_row,
        "max_col":    max_col,
        "cells":      cells,
        "merged":     merged,
        "col_widths": col_widths,
        "row_heights": row_heights,
        "dropdowns":  dropdowns,
    }
    _SHEET_GRID_CACHE[cache_key] = result
    return result


def read_named_expressions() -> list[Dict[str, Any]]:
    """Return the workbook's defined names (named ranges/expressions).

    These are formulas like `AT_Clear_Satin_w_Grit` that resolve to a
    cell or range. HyperFormula needs them registered explicitly,
    otherwise any formula that uses one returns #NAME?.

    Returns a list of {name, expression, scope} dicts where:
      - name: the defined name (e.g. "AT_Clear_Satin_w_Grit")
      - expression: the formula it resolves to (e.g. "='Stnd Alts'!$E$145")
      - scope: sheet name if scoped to a sheet, else None (workbook-wide)
    """
    wb = _load_template(data_only=False)
    out: list[Dict[str, Any]] = []
    try:
        defined_names = wb.defined_names
    except Exception:
        return out

    # openpyxl's defined_names is a dict-like collection
    try:
        items = list(defined_names.items())
    except AttributeError:
        # Older openpyxl: defined_names is an iterable of (name, DefinedName)
        items = [(dn.name, dn) for dn in defined_names.definedName]

    for name, dn in items:
        try:
            # `value` is the formula string, e.g. "Sheet1!$A$1" or "{2,3}"
            expression = getattr(dn, "value", None) or getattr(dn, "attr_text", None)
            if not expression:
                continue
            # Skip print areas, hidden ranges, etc. — only keep real names
            if name.startswith("_xlnm."):
                continue
            scope = None
            if hasattr(dn, "localSheetId") and dn.localSheetId is not None:
                try:
                    scope = wb.sheetnames[dn.localSheetId]
                except (IndexError, TypeError):
                    pass
            out.append({
                "name": name,
                "expression": "=" + expression if not expression.startswith("=") else expression,
                "scope": scope,
            })
        except Exception:
            continue
    return out


def list_sheet_names() -> list[str]:
    """List every sheet in the template (used by tab bar)."""
    wb = load_workbook(TEMPLATE_PATH, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _fill_hex(cell) -> str | None:
    """Return cell fill colour as '#RRGGBB' or None if no/default fill."""
    try:
        f = cell.fill
        if not f or f.patternType in (None, "none"):
            return None
        rgb = f.start_color.rgb if f.start_color else None
        if not rgb or not isinstance(rgb, str):
            return None
        # openpyxl returns 'AARRGGBB' — drop alpha
        if len(rgb) == 8:
            rgb = rgb[2:]
        if rgb == "000000" or rgb == "FFFFFF":
            # Treat plain black/white as "no fill" for visual clarity
            return None
        return f"#{rgb}"
    except Exception:
        return None


def _cell_borders(cell) -> dict | None:
    """Return per-side border info {side: {style, color}} or None if no borders.

    Excel border styles map roughly to CSS as:
        thin    → 1px solid
        medium  → 2px solid
        thick   → 3px solid
        double  → 3px double
        dashed  → 1px dashed
        dotted  → 1px dotted
        hair    → 1px dotted
        mediumDashed/dashDot/etc. → 2px dashed
    """
    try:
        b = cell.border
        if not b:
            return None
        out = {}
        for side in ("top", "right", "bottom", "left"):
            s = getattr(b, side, None)
            if not s or not s.style:
                continue
            color = None
            try:
                if s.color and s.color.type == "rgb" and isinstance(s.color.rgb, str):
                    rgb = s.color.rgb
                    if len(rgb) == 8:
                        rgb = rgb[2:]
                    if rgb and rgb != "000000":
                        color = f"#{rgb}"
            except Exception:
                pass
            out[side] = {"style": s.style, "color": color}
        return out or None
    except Exception:
        return None


def _font_color(cell) -> str | None:
    try:
        c = cell.font.color
        if not c:
            return None
        rgb = c.rgb if c.type == "rgb" else None
        if not rgb or not isinstance(rgb, str):
            return None
        if len(rgb) == 8:
            rgb = rgb[2:]
        if rgb in ("000000", "FFFFFF"):
            return None
        return f"#{rgb}"
    except Exception:
        return None


def read_totals(filled_xlsx_bytes: bytes) -> Dict[str, Dict[str, Any]]:
    """Read computed totals back out of a saved-and-reopened workbook.

    NOTE: openpyxl reads CACHED computed values (last time Excel saved).
    A workbook we just filled with openpyxl will have stale cached
    totals — the formulas weren't actually evaluated. Excel evaluates
    on next open.

    So this function is only useful AFTER a workbook has been opened in
    Excel and saved again, OR for previewing totals from a workbook
    Troy filled in Excel and uploaded.

    For Screen 2's live totals we use the `pricing` engine via /api/price
    (the Computed Bid panel), so this function is mostly here for
    completeness.
    """
    wb = load_workbook(io.BytesIO(filled_xlsx_bytes), data_only=True)
    out: Dict[str, Dict[str, Any]] = {"epoxy": {}, "polish": {}}
    for name, coord in EPOXY_TOTALS.items():
        out["epoxy"][name] = wb["Epoxy"][coord].value
    for name, coord in POLISH_TOTALS.items():
        out["polish"][name] = wb["Polish"][coord].value
    return out


# NOTE: the old `compute_estimate_totals` Python mirror was removed —
# it could never be accurate (it had no access to the system *selection*
# that determines price). The authoritative bid now comes from the
# `pricing` engine via /api/price (see backend/pricing.py), surfaced in
# the Estimate screen's Computed Bid panel and passed through generate.
