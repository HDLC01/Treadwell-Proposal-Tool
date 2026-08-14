"""The Analytics page as an .xlsx — one sheet per tab, plus Kyle's trailing-12 sheet.

WHY THE BROWSER SENDS THE NUMBERS. Every figure on that page comes from one place,
frontend/js/analytics-core.js, and the server has no idea what the viewer has filtered — the
dashboard fetches all the bids once and re-totals locally on every click. Re-deriving the totals
here would mean two engines computing the same win rates from the same rows, agreeing right up
until somebody fixes a definition in one of them. So the browser computes and this module only
formats. Its whole job is turning typed cells into a workbook.

WHY THE TRAILING SHEET IS DIFFERENT. Kyle keeps "Trailing 12TH MONTH.xlsx" — nineteen dated tabs,
each a manual BasisBoard pull. This sheet reproduces that layout cell for cell, including his
formulas, so the download drops into his archive as the next dated tab and still recalculates when
he edits a number. That is also why the derived cells are FORMULAS rather than values: a workbook
holding both our arithmetic and his would be two answers to one question.
"""
from __future__ import annotations

import datetime as dt
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Kyle's formats, read off his workbook ────────────────────────────────────
# Excel ACCOUNTING, not a plain currency format: the "$" is flush left, the digits align on the
# right, and a zero shows as "-". Copied verbatim so a pasted-in tab looks like the ones beside it.
FMT_ACC0 = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'
FMT_ACC2 = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
FMT_PCT = "0%"
FMT_DATE = "mm-dd-yy"

# Sheet-side formats for the four dashboard tabs, which are ours to design.
FMT_MONEY = '"$"#,##0'
FMT_MONEY2 = '"$"#,##0.00'
FMT_INT = "#,##0"

_HEAD_FILL = PatternFill("solid", fgColor="F3F1EA")
_TITLE_FONT = Font(bold=True, size=13)
_HEAD_FONT = Font(bold=True, size=10)

# Kyle's columns. The value column holds the number or formula; the label column holds the row
# names, repeated per trade exactly as his sheet does.
_T12_VALUE_COLS = ("C", "F", "I", "L")
_T12_LABEL_COLS = ("B", "E", "H", "K")

_T12_LOGIC = (
    "Logic: the 15-month window supplies the awards; the last 90 days of submissions come OUT of "
    "the denominator, because a bid submitted in the last three months has not been decided yet. "
    "What is left is a true trailing twelve months that ended 90 days ago."
)


def _cell_format(kind: str) -> str:
    return {"money": FMT_MONEY, "money2": FMT_MONEY2, "int": FMT_INT, "pct": FMT_PCT}.get(kind, "")


def _write_table(ws, at_row: int, table: Dict[str, Any]) -> int:
    """One titled table; returns the next free row."""
    r = at_row
    title = table.get("title")
    if title:
        ws.cell(row=r, column=1, value=title).font = _TITLE_FONT
        r += 1
    cols = table.get("columns") or []
    for i, col in enumerate(cols, start=1):
        c = ws.cell(row=r, column=i, value=col.get("label") or "")
        c.font = _HEAD_FONT
        c.fill = _HEAD_FILL
    r += 1
    for row in table.get("rows") or []:
        for i, cell in enumerate(row, start=1):
            if isinstance(cell, dict):
                # A typed numeric cell. `None` stays blank rather than becoming a misleading 0 —
                # the page shows "—" for the same value, and a zero would read as a real answer.
                ws.cell(row=r, column=i, value=cell.get("v"))
                fmt = _cell_format(str(cell.get("t") or ""))
                if fmt:
                    ws.cell(row=r, column=i).number_format = fmt
            else:
                ws.cell(row=r, column=i, value=cell)
        r += 1
    # Widths from the content, so nothing arrives as ####.
    for i, col in enumerate(cols, start=1):
        longest = len(str(col.get("label") or ""))
        for row in table.get("rows") or []:
            if i - 1 < len(row):
                v = row[i - 1]
                longest = max(longest, len(str(v.get("v")) if isinstance(v, dict) else str(v)))
        ws.column_dimensions[get_column_letter(i)].width = min(46, max(11, longest + 3))
    return r + 1


def _write_trailing12(ws, t12: Dict[str, Any], caption: str) -> None:
    """Kyle's layout, cell for cell.

    The one deliberate omission is his E1 note, "Fill in these cells" — an instruction to whoever
    was pasting numbers by hand, which is the job this replaces."""
    ws["A1"] = "BASIS BOARD DATA WIN % for last 12 months"
    ws["A1"].font = _TITLE_FONT

    as_of = str(t12.get("as_of") or "")
    try:
        ws["B2"] = dt.date.fromisoformat(as_of)
        ws["B2"].number_format = FMT_DATE
    except ValueError:
        ws["B2"] = as_of
    ws["A3"] = "Custom Date Range"
    ws["A3"].font = _HEAD_FONT

    # The two range descriptions live in column A, against the rows they govern — his wording.
    for row in (4, 5, 12, 13):
        ws.cell(row=row, column=1, value="15 months past until today")
    for row in (8, 16):
        ws.cell(row=row, column=1, value="3 months past until today")

    columns = t12.get("columns") or []
    for idx, col in enumerate(columns):
        if idx >= len(_T12_VALUE_COLS):
            break
        V, L = _T12_VALUE_COLS[idx], _T12_LABEL_COLS[idx]
        ws[L + "3"] = col.get("label") or ""
        ws[L + "3"].font = _HEAD_FONT

        # Labels carry their literal double quotes, exactly as he typed them.
        labels = {
            4: '"Won Amount"',
            5: '"Total Submitted Amount"',
            6: "Win % by volume",
            8: 'Last 90 days "Total Submitted Amount"',
            10: "Win % by Volume Excluding last 90 days",
            12: '"# Awarded Projects"',
            13: '"# Submitted Projects"',
            14: "Win % by number of projects",
            16: 'Last 90 days "# Submitted Projects"',
            18: "Win % by # Projects Excluding last 90 days",
            20: "Average Size of Bid",
            21: "Average Size of Win",
        }
        for row, text in labels.items():
            ws[L + str(row)] = text
        for row in (10, 18):                     # the two rows he actually reads
            ws[L + str(row)].font = Font(bold=True)

        # Raw sums — the only numbers that cross the wire.
        ws[V + "4"] = col.get("won_amount")
        ws[V + "4"].number_format = FMT_ACC0
        ws[V + "5"] = col.get("submitted_amount")
        ws[V + "5"].number_format = FMT_ACC0
        ws[V + "8"] = col.get("sub90_amount")
        ws[V + "8"].number_format = FMT_ACC2
        ws[V + "12"] = col.get("n_awarded")
        ws[V + "13"] = col.get("n_submitted")
        ws[V + "16"] = col.get("n_sub90")

        # Derived cells stay live formulas, so his edits recalculate and there is exactly one
        # answer in the file.
        ws[V + "6"] = "={0}4/{0}5".format(V)
        ws[V + "10"] = "={0}4/({0}5-{0}8)".format(V)
        ws[V + "14"] = "={0}12/{0}13".format(V)
        ws[V + "18"] = "={0}12/({0}13-{0}16)".format(V)
        ws[V + "20"] = "={0}5/{0}13".format(V)
        ws[V + "21"] = "={0}4/{0}12".format(V)
        for row in (6, 10, 14, 18):
            ws[V + str(row)].number_format = FMT_PCT
        for row in (20, 21):
            ws[V + str(row)].number_format = FMT_ACC2

    ws["A23"] = caption
    ws["A23"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A25"] = _T12_LOGIC
    ws["A25"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 30
    for L in _T12_LABEL_COLS:
        ws.column_dimensions[L].width = 42
    for V in _T12_VALUE_COLS:
        ws.column_dimensions[V].width = 16


def build_workbook(payload: Dict[str, Any]) -> bytes:
    """The whole page as a workbook. Pure: no app, no database, no clock."""
    wb = Workbook()
    wb.remove(wb.active)                     # the default sheet, replaced by the real ones

    filters = str(payload.get("filters") or "")
    generated = str(payload.get("generated_at") or "")
    header = "Filters: " + filters if filters else ""
    if generated:
        header = (header + " · " if header else "") + "Data as of " + generated
    if payload.get("truncated"):
        header += " · CAPPED: this org has more bids than the dashboard loads, and the ones " \
                  "dropped are not chosen by date — treat these as approximate."

    for tab in payload.get("tabs") or []:
        ws = wb.create_sheet(title=str(tab.get("name") or "Sheet")[:31])
        r = 1
        if header:
            ws.cell(row=r, column=1, value=header).alignment = Alignment(wrap_text=False)
            r += 2
        for table in tab.get("tables") or []:
            r = _write_table(ws, r, table)

    t12 = payload.get("trailing12")
    if t12:
        ws = wb.create_sheet(title="Trailing 12")
        caption = "All bids in the window; a project with two trades counts under each column, " \
                  "so the trade columns can add up to more than All Bids. Awards and submissions " \
                  "from {0} to {1}; the 90-day rows from {2}.".format(
                      t12.get("w15_from") or "?", t12.get("as_of") or "?",
                      t12.get("w90_from") or "?")
        _write_trailing12(ws, t12, caption)

    if not wb.sheetnames:                    # a payload with nothing in it still has to open
        wb.create_sheet(title="Analytics")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
