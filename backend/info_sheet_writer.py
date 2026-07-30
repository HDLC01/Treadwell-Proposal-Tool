"""Project Info Sheet — the ops hand-off workbook.

When a bid turns into a job, someone re-types everything the estimate already
knows into Kyle's `$Project Info Sheet`: who sold it, where it is, who to call,
the contract amount, the tax flags. Accounting then imports a row of it into
Foundation and bills off the Invoice tab. This module fills the parts we can
already answer and leaves the judgement calls to a human.

Two colours drive the whole design, taken from a completed sheet (FBC Oak
Grove) that Hanz marked up:

  * **chartreuse** — data the tool already holds. Prefilled here.
  * **pink** — a decision (market segment, payment terms, retainage). Left
    empty on purpose, but the dropdown has to work, which is the awkward part
    (see `prepare_info_sheet_template.py`).

Everything the estimator changes in the browser is stored per draft as
`data.info_cell_values`, keyed `"Info Sheet!B14"`, and layered over the prefill
at download time — an edit always wins, and it survives a re-download.

Only the Info Sheet is written. `Foundation Import`, `Invoice` and `Deposit`
populate themselves from it through their own cross-sheet formulas, so touching
them would just break the links.

The grid reader and value coercion are imported from `estimate_writer` rather
than copied: those helpers are pure (they take a cell, not a workbook), and the
`_coerce` path in particular carries the formula-injection defence that any cell
fed from an API needs.
"""
from __future__ import annotations

import io
import logging
import re
from pathlib import Path
from typing import Any, Dict, Optional

import openpyxl
from openpyxl.styles import PatternFill

import estimate_writer as ew
import leads
from estimate_writer import _coerce

log = logging.getLogger("proposal_tool.info_sheet")

TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "project_info_sheet.xlsx"
SHEET = "Info Sheet"

# `Lists` holds the dropdown source columns. It is hidden in the workbook and
# stays hidden here — it is never served, never written, and never the target
# of a structural op. Deleting rows from it would empty MarketList and friends,
# and every picker would silently go free-text again.
HIDDEN_SHEETS = frozenset({"Lists"})

# Excel caps an op list long before this; the bound just stops a malformed
# payload from replaying forever.
MAX_STRUCT_OPS = 500


# ─── Where each answer comes from ──────────────────────────────────────
# The estimate sheet's tax/wage flags live in the project-info block, whose
# rows differ per layout: gyp sits two rows lower than epoxy and keeps its own
# Taxable cell. Polish mirrors epoxy by formula, so it reads from epoxy.
_FLAG_CELLS = {
    #                    epoxy   gyp
    "prevailing_wage":  ("D5",  "D7"),
    "taxable":          ("B6",  "B8"),
    "remodel_tax":      ("D6",  "D8"),
}
_GYP_BASE = 'Gyp (USG 1-8")'

_STATES = {
    "KS": "KS - Kansas", "MO": "MO - Missouri", "OK": "OK - Oklahoma",
    "NE": "NE - Nebraska", "IA": "IA - Iowa", "AR": "AR - Arkansas",
    "IL": "IL - Illinois",
}

# `data.source` (the intake "How did they find us?" picker) → the Lead Source
# dropdown. Deliberately partial: intake's "referral" does not say whether it
# came from a partner, a past customer or a supplier, and a bid invite off
# BasisBoard is not any of the listed sources. Anything unmapped clears the
# cell rather than keeping the template's "Repeat Customer" default — a guessed
# lead source gets reported out of Foundation as if it were fact.
_LEAD_SOURCES = {
    "google_lead": "Online",
    "referral": "Referral - Other",
    "other": "Other",
    "treadwell_vehicle": "Other",
}

# Primary Floor, matched against the system names the estimator picked on the
# sheet. Order matters: a hybrid flake system ("Flake with hybrid blend") is a
# urethane-cement floor, so the urethane test has to run before the flake one.
_FLOORS = [
    (("urethane", "poly-crete", "polycrete", "hybri"), "Epoxy - Urethane Cement"),
    (("quartz",), "Epoxy - Quartz"),
    (("flake",), "Epoxy - Flake"),
]
_POLISH_FLOORS = [
    (("large aggregate", "aggregate"), "Polish - Large Aggregate"),
    (("s&p", "salt", "pepper"), "Polish - S&P"),
    (("cream",), "Polish - Cream"),
]


# ─── Template access ───────────────────────────────────────────────────
# The grid reader lives in estimate_writer and is shared, not forked. It was
# forked once and the copy immediately drifted: it lacked the border-symmetry
# pass, which on SOV alone decides 189 of 323 cells. SOV is a bordered table,
# so without it the tab renders visibly wrong.
def template_version() -> str:
    """ETag seed — changes whenever the committed template is replaced."""
    return str(TEMPLATE_PATH.stat().st_mtime_ns)


def visible_sheets() -> list[str]:
    """The tabs a user may see and edit, in workbook order.

    Derived from the file rather than hardcoded so the read gate and the write
    gate can never disagree about what was on screen. A test pins the exact
    five, so a master that unhides `Lists` or adds a tab fails CI instead of
    quietly exposing the dropdown source.
    """
    wb = ew._load_template(data_only=False, path=TEMPLATE_PATH)
    return [ws.title for ws in wb.worksheets
            if ws.sheet_state == "visible" and ws.title not in HIDDEN_SHEETS]


def read_sheet(sheet_name: str) -> Dict[str, Any]:
    """One tab as the JSON the grid renders. KeyError if it isn't a real tab.

    `parse_x14=False`: this workbook has no extension validations at all —
    `prepare_info_sheet_template.py` converted every one of them to a plain
    validation over a defined name — and the parser guesses a sheet's zip
    member by position, so there is nothing to gain by running it.
    """
    if sheet_name not in visible_sheets():
        raise KeyError(sheet_name)
    return ew.read_sheet_grid(sheet_name, path=TEMPLATE_PATH, parse_x14=False)


def read_workbook(prefill: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    """Every visible tab, with the prefill already merged into the cells.

    Merging server-side means the grid shows exactly what the download will
    write — there is no second source of truth for the client to reconcile, and
    a cell's `role` travels with it so a row insert shifts the colour key for
    free.
    """
    names = visible_sheets()
    sheets = {n: read_sheet(n) for n in names}
    _overlay_prefill(sheets.get(SHEET), prefill or {})
    return {
        "order": names,
        "sheets": sheets,
        "text_cells": sorted(f"{SHEET}!{a}" for a in TEXT_CELLS),
        "template_version": template_version(),
    }


def _overlay_prefill(grid: Optional[Dict[str, Any]], prefill: Dict[str, Any]) -> None:
    """Stamp the prefill onto the Info Sheet grid, tagging provenance.

    `role` is the colour key the estimator reads: chartreuse for an answer the
    tool supplied, pink for a decision it deliberately left alone. Attaching it
    to the cell rather than shipping a parallel address list means the client's
    insert/delete transform carries it along with everything else.
    """
    if not grid:
        return
    by_addr = {c["addr"]: c for c in grid["cells"]}

    def slot(addr: str) -> Dict[str, Any]:
        cell = by_addr.get(addr)
        if cell is None:
            m = re.match(r"^([A-Z]+)([0-9]+)$", addr)
            col = 0
            for ch in m.group(1):
                col = col * 26 + (ord(ch) - 64)
            cell = {"addr": addr, "row": int(m.group(2)), "col": col}
            grid["cells"].append(cell)
            by_addr[addr] = cell
        return cell

    for addr, value in (prefill or {}).items():
        cell = slot(addr)
        cell["role"] = "prefill"
        if value == "" or value is None:
            cell.pop("value", None)          # deliberately cleared a default
        else:
            cell["value"] = value
    for addr in PINK_CELLS:
        slot(addr)["role"] = "decision"



# ─── Prefill ───────────────────────────────────────────────────────────
def _txt(v) -> str:
    return "" if v is None else str(v).strip()


def _num(v) -> Optional[float]:
    try:
        n = float(str(v).replace(",", "").replace("$", "").strip())
    except (TypeError, ValueError):
        return None
    return n


def _yn(v) -> str:
    return "Y" if str(v).strip().lower() in ("yes", "y", "true", "1") else "N"


def _person(email: str) -> str:
    """"kyle@wetreadwell.com" → "Kyle". The same shorthand the CRM board and
    the Projects list show; real full names live in `profiles`, which this
    service does not read.

    A draft the lead autopilot created has no human owner yet — printing
    "Autopilot" as the Estimator / Sales Rep on an accounting document would
    read as a name. Blank, so whoever picks the job up fills it in.
    """
    local = (email or "").split("@")[0]
    if local.strip().lower() == leads.AUTOPILOT_ACTOR:
        return ""
    return " ".join(p.capitalize() for p in re.split(r"[._-]+", local) if p)


def _base_role(data: Dict[str, Any]) -> str:
    """Which layout the bid was priced on.

    The estimator's nominated base tab wins. Failing that, `work_type` decides —
    NOT "the first base-kind tab". On real drafts every priced tab carries
    `kind: "base"` (all seven of them, including the five gyp variants nobody
    priced), so picking the first would answer "epoxy" for every job.
    """
    tabs = [t for t in (data.get("priced_tabs") or []) if isinstance(t, dict)]
    base_id = data.get("base_tab_id")
    if base_id:
        for tab in tabs:
            if tab.get("id") == base_id:
                role = _txt(tab.get("role")).lower()
                if role:
                    return role

    work_type = _txt(data.get("work_type")).lower()
    if work_type in ("epoxy", "polish", "gyp"):
        return work_type
    if work_type == "combo":
        return "epoxy"        # block one of a combo is the resin system
    for tab in tabs:          # unknown work type: fall back to the first priced tab
        role = _txt(tab.get("role")).lower()
        if role:
            return role
    return "epoxy"


def _base_tab(data: Dict[str, Any], role: str) -> Optional[Dict[str, Any]]:
    """The priced tab the bid actually came off, so its own product names can be
    read without the other tabs' names mixed in."""
    tabs = [t for t in (data.get("priced_tabs") or []) if isinstance(t, dict)]
    base_id = data.get("base_tab_id")
    if base_id:
        for tab in tabs:
            if tab.get("id") == base_id:
                return tab
    for tab in tabs:
        if _txt(tab.get("role")).lower() == role:
            return tab
    return None


def _system_text(data: Dict[str, Any], role: str) -> str:
    """What the BASE bid installs, lowercased, for the Primary Floor match.

    Scoped to the base tab on purpose. Pooling every priced tab's `sys_names`
    let an alternate the customer never bought decide the answer: an epoxy job
    carrying a quartz option matched "quartz" before its own flake system, and
    B17 drives B18 Division, which is how the job is filed in Foundation.
    """
    parts = [_txt(data.get("system_name"))]
    tab = _base_tab(data, role)
    if tab:
        parts.append(_txt(tab.get("system_desc")))
        parts.extend(_txt(n) for n in (tab.get("sys_names") or []))
    elif not data.get("priced_tabs"):
        # No sheet snapshot at all (an older draft) — the proposal's system list
        # is the only description there is.
        for sysrow in data.get("sheet_systems") or []:
            if isinstance(sysrow, dict):
                parts.append(_txt(sysrow.get("name")))
    return " ".join(p for p in parts if p).lower()


def _primary_floor(data: Dict[str, Any]) -> str:
    role = _base_role(data)
    if role == "gyp":
        return "Gypsum Cement Underlayment"
    text = _system_text(data, role)
    table = _POLISH_FLOORS if role == "polish" else _FLOORS
    for keywords, label in table:
        if any(k in text for k in keywords):
            return label
    return ""


# Which takeoff fields make up a floor area, per layout. The sheet-resolved
# names come first; the second tuple is what intake called them, for drafts
# saved before the sheet snapshot existed.
_SF_KEYS = {
    "epoxy":  (("epoxy_sf", "epoxy_sf_2"), ("system_1_sf", "system_2_sf")),
    "polish": (("polish_sf",), ("polish_sf",)),
    "gyp":    (("gyp_soft_sf", "gyp_hard_sf", "gyp_corridor_sf"),) * 2,
}
_LF_KEYS = (("cove_lf", "cove_lf_2"), ("cove_1_lf", "cove_2_lf"))


def _sum_area(source: Dict[str, Any], keys) -> Optional[float]:
    vals = [v for v in (_num(source.get(k)) for k in keys) if v]
    return sum(vals) if vals else None


# Cove base is a resin product. Polish and gypsum underlayment do not carry it,
# and neither layout has cove cells on its estimate tab — so any cove figure on
# such a job is a leftover intake number from a different scope. Reporting it
# would have ops order material for work nobody bid.
_COVE_ROLES = {"epoxy", "combo"}


def _areas(data: Dict[str, Any], role: str) -> tuple[Optional[float], Optional[float]]:
    """(floor SF, cove LF) for one layout. Sheet-resolved areas win, because
    they follow the estimator's takeoff edits on the grid."""
    area = data.get("sheet_area") if isinstance(data.get("sheet_area"), dict) else {}
    sheet_keys, intake_keys = _SF_KEYS.get(role, _SF_KEYS["epoxy"])
    sf = _sum_area(area, sheet_keys) or _sum_area(data, intake_keys)
    if role not in _COVE_ROLES:
        return sf, None
    lf = _sum_area(area, _LF_KEYS[0]) or _sum_area(data, _LF_KEYS[1])
    return sf, lf


# Fallback wording for the second system block when a tab carries real area but
# no derived name — the sheet's own vocabulary for each layout.
_ROLE_LABELS = {"epoxy": "Epoxy", "polish": "Polished Concrete",
                "gyp": "Gypsum Underlayment"}


def _second_system(data: Dict[str, Any], base_role: str):
    """The other half of a combo bid, for the sheet's second system block.

    An epoxy+polish job prices both a resin and a polish tab. Only one can be
    block one, so without this the polish half never reaches the hand-off at all
    — the exact re-typing this page exists to remove.

    Two things disqualify a tab:

    * **No floor area.** Every draft carries all five gyp variants as priced tabs
      at zero square feet; picking one by tab order would put 'N12 1/8"' on the
      sheet as if somebody had bid it.
    * **Marked as a proposal OPTION.** An option is an alternate the customer was
      quoted and did not buy; B57 covers the base scope only. Listing it here
      would have ops order material and book crews for scope that was never sold.
    """
    opts = data.get("tab_opts") if isinstance(data.get("tab_opts"), dict) else {}
    for tab in data.get("priced_tabs") or []:
        if not isinstance(tab, dict):
            continue
        role = _txt(tab.get("role")).lower()
        if not role or role == base_role:
            continue
        opt = opts.get(tab.get("id"))
        if isinstance(opt, dict) and opt.get("is_option"):
            continue
        sf_src = tab.get("sf") if isinstance(tab.get("sf"), dict) else {}
        sheet_keys, _ = _SF_KEYS.get(role, _SF_KEYS["epoxy"])
        sf = _sum_area(sf_src, sheet_keys)
        if not sf:
            continue
        lf = _sum_area(sf_src, _LF_KEYS[0]) if role in _COVE_ROLES else None
        return (_txt(tab.get("system_desc")) or _ROLE_LABELS.get(role, ""), sf, lf)
    return None


def _flag(data: Dict[str, Any], flag: str) -> Optional[str]:
    """Read a Yes/No estimate flag out of the saved grid edits.

    On a gyp job two of the three flags may be read off the epoxy tab and one may
    NOT, and the difference is in the estimate template:

        Gyp (USG 1-8")!D7  =  '=Epoxy!D5'   Prevailing Wage — a mirror
        Gyp (USG 1-8")!D8  =  '=Epoxy!D6'   Remodel Tax     — a mirror
        Gyp (USG 1-8")!B8  =  'Yes'         Taxable         — its OWN literal

    So falling through to epoxy is right for the two mirrors and wrong for
    Taxable: the gyp bid's sales tax comes from `=IF($B$8="no",0,0.09475)` on the
    gyp tab, and nothing on the gyp tab reads Epoxy!B6. AI Autofill writes all
    seven flags to hardcoded `Epoxy!…` keys whatever the work type, so a gyp job
    for a school could hold Epoxy!B6="No" while the gyp tab still said "Yes" and
    the bid was priced WITH tax. Falling through then printed "Tax Exempt? Y" on
    a taxable job, fired the request-a-certificate instruction, and told
    Foundation the job was exempt.

    An untouched gyp Taxable cell means the template default stands (taxable), so
    returning None here leaves B66 alone rather than inventing an exemption.
    """
    cells = data.get("cell_values") if isinstance(data.get("cell_values"), dict) else {}
    epoxy_addr, gyp_addr = _FLAG_CELLS[flag]
    if _base_role(data) == "gyp":
        v = cells.get(f"{_GYP_BASE}!{gyp_addr}")
        if v not in (None, ""):
            return str(v)
        if flag == "taxable":
            return None          # gyp owns this cell; epoxy's answer is not it
    v = cells.get(f"Epoxy!{epoxy_addr}")
    return None if v in (None, "") else str(v)


def build_prefill(draft: Dict[str, Any], *, deposit_requested: bool = False) -> Dict[str, Any]:
    """Info Sheet cells we can answer from the draft, keyed by address.

    An empty string means "clear the template's default" — used where the
    template ships a guess that would read as fact once printed (Lead Source).
    A key that is simply absent leaves the template alone.

    Nothing pink is written. Manufacturer (B39) and Color/Blend (B41) are
    chartreuse on the marked-up sheet but stay blank too: the estimate does not
    record the supplier, and the blend is chosen with the customer after award.
    """
    data = draft.get("data") if isinstance(draft.get("data"), dict) else {}
    out: Dict[str, Any] = {}

    def put(addr: str, value) -> None:
        if value is None or value == "":
            return
        out[addr] = value

    # Who sold it. `estimator_name` is typed on the proposal screen; before that
    # the draft only knows the account that created it.
    put("B13", _txt(data.get("estimator_name")) or _person(draft.get("owner_email") or ""))
    put("B14", _txt(data.get("job_number")))
    put("B15", _txt(data.get("project_name")))
    put("B17", _primary_floor(data))

    state = _txt(data.get("state")).upper()
    put("B19", _STATES.get(state, state))
    put("B20", _txt(data.get("address")))
    put("B21", _txt(data.get("city")))
    put("D21", _txt(data.get("zip")))

    # Who gets billed. On a GC job that is the general contractor; direct work
    # is billed to the owner, whose name is the project name.
    is_gc = _txt(data.get("audience")).upper() == "GC" and bool(_txt(data.get("architect")))
    put("B23", _txt(data.get("architect")) if is_gc else _txt(data.get("project_name")))

    # The address block beneath B23 is the BILL-TO — the Invoice and Deposit tabs
    # print it. For direct work the customer is the owner of the site, so the job
    # address is right. On a GC job it is the contractor's office, which the draft
    # does not hold: filling the job site there would address an invoice to the GC
    # at the building they are constructing, and the prefill tick would tell the
    # estimator it had been checked. Left blank for a human instead.
    if not is_gc:
        put("B24", _txt(data.get("address")))
        put("B25", ", ".join(p for p in (
            _txt(data.get("city")),
            " ".join(p for p in (state, _txt(data.get("zip"))) if p)) if p))
        put("B26", _txt(data.get("contact_phone")))

    put("B29", _txt(data.get("contact_name")))
    put("B30", _txt(data.get("contact_phone")))
    put("B31", _txt(data.get("contact_email")))

    role = _base_role(data)
    put("B40", _txt(data.get("system_name")))
    sf, lf = _areas(data, role)
    put("B42", sf)
    put("D42", lf)

    second = _second_system(data, role)
    if second:
        put("B46", second[0])
        put("B48", second[1])
        put("D48", second[2])

    put("B57", _num(data.get("proposal_lump_sum")))
    costs = data.get("cost_snapshot") if isinstance(data.get("cost_snapshot"), dict) else {}
    put("B58", _num(costs.get("costs")))
    put("I58", _num(costs.get("man_hours")))

    out["B59"] = "Y" if deposit_requested else "N"

    # Blank beats a plausible guess here — see _LEAD_SOURCES.
    out["B62"] = _LEAD_SOURCES.get(_txt(data.get("source")).lower(), "")

    pw = _flag(data, "prevailing_wage")
    if pw is not None:
        out["B63"] = _yn(pw)
    taxable = _flag(data, "taxable")
    if taxable is not None:
        out["B66"] = "N" if _yn(taxable) == "Y" else "Y"   # exempt is taxable inverted
    remodel = _flag(data, "remodel_tax")
    if remodel is not None:
        out["B67"] = _yn(remodel)

    return out


# Cells the estimator may type into: everything the prefill can write, plus the
# pink decisions and the free-text blocks the sheet leaves for a human.
# Cells the sheet deliberately leaves to a human — the pink half of the colour
# key on Hanz's marked-up FBC Oak Grove sheet. Pink is a property of the CELL,
# not of whether we happened to write it: B16 is pink there *and* holds
# "Religious", because a person chose that. An estimator's entry never repaints
# one. `test_nothing_pink_is_prefilled` asserts the same set from the other
# side, so the test and the colour key are two views of one constant.
PINK_CELLS = frozenset({
    "B16",   # Market / Project Class
    "B39",   # Manufacturer
    "B41",   # Colour / Blend
    "B43",   # Special transition detail
    "B60",   # Payment terms
    "B61",   # New customer over $10k
    "B68",   # CCIP
    "B70",   # Retainage
    "F33",   # Bill platform / pay app
})

# Cells that must stay text even when they look numeric. A job number is the
# clearest case: "26.100" cast to a float is 26.1, and the Invoice and
# Foundation Import tabs both print it straight from B14. Phones lose their
# shape the same way. Info Sheet only — SOV has its own text columns, which the
# grid detects from the `@` number format instead.
TEXT_CELLS = frozenset({"B14", "B26", "B30", "F35", "B27"})

CHARTREUSE = "FFB3FF00"
PINK = "FFFFB0FF"
_FILL_PREFILL = PatternFill(fill_type="solid", start_color=CHARTREUSE, end_color=CHARTREUSE)
_FILL_DECISION = PatternFill(fill_type="solid", start_color=PINK, end_color=PINK)


# The characters Excel reads as the start of a formula or a DDE call. Same set
# `_coerce` guards; kept here because a text cell must skip _coerce's numeric
# casting and so cannot borrow its guard wholesale.
_INJECTION_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _as_text(v) -> str:
    """Keep the string, keep the injection guard `_coerce` would have applied."""
    s = "" if v is None else str(v)
    return "'" + s if s[:1] in _INJECTION_TRIGGERS else s


def _norm_ops(tab_structs, visible) -> list[dict]:
    """Validated structural ops, restricted to tabs the user can actually see.

    The `visible` filter is a security gate, not tidiness. `_apply_tab_structs`
    skips only sheets absent from the workbook, and `Lists` is present — so a
    crafted `{"sheet": "Lists", "kind": "delete_rows", "at": 4, "count": 20}`
    would wipe the dropdown source ranges. MarketList would resolve to blanks
    and every picker would go free-text: exactly the failure
    `prepare_info_sheet_template.py` exists to prevent, arriving through the
    front door.
    """
    ops = ew._norm_structs(tab_structs)[:MAX_STRUCT_OPS]
    return [op for op in ops if op["sheet"] in visible]


def resolve_addr(addr: str, tab_structs=None, sheet: str = SHEET) -> Optional[str]:
    """Where a template address ended up after the user's row/column edits.

    `None` if they deleted it. Callers outside this module need this because a
    hardcoded address stops meaning what it used to the moment a row goes in
    above it — the job-number mirror is the live example.
    """
    ops = [op for op in _norm_ops(tab_structs, {sheet}) if op["sheet"] == sheet]
    return ew._translate_addr(addr, ops)


def _apply_colour_key(ws, prefilled, ops) -> None:
    """Paint the provenance key Hanz colours by hand today.

    Chartreuse is *the set of addresses the prefill answered*, not "cells that
    ended up non-empty" — on the FBC sheet D42 is chartreuse and blank, because
    we knew the field and the job simply had no cove. Deriving it from emptiness
    would drop that and would also tick template defaults nobody chose.

    Runs after the structural replay, so every address has to be translated;
    pink goes on last so it wins any overlap (a test asserts there is none).
    Only ~35 addresses are ever stamped and none of them is a label, so the
    template's own header fills survive untouched. Assigning `.fill` replaces
    only that facet, so B57 keeps its currency format and B14 its `@`.
    """
    for addrs, fill in ((prefilled, _FILL_PREFILL), (PINK_CELLS, _FILL_DECISION)):
        for addr in addrs:
            moved = ew._translate_addr(addr, ops)
            if moved:
                ws[moved].fill = fill


# ─── Fill ──────────────────────────────────────────────────────────────
def fill_info_sheet(prefill: Dict[str, Any],
                    overrides: Optional[Dict[str, Any]] = None,
                    *, tab_structs=None) -> bytes:
    """Render the workbook.

    The ordering is the same invariant `fill_estimate` documents, and it is
    load-bearing: everything written in TEMPLATE coordinates goes down BEFORE
    the structural replay, and everything arriving in CURRENT coordinates after
    it. The prefill was authored against the pristine template, so it rides the
    shift like any other cell value; the estimator's overrides were typed
    against the grid they were looking at, so they are already current.

    Every cell is writable. The old EDITABLE whitelist is gone — a formula cell
    that gets typed over is replaced, exactly as in Excel.
    """
    wb = openpyxl.load_workbook(TEMPLATE_PATH)   # fresh — never the cached copy
    visible = set(visible_sheets())
    ops = _norm_ops(tab_structs, visible)
    ws = wb[SHEET]

    # Pass A — template coordinates.
    for addr, val in (prefill or {}).items():
        _write(ws, addr, val, text_cells=TEXT_CELLS)

    # Replay the user's row/column edits.
    ew._apply_tab_structs(wb, ops)
    info_ops = [op for op in ops if op["sheet"] == SHEET]

    # Pass B — current coordinates, whatever the estimator typed.
    for key, val in (overrides or {}).items():
        if "!" not in key:
            continue
        sheet_name, addr = key.split("!", 1)
        if sheet_name not in visible or not ew._CELL_SHAPE_RE.fullmatch(addr):
            log.info("info sheet: skipped override %r", key)
            continue
        _write(wb[sheet_name], addr, val,
               text_cells=TEXT_CELLS if sheet_name == SHEET else frozenset())

    # Pass C — the colour key, over the shifted addresses.
    _apply_colour_key(ws, list((prefill or {}).keys()), info_ops)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write(ws, addr: str, val, *, text_cells) -> None:
    """One cell, with the text rule and the formula-injection guard."""
    try:
        if val == "" or val is None:
            ws[addr] = None
        elif addr in text_cells:
            ws[addr] = _as_text(val)
        else:
            ws[addr] = _coerce(val)
    except Exception as exc:  # noqa: BLE001 — one bad address must not lose the file
        log.warning("info sheet: could not write %s!%s: %s", ws.title, addr, exc)
